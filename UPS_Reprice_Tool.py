# =====================================================================
# UPS Reprice Tool -- BUILD 2026-07-29-fix57
# (若右上角/標題看不到，可用此註解確認你跑的是最新版)
#
# 本版累積修正:
#   1. AHS-Weight 用真實過磅實重(非帳單膨脹值); pure ADJ+type9 抑制
#   2. AHS 優先序改 UPS 官方順序 (Weight>Dimension>Packaging)
#   3. AHS 類型: 有真實尺寸用門檻, 無尺寸靠發票code; Packaging只靠code
#   4. AHS 退費跟隨發票符號 (負值)
#   5. DAS pure-ADJ 改判用費率差額(帶符號)
#   6. pure ADJ 無AHS code 不無中生有 (含 C-A 分支)
#   7. LPS: 排除 1x1x1 佔位假尺寸, pure ADJ 大尺寸正確觸發
#   8. base rate: pure ADJ + HR佔位 + ent&bil都有值 -> 差額
#   9. base差額修正: 用原始billed/entered(避開min-billable污染);
#      hi腿用AG尺寸+floor, lo腿用HR原始尺寸+無floor (修type9差額塌陷)
#  10. LPS weight門檻用真實過磅實重(非dim灌高帳單重); 修123>110誤觸發LPS
#  11. pure_ca gate 加入LPS code(LPR), 修帶LPR的pure ADJ漏算LPS
#  12. 合併時: ADJ若無尺寸/重量(純附加費調整,如Reschedule Delivery),
#      不可用其0值覆蓋SHP的base rate/AHS/LPS; zone亦退回用SHP的有效zone
#  13. 修正12的偵測方式: 改用「ADJ有無FRT運費行」判定(不可看Length/
#      Weight欄, 那些會被SHP fallback填滿而誤判為有重算)
#  14. AHS類型優先序重整: 發票「正值」AH code 最優先(UPS實際過磅的
#      結果勝過我方門檻, 如申報1磅但UPS重秤後收AHW); 只有「負值退費」
#      或完全無code時才用門檻重算修正後包裹
#  15. LPS重量門檻: A包裹(HR)也套用真實過磅重規則; type-9 pure ADJ
#      無可信實重 -> 重量路徑不觸發, 由尺寸決定。修「尺寸差1吋造成
#      LPS升級」被A誤觸發而C-A淨額歸零的問題
#  16. AHS-Weight 不需尺寸即可判定(只看實重): 修 Residential->Commercial
#      改判(reclass)的ADJ無dims無code時, AHS-Weight被算成0而抹掉SHP
#      原本合法的重量附加費
#  17. AHS code 改歸類為 Dimension(原誤歸Weight): AHS = "Addl. Handling
#      second long. Side" 第二長邊>30吋, 是尺寸條件而非重量。已加入
#      registry 並修正 fallback 分類
#  28. normalize_service 猜測不再偽裝成判斷結果。描述比對不到任何服務
#      時, 它回傳 self.base_rate_service.get() -- 也就是 Import Files
#      分頁「Currently Editing」下拉的當前值。所以 Ground Hundredweight
#      會被顯示成 Ground Residential (或你當下選的任何服務), 看起來像
#      系統判斷, 其實只是撿了一個不相關的 UI 狀態。
#        a. 新增 service_is_recognised(), 比對不到時設 service_was_guessed
#        b. 可計價的行 -> Notes 標明「服務未識別, 沿用下拉預設 X, 請確認」
#        c. UNSUPPORTED 的行 -> Shipment Type 直接顯示發票原始描述,
#           不再顯示假的服務別
#      實測 586Y51: 123 / 16,926 條 FRT 會被標記。
#  27. 最低計費重量誤判為 artifact 的修正。原本只要「帳單重 == 設定的
#      oversize_min(90) 或 ahs_min(40)」且大於實重, 就一律當成 artifact
#      降回實重。但 UPS 本來就會照這些最低值實收 -- 實測 586Y51 有 188
#      條 FRT ($4,938.24) 命中, 其中 174 條 Billed Weight Type = 3
#      (UPS 自己標示「來自附加費最低重量」), 且全部帶 LPS/AHS 族代碼。
#      這些包裹在檔案中沒有尺寸, large_package_trigger 無法重新推導,
#      最低重量就此遺失 -> 例如 10 lb 實重的包裹 UPS 按 90 lb 收費,
#      工具卻按 10 lb 重算, 產生大額假超收。
#      改判依據 (任一成立即視為真實):
#        a. Billed Weight Type == 3
#        b. 帳單重 == oversize_min 且該筆帶 LPS 族代碼
#        c. 帳單重 == ahs_min     且該筆帶 AHS 族代碼
#      並以 minimum_billable_floor 帶到 billable_weight, 避免下游再遺失。
#  26. Ground/Air Hundredweight (CWT) 標記為不支援, 不再靜默錯算。
#      normalize_service 沒有 hundredweight 比對規則, 會落到最後的
#      return self.base_rate_service.get() -- 服務別由 Import Files 分頁
#      下拉的「當前選擇」決定, 等於用不相關的 UI 狀態決定計價基礎。
#      HWT 是整票貨總重按每百磅計價, 不是逐件包裹計價, 照 parcel 費率
#      重算會產生假的超收。改為沿用 18 的 UNSUPPORTED 標記路徑。
#      (實測 586Y51: 1 個 lead / 11 件 / $116.55)
#  25. AR 分類擴充: 過去只有 AR=="ACC" 會被計價與做未知代碼偵測, 所以
#      進口報關 (BRK) 與關稅 (GOV) 完全靜默 -- 不計價也不通報。改為
#      BILLABLE_AR_CODES = ACC/BRK/GOV (MSC 為帳戶層, 不計算), 自訂附加費與未知偵測皆適用;
#      FRT/FSC 刻意不提供 (基本運費與燃油由引擎自行計算, 註冊會重複計
#      算), INF 恆為 $0 不處理。Custom Surcharges 視窗加上「各欄位會出
#      現在系統哪裡」說明。
#      另: account_rows (無追蹤號的帳戶層費用) 同樣是收集後從未讀取,
#      現在輸出到 "Account Charges" 分頁 (實測 586Y51: 9 列 $7,589.88)。
#  24. TRACKING_COL 由 13 (Lead Shipment Number) 改為 20 (Tracking Number,
#      Excel U 欄)。多件貨 (MPS) 每個包裹共用同一個 lead, 舊分組會把
#      重量/尺寸/費率各不相同的實體包裹合併成一列重算, 其餘直接從報表
#      消失 (實測 586Y51: 7 個 lead 掛 25 個包裹, 18 件被吃掉)。col13
#      也不保證是追蹤號 -- 取件調整放取件記錄號、進口報關放報關批號,
#      會讓 ADJ 變成對不到 SHP 的孤兒層。切換前已確認 col13 有值而
#      col20 空的列數為 0。
#  23. Type 改名並真正生效 (原本三個選項行為完全相同, 只存不用):
#        REPRICE      - 用費率表重算, 計入 CAL Total
#        COPY INVOICE - 照抄帳單金額, 計入 CAL Total
#        REPORT ONLY  - 照抄帳單金額, 只顯示, 不計入 Total/燃油
#      舊值 TEMPLATE/RAW/AUDIT_ONLY 透過 DYNAMIC_TYPE_ALIASES 自動轉換,
#      舊 config 與舊匯出檔照常可用。另移除費率區的 ★ 標記與
#      Custom Surcharges 的 Load Setup 按鈕 (存檔即生效, 不需手動載入)。
#  22. 未知代碼現在會主動告知: unknown_rows 過去只被收集、從未被讀取,
#      報表也只寫 Shipment Detail 一張, 所以帳單有無法計價的代碼時
#      完全不會有任何提示, CAL Total 被低估卻看不出來。現在
#      (a) 報表多一張 "Unknown Codes" 明細分頁
#      (b) 跑完跳警告, 依金額絕對值排序列出 代碼/筆數/金額
#      (c) 狀態列顯示未知代碼數
#      (d) Code Lookup 掃描把未知代碼排到最上面並標 ⚠
#  21. Add Custom Surcharge 視窗: AR 由自由輸入改為下拉 (ACC/FRT/FSC/
#      INF), 只有 ACC 會重新計價; 新增 AS Code 空白、ACC 無名稱、
#      代碼重複三項擋下 (原本打錯 AR 會靜默存檔然後永遠不計價)。
#      (Type 的處理見 23)
#  20. Custom Surcharges UI 根本沒被掛上去: tab_dynamic_surcharge 有建
#      Frame 但 notebook.add() 沒加、_build_dynamic_surcharge_tab() 從未
#      被呼叫, 所以畫面上沒有任何入口可新增自訂 ACC。改為獨立視窗
#      (同燃油排程做法), 由 Pricing Rules 的「➕ Custom Surcharges...」
#      開啟; save/load 加上視窗已關閉的防護。
#  19. 自訂 ACC 真正計價: dynamic_surcharge_seen_codes 過去只被收集、
#      從未被讀取, 所以 Custom Surcharges 分頁新增的代碼一律算 $0。
#      現在依 shipment_type/zone 查費率寫入 row_result, 並自動加入
#      輸出欄位/CAL Total/報表版面; Fuel Eligible=NO 者計入 Total 但
#      不計入燃油基數。Rules 分頁費率由唯讀 Label 改為可輸入 Entry
#      並新增 Save Rates。
#  18. 不支援的服務/zone(如 Standard to Canada, zone 53)不再硬套
#      Ground Commercial + zone 8 算出錯誤數字: 改為清空所有計算欄位
#      並在 Notes 標記 *** UNSUPPORTED SERVICE/ZONE - NOT RATED ***
# =====================================================================
import json
import webbrowser
import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk
import pandas as pd
import math
import threading
from contextlib import contextmanager
import traceback
import gc

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("dark-blue")

DEFAULT_CONFIG_PATH = "ups_billing_tool_config.json"


def _app_dir():
    """Where this program lives -- NOT the current working directory.

    The config used to sit at os.getcwd(). That is whatever folder the
    launcher happened to start in: double-clicking, a shortcut with a
    different "Start in", or running from a terminal elsewhere all give
    different answers. Settings were then written to one place and read
    from another, which looks exactly like "nothing is being saved"."""
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except NameError:                      # frozen / interactive
        return os.getcwd()


def default_config_path():
    return os.path.join(_app_dir(), DEFAULT_CONFIG_PATH)
_config_lock = threading.Lock()

# =====================================================================
# MASTER ACC CODE TABLE  (source: "Adjustment 各種資料UPSR槽.xlsx" -> ACC Codes)
# =====================================================================
# AR = ACC accessorial Charge Description Codes (AS) and the surcharge each one
# means. This is the SINGLE source of truth for "is this code known?" -- it
# feeds handled_raw_codes below, the Code Lookup tab, and the built-in dynamic
# surcharge defaults. Codes sharing a description are the same billable
# surcharge charged on a different basis (e.g. LPS / LPR / LCC / LCR / LWC /
# LWR are all Large Package Surcharge: length+girth vs cubic vs weight, and
# commercial vs residential).
UPS_ACC_CODE_REFERENCE = {
    "AHG": "Additional Handling Surcharge - Dimension",
    "AHL": "Additional Handling Surcharge - Dimension",
    "AHV": "Additional Handling Surcharge - Dimension",
    "AHS": "Additional Handling Surcharge - Dimension",
    "AHW": "Additional Handling Surcharge - Weight",
    "AHC": "Additional Handling Surcharge - Packaging",
    "AHP": "Additional Handling Surcharge - Packaging",
    "CAC": "Remote Area Surcharge - Commercial",
    "CAR": "Remote Area Surcharge - Residential",
    "ERL": "Returns Electronic Label",
    "EVS": "Declared Value",
    "FTP": "Third Party Billing Service",
    "ISW": "Intercept Service Web",
    "LCR": "Large Package Surcharge",
    "LDC": "Delivery Area Surcharge - Extended - Commercial",
    "LDR": "Delivery Area Surcharge - Extended - Residential",
    "LPR": "Large Package Surcharge",
    "LPS": "Large Package Surcharge",
    "RDC": "Delivery Area Surcharge - Commercial",
    "RDR": "Delivery Area Surcharge - Residential",
    "RES": "Residential Surcharge",
    "ADS": "Adult Signature Required",
    "AFW": "Future Day Pickup - Alternate Address - Web Request",
    "CSG": "Delivery Confirmation Signature - Commercial",
    "DCS": "Delivery Confirmation Signature",
    "HIC": "Remote Area Surcharge - Commercial",
    "HIR": "Remote Area Surcharge - Residential",
    "IRW": "Reroute - Web Request",
    "LCC": "Large Package Surcharge",
    "LWC": "Large Package Surcharge",
    "LWR": "Large Package Surcharge",
    "OML": "Over Maximum Surcharge",
    "OVR": "Over Maximum Surcharge",
    "REP": "Residential Pickup surcharge",
    "AKR": "Remote Area Surcharge",
    "ART": "Returns 3 UPS Pickup Attempts",
    "NPF": "Not Previously Billed Missing PLD Fee",
    "SCF": "Not Previously Billed Shipping Correction Fee",
}

# Residential / Commercial signal carried by the code itself. Used for
# cross-period pure-ADJ service inference when the description is the generic
# "Shipping Charge Correction Ground".
ACC_RESIDENTIAL_CODES = {"LPR", "LCR", "LWR", "RDR", "LDR", "CAR", "HIR", "RES"}
ACC_COMMERCIAL_CODES = {"LPS", "LCT", "LCC", "LWC", "RDC", "LDC", "CAC", "HIC"}

# Fuel eligibility per surcharge (source: same workbook -> "Fuel Surcharge收錢").
# UPS applies the fuel percentage only to the YES items.
ACC_FUEL_ELIGIBLE = {
    "Residential Surcharge": True,
    "Additional Handling Surcharge - Dimension": True,
    "Additional Handling Surcharge - Weight": True,
    "Additional Handling Surcharge - Packaging": True,
    "Large Package Surcharge": True,
    "Over Maximum Surcharge": True,
    "Intercept Service Web": True,
    "Reroute - Web Request": True,
    "Future Day Pickup - Alternate Address - Web Request": True,
    "Residential Pickup surcharge": True,
    "Delivery Area Surcharge - Commercial": False,
    "Delivery Area Surcharge - Residential": False,
    "Delivery Area Surcharge - Extended - Commercial": False,
    "Delivery Area Surcharge - Extended - Residential": False,
    "Remote Area Surcharge": False,
    "Remote Area Surcharge - Commercial": False,
    "Remote Area Surcharge - Residential": False,
    "Adult Signature Required": False,
    "Delivery Confirmation Signature": False,
    "Delivery Confirmation Signature - Commercial": False,
}

# ACC codes that have NO dedicated reprice bucket in this tool. They are
# registered as built-in dynamic surcharges so they are recognised (never
# reported as Unknown) and so export_acc_template() gives each one a
# FLAT_CHARGES rate row to fill in.
DEFAULT_DYNAMIC_SURCHARGE_MAPPING = {
    code: {
        "ar": "ACC",
        "name": UPS_ACC_CODE_REFERENCE[code],
        "type": "REPRICE",
        "fuel_eligible": ACC_FUEL_ELIGIBLE.get(UPS_ACC_CODE_REFERENCE[code], True),
        "enabled": True,
    }
    for code in ["AFW", "ART", "REP"]
}

# NPF / SCF ("Not Previously Billed ..." fees) stay in UPS_ACC_CODE_REFERENCE so
# they are never reported as unknown codes, but they get NO rate row and are not
# repriced -- they are UPS billing catch-up fees, not a service surcharge.
ACC_CODES_NOT_REPRICED = {"NPF", "SCF"}


# =====================================================================
# FUEL SURCHARGE SCHEDULE
# =====================================================================
# UPS sets the fuel surcharge percentage by SHIP DATE (the invoice's
# Transaction Date), and republishes it weekly. A single global Fuel % therefore
# misprices any invoice that spans a rate change. The schedule below is a list
# of {"start", "end", "percent"} windows; ship date is matched inclusively
# against start..end.
def parse_schedule_date(value):
    """Parse anything UPS / Excel throws at us into a date. None if unusable."""
    if value is None:
        return None

    text = str(value).strip()
    if text == "" or text.lower() in ("nan", "nat", "none"):
        return None

    try:
        parsed = pd.to_datetime(text, errors="coerce")
    except Exception:
        return None

    if parsed is None or pd.isna(parsed):
        return None

    return parsed.date()


def normalize_fuel_schedule(raw_schedule):
    """Clean, validate and sort a fuel schedule.

    Returns (schedule, problems). Rows with an unparseable date, a missing end
    date, or end < start are dropped and reported -- a silently ignored bad row
    would quietly reprice a whole week at the fallback rate.
    """
    schedule = []
    problems = []

    for index, entry in enumerate(raw_schedule or [], start=1):
        start = parse_schedule_date(entry.get("start"))
        end = parse_schedule_date(entry.get("end"))

        try:
            percent = float(str(entry.get("percent", "")).replace("%", "").strip())
        except (ValueError, TypeError):
            percent = None

        if start is None or end is None:
            problems.append(f"Row {index}: unreadable date "
                            f"({entry.get('start')} ~ {entry.get('end')})")
            continue

        if end < start:
            problems.append(f"Row {index}: end date {end} is before start {start}")
            continue

        if percent is None:
            problems.append(f"Row {index}: unreadable Fuel % ({entry.get('percent')})")
            continue

        schedule.append({
            "start": start.isoformat(),
            "end": end.isoformat(),
            "percent": percent,
        })

    schedule.sort(key=lambda item: item["start"])

    # Overlapping windows are ambiguous: the earlier window wins on lookup, so
    # warn instead of picking silently.
    for earlier, later in zip(schedule, schedule[1:]):
        if later["start"] <= earlier["end"]:
            problems.append(
                f"Overlap: {earlier['start']}~{earlier['end']} ({earlier['percent']}%) "
                f"and {later['start']}~{later['end']} ({later['percent']}%)"
            )

    return schedule, problems


def fuel_percent_for_date(schedule, ship_date_value, default_percent):
    """Fuel percent for one shipment.

    Returns (percent_as_fraction, note). note is "" on a clean hit, otherwise a
    string for the row's Notes column so a fallback is never invisible.
    """
    if not schedule:
        return default_percent / 100.0, ""

    ship_date = parse_schedule_date(ship_date_value)

    if ship_date is None:
        return (default_percent / 100.0,
                f"Fuel: no ship date, used default {default_percent}%")

    for window in schedule:
        if window["start"] <= ship_date.isoformat() <= window["end"]:
            return float(window["percent"]) / 100.0, ""

    return (default_percent / 100.0,
            f"Fuel: no schedule for {ship_date.isoformat()}, "
            f"used default {default_percent}%")

# =====================================================================
# AR -- CHARGE CLASSIFICATION CODES
# =====================================================================
# Which AR values carry money and may therefore be registered as a custom
# surcharge. ACC is the usual accessorial bucket; BRK (brokerage) and GOV
# (duty / customs tax) show up on import entries; MSC covers miscellaneous
# billing adjustments. Anything here is also checked for unknown codes --
# previously only ACC was, so a brokerage or duty code was neither priced
# nor reported.
# MSC ("Billing Adjustment for W/E ...") is a WEEKLY, ACCOUNT-level
# adjustment. UPS sometimes staples a tracking number onto it, which used to
# dump a -$4,460 adjustment onto one parcel. It is not repriced; it is
# reported on the Account Charges sheet instead.
BILLABLE_AR_CODES = ("ACC", "BRK", "GOV")

# The engine computes these itself -- base rate and fuel. Registering them as
# a custom surcharge would double-count, so they are not offered.
ENGINE_AR_CODES = ("FRT", "FSC")

# Informational mirror rows. Always $0 net, nothing to price or report.
INFO_AR_CODES = ("INF",)

AR_CODE_DESCRIPTIONS = {
    "ACC": "Accessorial surcharge",
    "BRK": "Brokerage fee (import entry)",
    "GOV": "Government duty / customs tax",
    "MSC": "Miscellaneous billing adjustment",
}

AR_DROPDOWN_CHOICES = [f"{c} - {AR_CODE_DESCRIPTIONS[c]}" for c in BILLABLE_AR_CODES]


def normalize_ar_code(value):
    """'BRK - Brokerage fee (import entry)' -> 'BRK'."""
    raw = str(value or "").strip().upper()
    return raw.split(" - ")[0].strip()


# =====================================================================
# CUSTOM SURCHARGE "TYPE" -- what the tool does with the charge
# =====================================================================
# The old values (TEMPLATE / RAW / AUDIT_ONLY) were stored and exported but
# never read by any calculation, so all three behaved identically. They are
# renamed to say what they actually do, and now they actually do it. Old
# configs and old exported templates keep working via DYNAMIC_TYPE_ALIASES.
DYNAMIC_TYPE_CHOICES = [
    "REPRICE - use my rate table",
    "COPY INVOICE - use the billed amount",
    "REPORT ONLY - show it, exclude from Total",
]

DYNAMIC_TYPE_ALIASES = {
    "TEMPLATE": "REPRICE",
    "RAW": "COPY INVOICE",
    "AUDIT_ONLY": "REPORT ONLY",
    "AUDIT ONLY": "REPORT ONLY",
    "COPY_INVOICE": "COPY INVOICE",
    "REPORT_ONLY": "REPORT ONLY",
}


def normalize_dynamic_type(value):
    """Any old or new spelling -> REPRICE / COPY INVOICE / REPORT ONLY."""
    raw = str(value or "").strip().upper()
    raw = raw.split(" - ")[0].strip()
    raw = DYNAMIC_TYPE_ALIASES.get(raw, raw)
    if raw not in ("REPRICE", "COPY INVOICE", "REPORT ONLY"):
        return "REPRICE"
    return raw


def dynamic_type_choice_label(value):
    """Stored value -> the full combobox label."""
    mode = normalize_dynamic_type(value)
    for choice in DYNAMIC_TYPE_CHOICES:
        if choice.split(" - ")[0].strip() == mode:
            return choice
    return DYNAMIC_TYPE_CHOICES[0]


def _unknown_summary_text(unknown_rows, limit=12):
    """Aggregate unknown charge rows into a short code / count / $ summary."""
    agg = {}
    for r in unknown_rows:
        code = str(r.get("AS Charge Code", "")).strip() or "(blank)"
        desc = str(r.get("AT Description", "")).strip()
        entry = agg.setdefault(code, {"n": 0, "amt": 0.0, "desc": desc})
        entry["n"] += 1
        try:
            entry["amt"] += float(r.get("BA Net Amount", 0) or 0)
        except (TypeError, ValueError):
            pass
        if not entry["desc"] and desc:
            entry["desc"] = desc

    ordered = sorted(agg.items(), key=lambda kv: abs(kv[1]["amt"]), reverse=True)

    lines = []
    for code, info in ordered[:limit]:
        desc = info["desc"][:38]
        lines.append(f"  {code:<6} {info['n']:>6} rows   "
                     f"${info['amt']:>12,.2f}   {desc}")

    if len(ordered) > limit:
        lines.append(f"  ... and {len(ordered) - limit} more code(s)")

    total = sum(v["amt"] for v in agg.values())
    lines.append(f"\n  {len(agg)} unknown code(s), "
                 f"{len(unknown_rows)} row(s), ${total:,.2f} unpriced")
    return "\n".join(lines)


@contextmanager
def managed_workbook(path):
    """Context manager for proper workbook cleanup"""
    from openpyxl import load_workbook
    wb = None
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        yield wb
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

# ADJ detail codes (column AJ) that describe an accessorial, not a service.
# Their AR is FRT, which is why they have to be listed explicitly.
ACCESSORIAL_ADJ_DETAIL_CODES = {
    "ADC",   # Address Correction
    "DIN",   # 2nd mile / delivery intercept
}

UPS_FUEL_SURCHARGE_URL = (
    "https://www.ups.com/us/en/support/shipping-support/"
    "shipping-costs-rates/fuel-surcharges"
)


class UPSRepricingTool:

    def __init__(self, root):
        self.root = root
        self.root.title("UPS Reprice Platform")
        self.root.geometry("1220x760")
        self.root.minsize(1200, 740)
        self.root.configure(fg_color="#f3e9da")

        # FILE PATHS
        self.billing_path = tk.StringVar()
        self.config_path = tk.StringVar(value=default_config_path())

        # RULES - GLOBAL
        self.dim_factor = tk.StringVar(value="300")
        self.fuel_percent = tk.StringVar(value="18.75")

        # RULES - OVERSIZE
        self.use_oversize_longest = tk.BooleanVar(value=True)
        self.use_oversize_weight = tk.BooleanVar(value=True)
        self.use_oversize_cubic = tk.BooleanVar(value=True)
        self.use_oversize_lg = tk.BooleanVar(value=True)

        self.oversize_longest_side = tk.StringVar(value="96")
        self.oversize_actual_weight = tk.StringVar(value="110")
        self.oversize_cubic_inches = tk.StringVar(value="17280")
        self.oversize_length_girth = tk.StringVar(value="130")
        self.oversize_min_billable_weight = tk.StringVar(value="90")

        # RULES - AHS (WITH MANUAL AHS-DIMENSION THRESHOLD)
        self.ahs_weight_threshold = tk.StringVar(value="50")
        self.ahs_longest_side = tk.StringVar(value="48")
        self.ahs_second_side = tk.StringVar(value="30")
        self.ahs_cubic_inches = tk.StringVar(value="10368")
        self.ahs_lg_limit = tk.StringVar(value="105")
        # 15, not 40. Verified against invoice 0000XXXXXXX: across all
        # 690 packages carrying dimensions, billable weight is exactly
        #     max(ceil(entered), ceil(cubic / 300), floor)
        # with floor = 90 for LPS/LPR, 15 for AHS, 1 otherwise. A 40 here
        # inflates the recalculated freight of every small AHS parcel.
        self.ahs_min_billable_weight = tk.StringVar(value="15")

        # Fuel concessions. Some clients are sold fuel at a discount off the
        # published percentage, or with a ceiling on it. Blank / 0 = neither.
        #   discount 10  ->  18.00% becomes 16.20%   (18 x 0.90)
        #   cap      15  ->  18.00% becomes 15.00%
        # Discount is applied first, then the cap.
        #   mode "none"    no concession
        #        "percent"  N % OFF the published rate  (18% -10% -> 16.20%)
        #        "points"   N percentage POINTS off     (18% -2pt -> 16.00%)
        #   cap: an optional ceiling applied after the discount

        # RULES - ACCESSORIAL
        self.residential_fee = tk.StringVar(value="5")
        self.das_com_fee = tk.StringVar(value="7")
        self.das_res_fee = tk.StringVar(value="8")
        self.das_ext_com_fee = tk.StringVar(value="10")
        self.das_ext_res_fee = tk.StringVar(value="12")
        self.remote_com_fee = tk.StringVar(value="15")
        self.remote_res_fee = tk.StringVar(value="18")
        self.ahs_weight_fee = tk.StringVar(value="34")
        self.ahs_dimension_fee = tk.StringVar(value="25")
        self.ahs_packaging_fee = tk.StringVar(value="25")
        self.oversize_fee = tk.StringVar(value="205")
        self.ovr_fee = tk.StringVar(value="0")
        self.signature_fee = tk.StringVar(value="7")
        self.package_protection_fee = tk.StringVar(value="0")
        self.declared_value_markup = tk.StringVar(value="0")

        # Fuel surcharge by ship date. Empty list = use the global Fuel % for
        # everything (original behaviour).
        self.fuel_schedule = []

        self.rate_tables = {}
        self.last_report_path = ""
        self.last_base_rate_df = pd.DataFrame()
        self.accessorial_rate_table = {}
        self.warning_logs = []
        self.current_acc_template_path = ""

        # v2.3: DYNAMIC SURCHARGE CODE LINKING
        self.dynamic_surcharge_code_map = {}
        # Built-in ACC codes with no dedicated reprice bucket (AFW/ART/REP/
        # NPF/SCF). Seeded so a fresh install already knows them.
        self.dynamic_surcharge_mapping = {
            code: dict(cfg)
            for code, cfg in DEFAULT_DYNAMIC_SURCHARGE_MAPPING.items()
        }
        # Built-in default AHS/LPS code registry (from the Step 1 "Codes" table)
        # so testing doesn't require re-importing the code setup every time.
        # rc_class: Residential / Commercial / None (None = reclass-independent).
        self.ahs_lps_code_registry = {
            "AHG": {"charge_type": "AHS - Dimension", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Length+Girth"},
            "AHL": {"charge_type": "AHS - Dimension", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Longest Side"},
            "AHS": {"charge_type": "AHS - Dimension", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Second Longest Side"},
            "AHV": {"charge_type": "AHS - Dimension", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Volume / Dimension"},
            "AHW": {"charge_type": "AHS - Weight", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Weight"},
            "AHC": {"charge_type": "AHS - Packaging", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Packaging"},
            "AHP": {"charge_type": "AHS - Packaging", "category": "AHS", "rc_class": "None", "description": "Additional Handling - Packaging"},
            "LPS": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Commercial", "description": "Large Package Surcharge (Commercial)"},
            "LPR": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Residential", "description": "Large Package Surcharge (Residential)"},
            "LCR": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Residential", "description": "Large Package Surcharge Resi - Cubic"},
            "LCT": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Commercial", "description": "Large Package Surcharge - Cubic"},
            "LCC": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Commercial", "description": "Large Package Surcharge Comm - Cubic Volume"},
            "LWC": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Commercial", "description": "Large Package Surcharge Comm - Weight"},
            "LWR": {"charge_type": "Large Package", "category": "LPS", "rc_class": "Residential", "description": "Large Package Surcharge Resi - Weight"},
        }
        
        style = ttk.Style()
        style.theme_use("default")

        # ---- Dark milk-tea palette ----
        BG = "#cbb189"        # medium milk-tea background
        PANEL = "#dbc6a4"      # lighter latte panel / card
        TAB_BG = "#bfa477"     # tab strip
        ACCENT = "#a9743f"     # caramel accent (buttons / selected tab)
        TEXT = "#3b2f23"       # dark espresso text (readable on light bg)
        LINE = "#a98c63"       # subtle border
        self.MILKTEA = dict(BG=BG, PANEL=PANEL, TAB_BG=TAB_BG,
                            ACCENT=ACCENT, TEXT=TEXT, LINE=LINE)

        self.root.configure(fg_color=BG)

        style.configure("TFrame", background=BG)
        style.configure("TLabelframe", background=BG, foreground=TEXT,
                        bordercolor=LINE, relief="groove")
        style.configure("TLabelframe.Label", background=BG, foreground="#7a4f24",
                        font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", background=BG, foreground=TEXT,
                        font=("Segoe UI", 10))
        style.configure("TButton", background=TAB_BG, foreground=TEXT,
                        font=("Segoe UI", 9, "bold"), padding=[8, 4])
        style.map("TButton", background=[("active", ACCENT)],
                  foreground=[("active", "white")])
        style.configure("TEntry", fieldbackground=PANEL)
        style.configure("TCombobox", fieldbackground=PANEL, background=PANEL)

        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", background=TAB_BG, foreground=TEXT,
                        padding=[10, 5], font=("Segoe UI", 9, "bold"))
        style.map("TNotebook.Tab", background=[("selected", ACCENT)],
                  foreground=[("selected", "white")])
        style.configure("Treeview", background=PANEL, foreground=TEXT,
                        rowheight=22, fieldbackground=PANEL, borderwidth=0,
                        font=("Segoe UI", 10))
        style.configure("Treeview.Heading", background=ACCENT,
                        foreground="white", font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[("selected", "#e3c79a")],
                  foreground=[("selected", TEXT)])

        header = ctk.CTkLabel(self.root, text="📦 UPS Reprice Platform\nDeveloped by Sandy Liu", font=("Segoe UI", 14, "bold"), text_color="#7a4f24", justify="center")
        header.pack(pady=(14, 6))

        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_files = ttk.Frame(notebook)
        self.tab_rules = ttk.Frame(notebook)
        self.tab_rates = ttk.Frame(notebook)
        self.tab_ahs_lps_codes = ttk.Frame(notebook)
        self.tab_dynamic_surcharge = ttk.Frame(notebook)
        self.tab_surcharge_codes = ttk.Frame(notebook)
        self.tab_run = ttk.Frame(notebook)
        self.tab_code_lookup = ttk.Frame(notebook)

        notebook.add(self.tab_files, text="📁 Import Files")
        notebook.add(self.tab_rules, text="⚙ Pricing Rules")
        notebook.add(self.tab_code_lookup, text="🔎 Code Lookup")

        self._build_files_tab()
        self._build_rules_tab()
        self._build_code_lookup_tab()

    # =========================================================
    # TAB 1: FILES
    # =========================================================
    def _build_files_tab(self):
        frm = ttk.LabelFrame(self.tab_files, text="📁 Import Files")
        frm.pack(fill="x", padx=12, pady=12)

        fields = [
            ("Billing file", self.billing_path, self._pick_billing),
        ]

        for i, (label, var, cmd) in enumerate(fields):
            ttk.Label(frm, text=label).grid(row=i, column=0, sticky="w", padx=8, pady=8)
            ttk.Entry(frm, textvariable=var, width=88).grid(row=i, column=1, padx=8, pady=8)
            ttk.Button(frm, text="Browse", command=cmd).grid(row=i, column=2, padx=8, pady=8)

        self.base_rate_service = tk.StringVar(value="Ground Commercial")
        row1 = ttk.Frame(frm)
        row1.grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=6)

        service_combo = ttk.Combobox(row1, textvariable=self.base_rate_service, values=[
            "Ground Commercial", "Ground Residential", "SurePost", "3 Day Select", "2nd Day Air", "2nd Day Air AM",
            "Next Day Air Saver", "Next Day Air", "Next Day Air Early", "Standard", "Worldwide Expedited",
            "Worldwide Saver", "Worldwide Express", "Worldwide Express Plus", "Worldwide Economy"
        # 24, not 18: "Worldwide Express Plus" is 22 characters and was being
        # clipped in the closed box.
        ], state="readonly", width=24)
        service_combo.pack(side="left", padx=6)

        self.current_service_label = ttk.Label(row1, text="Currently Editing: Ground Commercial", font=("Segoe UI", 9, "bold"))
        self.current_service_label.pack(side="left", padx=10)
        service_combo.bind("<<ComboboxSelected>>", self.on_service_changed)

        rate_frame = ttk.LabelFrame(frm, text="🚛 Base Rates")
        rate_frame.grid(row=3, column=0, columnspan=3, sticky="w", padx=8, pady=10)

        ttk.Button(rate_frame, text="Export Base Rate", width=18, command=self.export_base_rate_template).grid(row=0, column=0, padx=4, pady=4)
        ttk.Button(rate_frame, text="Import Base Rate", width=18, command=self.import_base_rate_template).grid(row=0, column=1, padx=4, pady=4)

        bottom_row = ttk.Frame(frm)
        bottom_row.grid(row=4, column=0, columnspan=3, sticky="w", padx=8, pady=10)

        ttk.Button(bottom_row, text="Save Settings", command=self.save_config).pack(side="left", padx=4)
        ttk.Button(bottom_row, text="Load Settings", command=self.load_config).pack(side="left", padx=4)

        # ---- Generate (moved here from the old Generate Report tab) ----
        gen_frame = ttk.Frame(self.tab_files)
        gen_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        ctk.CTkButton(gen_frame, text="🚚 Generate Reprice Report",
                      fg_color="#a9743f", hover_color="#8c5e30",
                      corner_radius=12, height=42,
                      font=("Segoe UI", 13, "bold"),
                      command=self.run_repricing).pack(anchor="w", pady=12)

        ctk.CTkButton(gen_frame, text="📊 Generate Profit Report",
                      fg_color="#8c5a4a", hover_color="#74473a",
                      corner_radius=12, height=42,
                      font=("Segoe UI", 13, "bold"),
                      command=self.run_profit_report).pack(anchor="w",
                                                           pady=(0, 12))

        self.status_var = tk.StringVar(value="System Ready")
        ttk.Label(gen_frame, textvariable=self.status_var).pack(anchor="w")

        # Hidden preview holder (engine writes here; not shown as a tab)
        self.preview = ttk.Treeview(gen_frame, columns=("info",),
                                    show="headings", height=1)
        self.preview.heading("info", text="")

    def on_service_changed(self, event=None):
        selected = self.base_rate_service.get()
        self.current_service_label.config(text=f"Currently Editing: {selected}")


    # =========================================================
    # TAB 2: RULES
    # =========================================================
    def _build_rules_tab(self):

        main_container = ttk.Frame(self.tab_rules)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        top_frame = ttk.Frame(main_container)
        top_frame.pack(fill="x", pady=6)

        global_frame = ttk.LabelFrame(top_frame, text="⚙ Global Settings")
        global_frame.pack(side="left", fill="both", expand=True, padx=6)

        ttk.Label(global_frame, text="DIM Factor").grid(row=0, column=0, padx=8, pady=8, sticky="w")
        ttk.Entry(global_frame, textvariable=self.dim_factor, width=12).grid(row=0, column=1, padx=8, pady=8)

        # Fuel % is NOT a global setting any more -- it lives inside the fuel
        # schedule window as the "no range matched" default, so there is only
        # one place that decides fuel. self.fuel_percent is kept for config
        # backward compatibility.

        oversize_frame = ttk.LabelFrame(top_frame, text="📏 Oversize Rules")
        oversize_frame.pack(side="left", fill="both", expand=True, padx=6)

        ttk.Checkbutton(oversize_frame, text="Enable Longest Side", variable=self.use_oversize_longest).grid(row=0, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(oversize_frame, textvariable=self.oversize_longest_side, width=12).grid(row=0, column=1, padx=8)

        ttk.Checkbutton(oversize_frame, text="Enable Actual Weight", variable=self.use_oversize_weight).grid(row=1, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(oversize_frame, textvariable=self.oversize_actual_weight, width=12).grid(row=1, column=1, padx=8)

        ttk.Checkbutton(oversize_frame, text="Enable Cubic Inches", variable=self.use_oversize_cubic).grid(row=2, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(oversize_frame, textvariable=self.oversize_cubic_inches, width=12).grid(row=2, column=1, padx=8)

        ttk.Checkbutton(oversize_frame, text="Enable Length + Girth", variable=self.use_oversize_lg).grid(row=3, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(oversize_frame, textvariable=self.oversize_length_girth, width=12).grid(row=3, column=1, padx=8)

        ttk.Label(oversize_frame, text="Oversize Min Billable Weight").grid(row=4, column=0, padx=8, pady=5, sticky="w")
        ttk.Entry(oversize_frame, textvariable=self.oversize_min_billable_weight, width=12).grid(row=4, column=1, padx=8)

        ahs_frame = ttk.LabelFrame(top_frame, text="📦 AHS Rules")
        ahs_frame.pack(side="left", fill="both", expand=True, padx=6)

        ttk.Label(ahs_frame, text="AHS Weight Threshold").grid(row=0, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_weight_threshold, width=12).grid(row=0, column=1, padx=8)

        ttk.Label(ahs_frame, text="AHS-Dimension Longest Side (inches)").grid(row=1, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_longest_side, width=12).grid(row=1, column=1, padx=8)

        ttk.Label(ahs_frame, text="AHS-Dimension Second Longest Side").grid(row=2, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_second_side, width=12).grid(row=2, column=1, padx=8)

        ttk.Label(ahs_frame, text="AHS Cubic Inches").grid(row=3, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_cubic_inches, width=12).grid(row=3, column=1, padx=8)

        ttk.Label(ahs_frame, text="AHS L+G Limit").grid(row=4, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_lg_limit, width=12).grid(row=4, column=1, padx=8)

        ttk.Label(ahs_frame, text="AHS Min Billable Weight").grid(row=5, column=0, sticky="w", padx=8, pady=5)
        ttk.Entry(ahs_frame, textvariable=self.ahs_min_billable_weight, width=12).grid(row=5, column=1, padx=8)

        # Fuel schedule lives in its own window so this tab's layout is
        # untouched -- only one button and a count label are added here.
        ttk.Label(global_frame, text="Fuel Surcharge").grid(
            row=1, column=0, padx=8, pady=8, sticky="w")

        fuel_btn_cell = ttk.Frame(global_frame)
        fuel_btn_cell.grid(row=1, column=1, padx=8, pady=8, sticky="w")

        ttk.Button(fuel_btn_cell, text="⛽ Set Schedule...", width=18,
                   command=self.open_fuel_schedule_window).pack(side="left")

        self.fuel_schedule_summary = tk.StringVar()
        ttk.Label(fuel_btn_cell, textvariable=self.fuel_schedule_summary,
                  font=("Segoe UI", 8)).pack(side="left", padx=6)

        # UPS republishes the ground fuel percentage every Monday, so the
        # schedule goes stale on its own. Its own row: side by side this
        # widened Global Settings to 687px and pushed AHS Rules off the
        # right edge of the window.
        ttk.Button(global_frame, text="🌐 UPS Fuel Rates", width=18,
                   command=lambda: self.open_url(UPS_FUEL_SURCHARGE_URL)
                   ).grid(row=2, column=1, padx=8, pady=(0, 8), sticky="w")

        acc_frame = ttk.LabelFrame(main_container, text="💵 Surcharge Rates")
        acc_frame.pack(fill="both", expand=True, pady=10)

        btn_frame = ttk.Frame(acc_frame)
        btn_frame.pack(fill="x", padx=10, pady=8)

        ttk.Button(btn_frame, text="➕ Custom Surcharges...", width=24, command=self.open_custom_surcharge_window).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="💾 Save Rates", width=20, command=self.save_accessorial_rates).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="Export Surcharge Rates", width=30, command=self.export_acc_template).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="Import Surcharge Rates", width=30, command=self.import_acc_template).pack(side="left", padx=4)

        table_container = ttk.Frame(acc_frame)
        table_container.pack(fill="both", expand=True, padx=10, pady=10)

        self.acc_canvas = tk.Canvas(table_container, bg="#cbb189", highlightthickness=0)
        self.acc_y_scroll = ttk.Scrollbar(table_container, orient="vertical", command=self.acc_canvas.yview)
        self.acc_x_scroll = ttk.Scrollbar(table_container, orient="horizontal", command=self.acc_canvas.xview)

        self.acc_y_scroll.pack(side="right", fill="y")
        self.acc_x_scroll.pack(side="bottom", fill="x")
        self.acc_canvas.pack(side="left", fill="both", expand=True)

        self.acc_canvas.configure(xscrollcommand=self.acc_x_scroll.set, yscrollcommand=self.acc_y_scroll.set)

        self.acc_display_frame = ttk.Frame(self.acc_canvas)
        self.acc_canvas.create_window((0, 0), window=self.acc_display_frame, anchor="nw")

        def on_frame_configure(event):
            self.acc_canvas.configure(scrollregion=self.acc_canvas.bbox("all"))

        self.acc_display_frame.bind("<Configure>", on_frame_configure)
        self.acc_canvas.bind("<Configure>", lambda e: self.acc_canvas.configure(scrollregion=self.acc_canvas.bbox("all")))

        self.accessorial_schema = {
            "AHS - Weight": ["Zone 2", "Zone 3-4", "Zone 5-6", "Zone 7+"],
            "AHS - Dimension": ["Zone 2", "Zone 3-4", "Zone 5-6", "Zone 7+"],
            "AHS - Packaging": ["Zone 2", "Zone 3-4", "Zone 5-6", "Zone 7+"],
        }

        self.accessorial_rates = {}
        self.render_accessorial_tables()

    def dynamic_acc_columns(self):
        """
        Enabled AR=ACC custom surcharges as an ordered list of
        (as_code, surcharge_name, fuel_eligible).

        This is the SINGLE place that decides which custom surcharge columns
        exist. The calc loop, the CAL Total base, the fuel base and every
        export layout all read from here, so a code added on the Custom
        Surcharges tab can never appear in one of them and be missing from
        another.
        """
        out = []
        for as_code, cfg in (getattr(self, "dynamic_surcharge_mapping", {}) or {}).items():
            if not isinstance(cfg, dict):
                continue
            if normalize_ar_code(cfg.get("ar", "")) not in BILLABLE_AR_CODES:
                continue
            if not cfg.get("enabled", True):
                continue
            name = str(cfg.get("name", "")).strip()
            if name == "":
                continue
            out.append((
                str(as_code).strip().upper(),
                name,
                bool(cfg.get("fuel_eligible", True)),
                normalize_dynamic_type(cfg.get("type", "REPRICE")),
            ))
        return out

    def dynamic_acc_report_only_names(self):
        """Custom ACC names set to REPORT ONLY -- shown, but not in CAL Total."""
        return {name for _as, name, _fuel, mode in self.dynamic_acc_columns()
                if mode == "REPORT ONLY"}

    def dynamic_acc_names(self, exclude=()):
        """Custom ACC column names, de-duplicated and minus anything in `exclude`."""
        excluded = set(exclude)
        names = []
        for _as_code, name, _fuel, _mode in self.dynamic_acc_columns():
            if name not in excluded and name not in names:
                names.append(name)
        return names

    def dynamic_acc_nonfuel_names(self):
        """Custom ACC names explicitly flagged Fuel Eligible = NO."""
        blocked = set()
        for _as_code, name, fuel_ok, _mode in self.dynamic_acc_columns():
            if not fuel_ok:
                blocked.add(name)
        return blocked

    def save_accessorial_rates(self):
        """Read every rate Entry back into accessorial_rate_table and persist."""
        table = getattr(self, "accessorial_rate_table", {})
        bad = []

        for (key, zone), var in getattr(self, "acc_rate_vars", {}).items():
            raw = var.get().strip()
            if raw == "":
                raw = "0"
            try:
                amount = float(str(raw).replace("$", "").replace(",", ""))
            except Exception:
                bad.append(f"{key}" + (f" / Zone {zone}" if zone is not None else ""))
                continue

            if zone is None:
                table[key] = amount
            else:
                if not isinstance(table.get(key), dict):
                    table[key] = {}
                table[key][zone] = amount

        self.accessorial_rate_table = table

        try:
            self.save_config()
        except Exception as e:
            print("Auto save config after surcharge rate edit failed:", e)

        self.render_accessorial_tables()
        try:
            self.acc_canvas.update_idletasks()
        except Exception:
            pass

        if bad:
            messagebox.showwarning(
                "Saved with warnings",
                "These fields were not numeric and were left unchanged:\n\n"
                + "\n".join(bad[:15])
            )
        else:
            self.set_status("Surcharge rates saved.")
    def render_accessorial_tables(self):

        for widget in self.acc_display_frame.winfo_children():
            widget.destroy()

        default_table = {
            ("AHS Weight", "Ground Commercial"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("AHS Weight", "Ground Residential"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("AHS Dimension", "Ground Commercial"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("AHS Dimension", "Ground Residential"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("AHS Packaging", "Ground Commercial"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("AHS Packaging", "Ground Residential"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("Large Package", "Ground Commercial"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            ("Large Package", "Ground Residential"): {"2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0, "8": 0, "44": 0, "45": 0, "46": 0},
            "DAS Commercial": 0,
            "DAS Residential": 0,
            "DAS Extended Commercial": 0,
            "DAS Extended Residential": 0,
            "Remote Area Commercial": 0,
            "Remote Area Residential": 0,
            "Remote Area - AK": 0,
            "Remote Area - HI": 0,
            "Over Maximum Size Surcharge": 0,
            "Residential Surcharge": 0,
            "Return To Sender": 0,
            "Reroute": 0,
            "Reschedule Delivery": 0,
            "Returns Electronic Label": 0,
            "Returns Print Label": 0,
            "Signature": 0,
            "Adult Signature": 0,
            "Package Protection": 0,
            "Address Correction": 0,
            "Direct Delivery Only": 0,
            "Declared Value Markup %": 0,
        }

        existing_table = getattr(self, "accessorial_rate_table", {})

        if not isinstance(existing_table, dict):
            existing_table = {}

        merged_table = {}

        for key, value in default_table.items():
            if key in existing_table:
                merged_table[key] = existing_table[key]
            else:
                merged_table[key] = value

        for key, value in existing_table.items():
            if key not in merged_table:
                merged_table[key] = value

        for as_code, cfg in getattr(self, "dynamic_surcharge_mapping", {}).items():
            surcharge_name = str(cfg.get("name", "")).strip()
            if surcharge_name and surcharge_name not in merged_table:
                merged_table[surcharge_name] = 0

        self.accessorial_rate_table = merged_table
        table = self.accessorial_rate_table

        # Rate cells are editable Entry widgets (they used to be read-only
        # Labels that merely looked like inputs). Every widget is registered in
        # acc_rate_vars so "Save Rates" can read them all back.
        self.acc_rate_vars = {}

        for key, value in table.items():
            frame = ttk.LabelFrame(self.acc_display_frame, text=str(key))
            frame.pack(side="left", padx=10, pady=10, anchor="n")

            if isinstance(value, dict):
                ttk.Label(frame, text="Zone", width=12).grid(row=0, column=0, padx=4, pady=4)
                ttk.Label(frame, text="Rate", width=12).grid(row=0, column=1, padx=4, pady=4)

                row_idx = 1
                for zone, rate in value.items():
                    ttk.Label(frame, text=f"Zone {zone}", width=12).grid(row=row_idx, column=0, padx=4, pady=3)
                    var = tk.StringVar(value=str(rate))
                    ttk.Entry(frame, textvariable=var, width=12).grid(row=row_idx, column=1, padx=4, pady=3)
                    self.acc_rate_vars[(key, zone)] = var
                    row_idx += 1
            else:
                ttk.Label(frame, text="Amount", width=12).grid(row=0, column=0, padx=4, pady=4)
                var = tk.StringVar(value=str(value))
                ttk.Entry(frame, textvariable=var, width=12).grid(row=1, column=0, padx=4, pady=3)
                self.acc_rate_vars[(key, None)] = var


    # =========================================================
    # TAB 3: RATE TABLES
    # =========================================================
    def _build_rates_tab(self):
        mt = self.MILKTEA
        frm = ttk.Frame(self.tab_rates)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        # ---- Header card ----
        card = ctk.CTkFrame(frm, fg_color=mt["PANEL"], corner_radius=14)
        card.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(card, text="📊  Base Rate Table",
                     font=("Segoe UI", 15, "bold"),
                     text_color="#7a4f24").pack(anchor="w", padx=16, pady=(12, 2))
        ctk.CTkLabel(card,
                     text="Freight rates by billable weight (rows) and zone "
                          "(columns). Import a template to load, or filter below.",
                     font=("Segoe UI", 10),
                     text_color=mt["TEXT"]).pack(anchor="w", padx=16, pady=(0, 4))

        toolbar = ctk.CTkFrame(card, fg_color="transparent")
        toolbar.pack(fill="x", padx=12, pady=(0, 12))

        ctk.CTkButton(toolbar, text="📤 Export Template",
                      fg_color="#8b5e34", hover_color="#6f4a28",
                      corner_radius=10, height=34,
                      command=self.export_base_rate_template
                      ).pack(side="left", padx=4)

        ctk.CTkLabel(toolbar, text="🔍 Filter weight:",
                     font=("Segoe UI", 10), text_color=mt["TEXT"]
                     ).pack(side="left", padx=(16, 4))
        self.rate_filter_var = tk.StringVar()
        ent = ctk.CTkEntry(toolbar, textvariable=self.rate_filter_var,
                           width=110, placeholder_text="e.g. 15")
        ent.pack(side="left", padx=4)
        ent.bind("<KeyRelease>", lambda e: self._apply_rate_filter())
        ctk.CTkButton(toolbar, text="Clear", width=60,
                      fg_color="#b89a6e", hover_color="#9c8055",
                      corner_radius=10, height=34,
                      command=self._clear_rate_filter).pack(side="left", padx=4)

        self.rate_status = ctk.CTkLabel(
            toolbar, text="No base rates loaded.",
            font=("Segoe UI", 10, "italic"), text_color="#8b6a45")
        self.rate_status.pack(side="right", padx=10)

        # ---- Table with both scrollbars ----
        table_wrap = ttk.Frame(frm)
        table_wrap.pack(fill="both", expand=True)

        cols = ["Service", "Weight", "Zone 2", "Zone 3", "Zone 4", "Zone 5",
                "Zone 6", "Zone 7", "Zone 8", "Zone 44", "Zone 45", "Zone 46"]
        self.rate_preview = ttk.Treeview(table_wrap, columns=cols,
                                         show="headings", height=20)
        for c in cols:
            self.rate_preview.heading(c, text=c)
            self.rate_preview.column(
                c, width=130 if c == "Service" else 80,
                anchor="w" if c == "Service" else "e")

        vsb = ttk.Scrollbar(table_wrap, orient="vertical",
                            command=self.rate_preview.yview)
        hsb = ttk.Scrollbar(table_wrap, orient="horizontal",
                            command=self.rate_preview.xview)
        self.rate_preview.configure(yscrollcommand=vsb.set,
                                    xscrollcommand=hsb.set)
        self.rate_preview.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_wrap.rowconfigure(0, weight=1)
        table_wrap.columnconfigure(0, weight=1)

        # zebra striping tags
        self.rate_preview.tag_configure("odd", background="#dbc6a4")
        self.rate_preview.tag_configure("even", background="#d0bb98")

    def _populate_rate_preview(self, df):
        """Fill the rate table from a dataframe with zebra striping."""
        # The on-screen rate preview was removed; skip safely if absent.
        if not hasattr(self, "rate_preview"):
            return
        self.rate_preview.delete(*self.rate_preview.get_children())
        cols = list(self.rate_preview["columns"])
        for i, (_, r) in enumerate(df.iterrows()):
            vals = [r.get(c, "") for c in cols]
            tag = "even" if i % 2 == 0 else "odd"
            self.rate_preview.insert("", "end", values=vals, tags=(tag,))
        if hasattr(self, "rate_status"):
            self.rate_status.configure(
                text=f"{len(df)} rows loaded.")

    def _apply_rate_filter(self):
        q = self.rate_filter_var.get().strip()
        df = getattr(self, "last_base_rate_df", None)
        if df is None or df.empty:
            return
        if q == "":
            self._populate_rate_preview(df.head(500))
            return
        try:
            wt = int(float(q))
            sub = df[pd.to_numeric(df["Weight"], errors="coerce") == wt]
            self._populate_rate_preview(sub)
        except ValueError:
            self._populate_rate_preview(df.head(500))

    def _clear_rate_filter(self):
        self.rate_filter_var.set("")
        df = getattr(self, "last_base_rate_df", None)
        if df is not None and not df.empty:
            self._populate_rate_preview(df.head(500))

    # =========================================================
    # TAB 4: AHS/LPS CODES
    # =========================================================
    def _build_ahs_lps_codes_tab(self):
        
        main = ttk.Frame(self.tab_ahs_lps_codes)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        title = ttk.Label(main, text="Charge Code Setup", font=("Segoe UI", 13, "bold"))
        title.pack(anchor="w", pady=(0, 10))

        desc = ttk.Label(main, text="Set which UPS charge codes should be treated as AHS or Large Package charges.")
        desc.pack(anchor="w", pady=(0, 12))

        self.ahs_lps_code_tree = ttk.Treeview(
            main, 
            columns=("CODE", "CHARGE_TYPE", "CATEGORY", "RC_CLASS", "DESCRIPTION"), 
            show="headings", 
            height=14
        )

        cols = [
            ("CODE", "UPS Charge Code", 120),
            ("CHARGE_TYPE", "Maps To", 180),
            ("CATEGORY", "Category", 110),
            ("RC_CLASS", "R/C Class", 110),
            ("DESCRIPTION", "Description", 240),
        ]

        for key, title_text, width in cols:
            self.ahs_lps_code_tree.heading(key, text=title_text)
            self.ahs_lps_code_tree.column(key, width=width)

        self.ahs_lps_code_tree.pack(fill="both", expand=True, pady=10)

        btn_row = ttk.Frame(main)
        btn_row.pack(fill="x", pady=10)

        ttk.Button(btn_row, text="+ Add Code", command=self.open_ahs_lps_popup).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Delete Selected", command=self.delete_ahs_lps_code).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Save Setup", command=self.save_ahs_lps_registry).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Load Setup", command=self.load_ahs_lps_registry).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Export List", command=self.export_ahs_lps_code_template).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Import List", command=self.import_ahs_lps_code_template).pack(side="left", padx=4)

        # Show the built-in default registry immediately (no import needed).
        try:
            self.load_ahs_lps_registry()
        except Exception as e:
            print("populate default AHS/LPS registry failed:", e)

    def open_ahs_lps_popup(self):
        
        popup = tk.Toplevel(self.root)
        popup.title("Add Charge Code")
        popup.geometry("520x380")

        container = ttk.Frame(popup)
        container.pack(fill="both", expand=True, padx=16, pady=16)

        ttk.Label(container, text="UPS Charge Code").grid(row=0, column=0, sticky="w", pady=8)
        code_entry = ttk.Entry(container, width=35)
        code_entry.grid(row=0, column=1, padx=8, pady=8)

        ttk.Label(container, text="Maps To").grid(row=1, column=0, sticky="w", pady=8)
        charge_type_var = tk.StringVar(value="AHS - Weight")
        charge_type_combo = ttk.Combobox(
            container, 
            textvariable=charge_type_var,
            values=[
                "AHS - Weight",
                "AHS - Dimension",
                "AHS - Packaging",
                "Large Package",
            ],
            state="readonly",
            width=32
        )
        charge_type_combo.grid(row=1, column=1, padx=8, pady=8)

        ttk.Label(container, text="Category").grid(row=2, column=0, sticky="w", pady=8)
        category_var = tk.StringVar(value="AHS")
        category_combo = ttk.Combobox(
            container,
            textvariable=category_var,
            values=["AHS", "LPS", "OTHER"],
            state="readonly",
            width=32
        )
        category_combo.grid(row=2, column=1, padx=8, pady=8)

        ttk.Label(container, text="R/C Class").grid(row=3, column=0, sticky="w", pady=8)
        rc_class_var = tk.StringVar(value="None")
        rc_class_combo = ttk.Combobox(
            container,
            textvariable=rc_class_var,
            values=["Residential", "Commercial", "None"],
            state="readonly",
            width=32
        )
        rc_class_combo.grid(row=3, column=1, padx=8, pady=8)

        ttk.Label(container, text="Description (optional)").grid(row=4, column=0, sticky="w", pady=8)
        description_entry = ttk.Entry(container, width=35)
        description_entry.grid(row=4, column=1, padx=8, pady=8)

        def save_popup():
            ups_code = code_entry.get().strip().upper()
            maps_to = charge_type_var.get()
            category = category_var.get()
            rc_class = rc_class_var.get()
            description = description_entry.get().strip()

            if ups_code == "":
                messagebox.showerror("Error", "UPS Charge Code cannot be empty")
                return

            self.ahs_lps_code_tree.insert("", "end", values=(
                ups_code,
                maps_to,
                category,
                rc_class,
                description
            ))

            try:
                self.save_ahs_lps_registry()
            except Exception as e:
                print("Auto save AHS/LPS registry failed:", e)

            popup.destroy()

        ttk.Button(container, text="Save Code", command=save_popup).grid(row=5, column=0, columnspan=2, pady=20)

    def delete_ahs_lps_code(self):
        selected = self.ahs_lps_code_tree.selection()
        for item in selected:
            self.ahs_lps_code_tree.delete(item)
        try:
            self.save_ahs_lps_registry()
        except Exception as e:
            print("Auto save failed:", e)

    def save_ahs_lps_registry(self):
        
        self.ahs_lps_code_registry = {}

        for item in self.ahs_lps_code_tree.get_children():
            vals = self.ahs_lps_code_tree.item(item, "values")
            if not vals or len(vals) < 2:
                continue

            ups_code = str(vals[0]).strip().upper()
            maps_to = str(vals[1]).strip()

            if ups_code == "":
                continue

            self.ahs_lps_code_registry[ups_code] = {
                "charge_type": maps_to,
                "category": str(vals[2]).strip() if len(vals) > 2 else "AHS",
                "rc_class": str(vals[3]).strip() if len(vals) > 3 else "None",
                "description": str(vals[4]).strip() if len(vals) > 4 else ""
            }

        try:
            self.save_config()
        except Exception as e:
            print("Auto save config failed:", e)

        self.set_status("Charge code setup saved.")
    def load_ahs_lps_registry(self):
        
        self.ahs_lps_code_tree.delete(*self.ahs_lps_code_tree.get_children())

        for ups_code, cfg in self.ahs_lps_code_registry.items():
            self.ahs_lps_code_tree.insert("", "end", values=(
                ups_code,
                cfg.get("charge_type", "AHS - Weight"),
                cfg.get("category", "AHS"),
                cfg.get("rc_class", "None"),
                cfg.get("description", "")
            ))

    def export_ahs_lps_code_template(self):
        """Export AHS/LPS code registry as template"""
        
        path = filedialog.asksaveasfilename(
            initialfile=self.export_name_for("ahs_lps", "AHS_LPS_Code_Template.xlsx")[0],
            initialdir=self.export_name_for("ahs_lps", "AHS_LPS_Code_Template.xlsx")[1] or ".",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        rows = []

        for ups_code, cfg in self.ahs_lps_code_registry.items():
            rows.append({
                "UPS Charge Code": ups_code,
                "Maps To System Charge": cfg.get("charge_type", "AHS - Weight"),
                "Category": cfg.get("category", "AHS"),
                "R/C Class": cfg.get("rc_class", "None"),
                "Description": cfg.get("description", "")
            })

        if not rows:
            rows = [
                {
                    "UPS Charge Code": "AHG",
                    "Maps To System Charge": "AHS - Weight",
                    "Category": "AHS",
                    "Description": "AHS Weight - Group Rate"
                },
                {
                    "UPS Charge Code": "AHW",
                    "Maps To System Charge": "AHS - Weight",
                    "Category": "AHS",
                    "Description": "AHS Weight - Weighted Rate"
                },
                {
                    "UPS Charge Code": "AHL",
                    "Maps To System Charge": "AHS - Weight",
                    "Category": "AHS",
                    "Description": "AHS Weight - List Rate"
                },
                {
                    "UPS Charge Code": "AHD",
                    "Maps To System Charge": "AHS - Dimension",
                    "Category": "AHS",
                    "Description": "AHS Dimension"
                },
                {
                    "UPS Charge Code": "AHB",
                    "Maps To System Charge": "AHS - Dimension",
                    "Category": "AHS",
                    "Description": "AHS Dimension (Alt)"
                },
                {
                    "UPS Charge Code": "AHP",
                    "Maps To System Charge": "AHS - Packaging",
                    "Category": "AHS",
                    "Description": "AHS Packaging"
                },
                {
                    "UPS Charge Code": "AHC",
                    "Maps To System Charge": "AHS - Packaging",
                    "Category": "AHS",
                    "Description": "AHS Packaging (Alt)"
                },
                {
                    "UPS Charge Code": "LPS",
                    "Maps To System Charge": "Large Package",
                    "Category": "LPS",
                    "Description": "Large Package Surcharge"
                },
            ]

        pd.DataFrame(rows).to_excel(path, index=False)
        self.set_status(f"Charge code setup exported:\n{path}")
    def import_ahs_lps_code_template(self):
        """Import AHS/LPS code registry from template"""
        
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])

        if not path:
            return
        self.remember_import("ahs_lps", path)

        try:
            df = pd.read_excel(path).fillna("")

            required = ["UPS Charge Code", "Maps To System Charge"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise Exception(f"Charge code setup file missing column(s): {missing}")

            self.ahs_lps_code_registry = {}
            self.ahs_lps_code_tree.delete(*self.ahs_lps_code_tree.get_children())

            for _, row in df.iterrows():
                ups_code = str(row.get("UPS Charge Code", "")).strip().upper()

                if ups_code == "":
                    continue

                cfg = {
                    "charge_type": str(row.get("Maps To System Charge", "AHS - Weight")).strip(),
                    "category": str(row.get("Category", "AHS")).strip(),
                    "rc_class": str(row.get("R/C Class", "None")).strip() or "None",
                    "description": str(row.get("Description", "")).strip()
                }

                self.ahs_lps_code_registry[ups_code] = cfg

                self.ahs_lps_code_tree.insert("", "end", values=(
                    ups_code,
                    cfg["charge_type"],
                    cfg["category"],
                    cfg["rc_class"],
                    cfg["description"]
                ))

            try:
                self.save_config()
            except Exception as e:
                print("Auto save config failed:", e)

            self.confirm("Import complete",
                         f"Charge code setup imported.\n\n"
                         f"{os.path.basename(path)}\n"
                         f"Total codes: {len(self.ahs_lps_code_registry)}")
        except Exception as e:
            messagebox.showerror("Import Failed", f"Error: {e}")


    # =========================================================
    # TAB 5: DYNAMIC SURCHARGE
    # =========================================================
    def open_custom_surcharge_window(self):
        """
        Custom Surcharges UI. It used to live on its own notebook tab, but that
        tab was never added to the notebook and its builder was never called,
        so there was no way to reach it. It now opens as its own window --
        same pattern as the fuel schedule -- leaving the Pricing Rules layout
        untouched.
        """
        win = tk.Toplevel(self.root)
        win.title("Custom Surcharges")
        win.geometry("1000x580")
        self._build_dynamic_surcharge_tab(parent=win)
        try:
            self.load_dynamic_surcharge_registry()
        except Exception as e:
            print("Populating custom surcharge window failed:", e)
        win.transient(self.root)
        win.focus_set()

    def _build_dynamic_surcharge_tab(self, parent=None):

        main = ttk.Frame(parent if parent is not None else self.tab_dynamic_surcharge)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        title = ttk.Label(main, text="Custom Surcharges", font=("Segoe UI", 13, "bold"))
        title.pack(anchor="w", pady=(0, 10))

        ttk.Label(main, font=("Segoe UI", 8), foreground="#555", justify="left",
                  text="AS Code must match the invoice exactly. Surcharge Name "
                       "becomes a report column; set its rate on Pricing Rules "
                       "-> Surcharge Rates."
                  ).pack(anchor="w", pady=(0, 8))

        # Buttons are packed to the bottom FIRST so a tall list can never push
        # them off the window.
        btn_row = ttk.Frame(main)
        btn_row.pack(side="bottom", fill="x", pady=(10, 0))

        ttk.Button(btn_row, text="+ Add Surcharge", command=self.add_dynamic_surcharge_row).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Delete Selected", command=self.delete_dynamic_surcharge_row).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Save Setup", command=self.save_dynamic_surcharge_registry).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Export List", command=self.export_dynamic_surcharge_template).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Import List", command=self.import_dynamic_surcharge_template).pack(side="left", padx=4)

        tree_wrap = ttk.Frame(main)
        tree_wrap.pack(fill="both", expand=True)

        self.dynamic_surcharge_tree = ttk.Treeview(
            tree_wrap, columns=("AS", "AR", "NAME", "TYPE", "FUEL", "ENABLED"),
            show="headings", height=12)

        cols = [("AS", "AS Code", 100), ("AR", "AR", 70), ("NAME", "Surcharge Name", 240),
                ("TYPE", "Type", 130), ("FUEL", "Fuel Eligible", 110), ("ENABLED", "Enabled", 90)]

        for key, title, width in cols:
            self.dynamic_surcharge_tree.heading(key, text=title)
            self.dynamic_surcharge_tree.column(key, width=width)

        tree_scroll = ttk.Scrollbar(tree_wrap, orient="vertical",
                                    command=self.dynamic_surcharge_tree.yview)
        self.dynamic_surcharge_tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="right", fill="y")
        self.dynamic_surcharge_tree.pack(side="left", fill="both", expand=True)

    # =========================================================
    # FUEL SCHEDULE (by ship date) -- own window, own date picker
    # =========================================================
    def update_fuel_schedule_summary(self):
        """One short line on the Rules tab: how many ranges and what they cover."""
        var = getattr(self, "fuel_schedule_summary", None)
        if var is None:
            return

        if not self.fuel_schedule:
            var.set("")
            return

        var.set(f"{len(self.fuel_schedule)} range(s): "
                f"{self.fuel_schedule[0]['start']} ~ {self.fuel_schedule[-1]['end']}")

    def pick_date_into(self, target_var, parent=None):
        """Small built-in month calendar. No third-party dependency."""
        import calendar as _calendar
        from datetime import date as _date

        start = parse_schedule_date(target_var.get()) or _date.today()
        state = {"year": start.year, "month": start.month}

        win = tk.Toplevel(parent or self.root)
        win.title("Pick a date")
        win.transient(parent or self.root)
        win.resizable(False, False)
        win.grab_set()

        header = ttk.Frame(win)
        header.pack(fill="x", padx=8, pady=(8, 2))

        title_var = tk.StringVar()
        grid_frame = ttk.Frame(win)
        grid_frame.pack(padx=8, pady=8)

        def choose(day):
            target_var.set(_date(state["year"], state["month"], day).isoformat())
            win.destroy()

        def redraw():
            title_var.set(f"{_calendar.month_name[state['month']]} {state['year']}")

            for widget in grid_frame.winfo_children():
                widget.destroy()

            for col, name in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
                ttk.Label(grid_frame, text=name, width=4,
                          anchor="center").grid(row=0, column=col, padx=1, pady=1)

            for r, week in enumerate(_calendar.Calendar().monthdayscalendar(
                    state["year"], state["month"]), start=1):
                for c, day in enumerate(week):
                    if day == 0:
                        continue
                    ttk.Button(grid_frame, text=str(day), width=4,
                               command=lambda d=day: choose(d)).grid(
                        row=r, column=c, padx=1, pady=1)

        def shift(months):
            total = state["year"] * 12 + (state["month"] - 1) + months
            state["year"] = total // 12
            state["month"] = total % 12 + 1
            redraw()

        ttk.Button(header, text="<<", width=4,
                   command=lambda: shift(-12)).pack(side="left")
        ttk.Button(header, text="<", width=3,
                   command=lambda: shift(-1)).pack(side="left", padx=2)
        ttk.Label(header, textvariable=title_var, width=18,
                  anchor="center").pack(side="left", padx=6)
        ttk.Button(header, text=">", width=3,
                   command=lambda: shift(1)).pack(side="left", padx=2)
        ttk.Button(header, text=">>", width=4,
                   command=lambda: shift(12)).pack(side="left")

        redraw()

    def open_fuel_schedule_window(self):
        """Fuel schedule editor. Holds any number of ranges (scrollable)."""
        existing = getattr(self, "_fuel_window", None)
        if existing is not None:
            try:
                if existing.winfo_exists():
                    existing.lift()
                    return
            except Exception:
                pass

        win = tk.Toplevel(self.root)
        self._fuel_window = win
        win.title("Fuel Surcharge by Ship Date")
        win.geometry("820x600")
        win.transient(self.root)

        top = ttk.Frame(win)
        top.pack(fill="x", padx=12, pady=(12, 4))

        ttk.Label(top, text="Default Fuel %").pack(side="left")
        ttk.Entry(top, textvariable=self.fuel_percent,
                  width=9).pack(side="left", padx=(4, 0))

        # Set ranges. From/To spanning one week or less produces a single range,
        # so this covers both the one-off and the whole-year case.
        setter = ttk.Frame(win)
        setter.pack(fill="x", padx=12, pady=6)

        self.fuel_bulk_start_var = tk.StringVar()
        self.fuel_bulk_end_var = tk.StringVar()
        self.fuel_bulk_pct_var = tk.StringVar()

        ttk.Label(setter, text="From").pack(side="left")
        ttk.Entry(setter, textvariable=self.fuel_bulk_start_var,
                  width=13).pack(side="left", padx=(4, 2))
        ttk.Button(setter, text="\U0001F4C5", width=3,
                   command=lambda: self.pick_date_into(
                       self.fuel_bulk_start_var, win)).pack(side="left")

        ttk.Label(setter, text="To").pack(side="left", padx=(12, 0))
        ttk.Entry(setter, textvariable=self.fuel_bulk_end_var,
                  width=13).pack(side="left", padx=(4, 2))
        ttk.Button(setter, text="\U0001F4C5", width=3,
                   command=lambda: self.pick_date_into(
                       self.fuel_bulk_end_var, win)).pack(side="left")

        ttk.Label(setter, text="Fuel %").pack(side="left", padx=(12, 0))
        ttk.Entry(setter, textvariable=self.fuel_bulk_pct_var,
                  width=9).pack(side="left", padx=4)

        ttk.Button(setter, text="Add Weekly Ranges",
                   command=self.generate_weekly_fuel_rows).pack(side="left", padx=8)

        tree_wrap = ttk.Frame(win)
        tree_wrap.pack(fill="both", expand=True, padx=12, pady=6)

        self.fuel_schedule_tree = ttk.Treeview(
            tree_wrap, columns=("START", "END", "PCT"), show="headings")

        for key, title, width in [("START", "Start Date", 220),
                                  ("END", "End Date", 220),
                                  ("PCT", "Fuel %", 120)]:
            self.fuel_schedule_tree.heading(key, text=title)
            self.fuel_schedule_tree.column(key, width=width, anchor="center")

        fuel_scroll = ttk.Scrollbar(tree_wrap, orient="vertical",
                                    command=self.fuel_schedule_tree.yview)
        self.fuel_schedule_tree.configure(yscrollcommand=fuel_scroll.set)

        fuel_scroll.pack(side="right", fill="y")
        self.fuel_schedule_tree.pack(side="left", fill="both", expand=True)

        # Double-click a row to correct just that week's %.
        self.fuel_schedule_tree.bind("<Double-1>", self.edit_fuel_schedule_percent)

        actions = ttk.Frame(win)
        actions.pack(fill="x", padx=12, pady=(4, 12))

        ttk.Button(actions, text="Delete Selected",
                   command=self.delete_fuel_schedule_row).pack(side="left", padx=4)
        ttk.Button(actions, text="Clear All",
                   command=self.clear_fuel_schedule).pack(side="left", padx=4)
        ttk.Button(actions, text="Save Schedule",
                   command=self.save_fuel_schedule).pack(side="right", padx=4)

        self.refresh_fuel_schedule_tree()

    def refresh_fuel_schedule_tree(self):
        """Redraw the fuel schedule table from self.fuel_schedule."""
        tree = getattr(self, "fuel_schedule_tree", None)

        if tree is not None:
            try:
                tree.delete(*tree.get_children())

                for window in self.fuel_schedule:
                    tree.insert("", "end", values=(
                        window["start"],
                        window["end"],
                        f"{window['percent']}",
                    ))
            except Exception:
                # Window was closed; the schedule itself is unaffected.
                self.fuel_schedule_tree = None

        self.update_fuel_schedule_summary()

    def generate_weekly_fuel_rows(self):
        """Expand From/To into one 7-day range per week."""
        from datetime import timedelta as _timedelta

        start = parse_schedule_date(self.fuel_bulk_start_var.get())
        end = parse_schedule_date(self.fuel_bulk_end_var.get())

        if start is None or end is None:
            messagebox.showerror("Cannot generate",
                                 "Enter a readable From and To date first.")
            return

        if end < start:
            messagebox.showerror("Cannot generate",
                                 "To date is before From date.")
            return

        percent_text = self.fuel_bulk_pct_var.get().strip()

        generated = []
        cursor = start

        while cursor <= end:
            week_end = min(cursor + _timedelta(days=6), end)
            generated.append({
                "start": cursor.isoformat(),
                "end": week_end.isoformat(),
                "percent": percent_text,
            })
            cursor = week_end + _timedelta(days=1)

        merged, problems = normalize_fuel_schedule(
            list(self.fuel_schedule) + generated)

        added = len(merged) - len(self.fuel_schedule)

        if added <= 0:
            messagebox.showerror(
                "Nothing generated",
                "\n".join(problems or ["Check the Fuel % value."]))
            return

        self.fuel_schedule = merged
        self.refresh_fuel_schedule_tree()

        self.fuel_bulk_start_var.set("")
        self.fuel_bulk_end_var.set("")
        self.fuel_bulk_pct_var.set("")

        message = f"Added {added} weekly range(s)."
        if problems:
            message += "\n\nWarnings:\n" + "\n".join(problems)

        self.set_status(message)
    def edit_fuel_schedule_percent(self, event=None):
        """Double-click a row to correct just that week's Fuel %."""
        selected = self.fuel_schedule_tree.selection()

        if not selected:
            return

        values = self.fuel_schedule_tree.item(selected[0], "values")

        if not values:
            return

        row_start, row_end = str(values[0]), str(values[1])

        win = tk.Toplevel(self._fuel_window)
        win.title("Edit Fuel %")
        win.transient(self._fuel_window)
        win.resizable(False, False)
        win.grab_set()

        ttk.Label(win, text=f"{row_start}  ~  {row_end}").pack(padx=14, pady=(14, 4))

        pct_var = tk.StringVar(value=str(values[2]))
        entry = ttk.Entry(win, textvariable=pct_var, width=12)
        entry.pack(padx=14, pady=4)
        entry.focus_set()

        def apply_value():
            try:
                new_percent = float(pct_var.get().replace("%", "").strip())
            except (ValueError, TypeError):
                messagebox.showerror("Invalid", "Fuel % must be a number.")
                return

            for window in self.fuel_schedule:
                if window["start"] == row_start and window["end"] == row_end:
                    window["percent"] = new_percent

            win.destroy()
            self.refresh_fuel_schedule_tree()

        buttons = ttk.Frame(win)
        buttons.pack(padx=14, pady=(6, 14))
        ttk.Button(buttons, text="OK", width=10,
                   command=apply_value).pack(side="left", padx=4)
        ttk.Button(buttons, text="Cancel", width=10,
                   command=win.destroy).pack(side="left", padx=4)

        win.bind("<Return>", lambda e: apply_value())

    def delete_fuel_schedule_row(self):
        selected = self.fuel_schedule_tree.selection()

        if not selected:
            messagebox.showwarning("Delete", "Select a row first.")
            return

        drop = set()
        for item in selected:
            values = self.fuel_schedule_tree.item(item, "values")
            if values:
                drop.add((str(values[0]), str(values[1])))

        self.fuel_schedule = [
            window for window in self.fuel_schedule
            if (window["start"], window["end"]) not in drop
        ]

        self.refresh_fuel_schedule_tree()

    def clear_fuel_schedule(self):
        if not self.fuel_schedule:
            return

        if not messagebox.askyesno(
                "Clear All",
                f"Remove all {len(self.fuel_schedule)} range(s)?\n\n"
                "Every shipment then uses the global Fuel %."):
            return

        self.fuel_schedule = []
        self.refresh_fuel_schedule_tree()

    def save_fuel_schedule(self):
        self.fuel_schedule, problems = normalize_fuel_schedule(self.fuel_schedule)
        self.refresh_fuel_schedule_tree()

        try:
            self.save_config()
        except Exception as e:
            messagebox.showerror("Save failed", f"Could not save config: {e}")
            return

        if problems:
            messagebox.showwarning(
                "Fuel schedule saved with warnings", "\n".join(problems))
        else:
            self.set_status(f"Fuel schedule saved. Ranges: {len(self.fuel_schedule)}")
    def export_dynamic_surcharge_template(self):
        """Export dynamic surcharge registry as template"""
        
        path = filedialog.asksaveasfilename(
            initialfile=self.export_name_for("dynamic", "Dynamic_Surcharge_Template.xlsx")[0],
            initialdir=self.export_name_for("dynamic", "Dynamic_Surcharge_Template.xlsx")[1] or ".",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        rows = []

        for as_code, cfg in self.dynamic_surcharge_mapping.items():
            rows.append({
                "AS Code": as_code,
                "AR": cfg.get("ar", "ACC"),
                "Surcharge Name": cfg.get("name", ""),
                "Type": normalize_dynamic_type(cfg.get("type", "REPRICE")),
                "Fuel Eligible": "YES" if cfg.get("fuel_eligible", True) else "NO",
                "Enabled": "YES" if cfg.get("enabled", True) else "NO"
            })

        if not rows:
            rows = [
    {
        "AS Code": "ARK",
        "AR": "ACC",
        "Surcharge Name": "Remote Area - AK",
        "Type": "REPRICE",
        "Fuel Eligible": "NO",
        "Enabled": "YES"
    },
    {
        "AS Code": "DDO",
        "AR": "ACC",
        "Surcharge Name": "Direct Delivery Only",
        "Type": "REPRICE",
        "Fuel Eligible": "YES",
        "Enabled": "YES"
    }
]

        pd.DataFrame(rows).to_excel(path, index=False)
        self.set_status(f"Custom surcharge list exported:\n{path}")
    def import_dynamic_surcharge_template(self):
        """Import dynamic surcharge registry from template"""
        
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])

        if not path:
            return
        self.remember_import("dynamic", path)

        try:
            df = pd.read_excel(path).fillna("")

            required = ["AS Code", "AR", "Surcharge Name"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise Exception(f"Custom surcharge file missing column(s): {missing}")

            self.dynamic_surcharge_mapping = {}
            self.dynamic_surcharge_tree.delete(*self.dynamic_surcharge_tree.get_children())

            for _, row in df.iterrows():
                as_code = str(row.get("AS Code", "")).strip().upper()

                if as_code == "":
                    continue

                cfg = {
                    "ar": normalize_ar_code(row.get("AR", "ACC")),
                    "name": str(row.get("Surcharge Name", "")).strip(),
                    "type": normalize_dynamic_type(row.get("Type", "REPRICE")),
                    "fuel_eligible": str(row.get("Fuel Eligible", "YES")).strip().upper() == "YES",
                    "enabled": str(row.get("Enabled", "YES")).strip().upper() == "YES"
                }

                self.dynamic_surcharge_mapping[as_code] = cfg

                self.dynamic_surcharge_tree.insert("", "end", values=(
                    as_code,
                    cfg["ar"],
                    cfg["name"],
                    cfg["type"],
                    "YES" if cfg["fuel_eligible"] else "NO",
                    "YES" if cfg["enabled"] else "NO"
                ))

            try:
                self.render_accessorial_tables()
                self.acc_canvas.update_idletasks()
            except Exception as e:
                print("Accessorial table render failed:", e)

            try:
                self.save_config()
            except Exception as e:
                print("Auto save config failed:", e)

            self.confirm("Import complete",
                         f"Custom surcharge list imported.\n\n"
                         f"{os.path.basename(path)}\n"
                         f"Total surcharges: {len(self.dynamic_surcharge_mapping)}")
        except Exception as e:
            messagebox.showerror("Import Failed", f"Error: {e}")


    # =========================================================
    # TAB 6: SURCHARGE CODE LINKS
    # =========================================================
    def _build_surcharge_codes_tab(self):
        """
        Allow users to link multiple AS codes to existing surcharges.
        """

        main = ttk.Frame(self.tab_surcharge_codes)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        title = ttk.Label(main, text="Charge Code Mapping", font=("Segoe UI", 13, "bold"))
        title.pack(anchor="w", pady=(0, 10))

        desc = ttk.Label(main, text="Map UPS charge codes to the correct surcharge type used in the report.\nThis allows the system to automatically identify invoice charges and assign them to the correct surcharge category for repricing and audit reporting.")
        desc.pack(anchor="w", pady=(0, 12))

        self.surcharge_code_tree = ttk.Treeview(
            main,
            columns=("AS_CODE", "SURCHARGE_NAME", "DESCRIPTION"),
            show="headings",
            height=15
        )

        cols = [
            ("AS_CODE", "AS Code (UPS)", 150),
            ("SURCHARGE_NAME", "Links To System Surcharge", 350),
            ("DESCRIPTION", "Notes", 300),
        ]

        for key, title_text, width in cols:
            self.surcharge_code_tree.heading(key, text=title_text)
            self.surcharge_code_tree.column(key, width=width)

        self.surcharge_code_tree.pack(fill="both", expand=True, pady=10)

        btn_row = ttk.Frame(main)
        btn_row.pack(fill="x", pady=10)

        ttk.Button(btn_row, text="+ Add Mapping", command=self.open_surcharge_code_popup).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Delete Selected", command=self.delete_surcharge_code).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Save Mapping", command=self.save_surcharge_codes).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Load Mapping", command=self.load_surcharge_codes).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Export List", command=self.export_surcharge_code_template).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Import List", command=self.import_surcharge_code_template).pack(side="left", padx=4)

    def open_surcharge_code_popup(self):
        """Add new code → surcharge link"""
        
        popup = tk.Toplevel(self.root)
        popup.title("Add Charge Code Mapping")
        popup.geometry("600x280")

        container = ttk.Frame(popup)
        container.pack(fill="both", expand=True, padx=16, pady=16)

        ttk.Label(container, text="UPS AS Code").grid(row=0, column=0, sticky="w", pady=8)
        code_entry = ttk.Entry(container, width=35)
        code_entry.grid(row=0, column=1, padx=8, pady=8)

        ttk.Label(container, text="Maps To").grid(row=1, column=0, sticky="w", pady=8)

        surcharge_list = []
        
        for key in self.accessorial_rate_table.keys():
            if isinstance(key, str):
                surcharge_list.append(key)

        for as_code, cfg in self.dynamic_surcharge_mapping.items():
            name = cfg.get("name", "").strip()
            if name and name not in surcharge_list:
                surcharge_list.append(name)

        surcharge_list = sorted(list(set(surcharge_list)))

        surcharge_var = tk.StringVar(value="Remote Area - AK" if "Remote Area - AK" in surcharge_list else (surcharge_list[0] if surcharge_list else ""))
        surcharge_combo = ttk.Combobox(
            container,
            textvariable=surcharge_var,
            values=surcharge_list,
            state="readonly",
            width=32
        )
        surcharge_combo.grid(row=1, column=1, padx=8, pady=8)

        ttk.Label(container, text="Description / Notes (optional)").grid(row=2, column=0, sticky="w", pady=8)
        description_entry = ttk.Entry(container, width=35)
        description_entry.grid(row=2, column=1, padx=8, pady=8)

        def save_popup():
            as_code = code_entry.get().strip().upper()
            surcharge_name = surcharge_var.get().strip()
            description = description_entry.get().strip()

            if as_code == "":
                messagebox.showerror("Error", "AS Code cannot be empty")
                return

            if surcharge_name == "":
                messagebox.showerror("Error", "Surcharge name cannot be empty")
                return

            self.surcharge_code_tree.insert("", "end", values=(
                as_code,
                surcharge_name,
                description
            ))

            try:
                self.save_surcharge_codes()
            except Exception as e:
                print("Auto save surcharge codes failed:", e)

            popup.destroy()

        ttk.Button(container, text="Save Link", command=save_popup).grid(row=3, column=0, columnspan=2, pady=20)

    def delete_surcharge_code(self):
        """Delete selected code link"""
        selected = self.surcharge_code_tree.selection()
        for item in selected:
            self.surcharge_code_tree.delete(item)
        try:
            self.save_surcharge_codes()
        except Exception as e:
            print("Auto save failed:", e)

    def save_surcharge_codes(self):
        """Save all surcharge code links to memory"""
        
        self.dynamic_surcharge_code_map = {}

        for item in self.surcharge_code_tree.get_children():
            vals = self.surcharge_code_tree.item(item, "values")
            if not vals or len(vals) < 2:
                continue

            as_code = str(vals[0]).strip().upper()
            surcharge_name = str(vals[1]).strip()

            if as_code == "" or surcharge_name == "":
                continue

            self.dynamic_surcharge_code_map[as_code] = {
                "surcharge": surcharge_name,
                "description": str(vals[2]).strip() if len(vals) > 2 else ""
            }

        try:
            self.save_config()
        except Exception as e:
            print("Auto save config failed:", e)

        self.set_status(f"Charge code mapping saved.\nTotal links: {len(self.dynamic_surcharge_code_map)}")
    def load_surcharge_codes(self):
        """Load surcharge code links from memory"""
        
        self.surcharge_code_tree.delete(*self.surcharge_code_tree.get_children())

        for as_code, cfg in self.dynamic_surcharge_code_map.items():
            self.surcharge_code_tree.insert("", "end", values=(
                as_code,
                cfg.get("surcharge", ""),
                cfg.get("description", "")
            ))

    def export_surcharge_code_template(self):
        """Export surcharge code template"""
        
        path = filedialog.asksaveasfilename(
            initialfile=self.export_name_for("code_links", "Surcharge_Code_Links_Template.xlsx")[0],
            initialdir=self.export_name_for("code_links", "Surcharge_Code_Links_Template.xlsx")[1] or ".",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        rows = []

        for as_code, cfg in self.dynamic_surcharge_code_map.items():
            rows.append({
                "AS Code": as_code,
                "Surcharge Name": cfg.get("surcharge", ""),
                "Description": cfg.get("description", "")
            })

        if not rows:
            rows = [
                {"AS Code": "RAK", "Surcharge Name": "Remote Area - AK", "Description": "Alaska remote area surcharge variant 1"},
                {"AS Code": "RAS", "Surcharge Name": "Remote Area - AK", "Description": "Alaska remote area surcharge variant 2"},
                {"AS Code": "RSK", "Surcharge Name": "Remote Area - AK", "Description": "Alaska remote area surcharge variant 3"},
            ]

        pd.DataFrame(rows).to_excel(path, index=False)
        self.set_status(f"Charge code mapping exported:\n{path}")
    def import_surcharge_code_template(self):
        """Import surcharge code template"""
        
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])

        if not path:
            return
        self.remember_import("code_links", path)

        try:
            df = pd.read_excel(path).fillna("")

            self.dynamic_surcharge_code_map = {}
            self.surcharge_code_tree.delete(*self.surcharge_code_tree.get_children())

            for _, row in df.iterrows():
                as_code = str(row.get("AS Code", "")).strip().upper()

                if as_code == "":
                    continue

                surcharge_name = str(row.get("Surcharge Name", "")).strip()

                cfg = {
                    "surcharge": surcharge_name,
                    "description": str(row.get("Description", "")).strip()
                }

                self.dynamic_surcharge_code_map[as_code] = cfg

                self.surcharge_code_tree.insert("", "end", values=(
                    as_code,
                    surcharge_name,
                    cfg["description"]
                ))

            try:
                self.save_config()
            except Exception as e:
                print("Auto save config failed:", e)

            self.confirm("Import complete",
                         f"Surcharge code links imported and saved.\n\n"
                         f"{os.path.basename(path)}\n"
                         f"Total links: {len(self.dynamic_surcharge_code_map)}")
        except Exception as e:
            messagebox.showerror("Import Failed", f"Error: {e}")

    # =========================================================
    # TAB 7: RUN / PREVIEW
    # =========================================================
    def _build_run_tab(self):
        frm = ttk.Frame(self.tab_run)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkButton(frm, text="🚚 Generate Reprice Report", fg_color="#c98b47", hover_color="#b7792f", corner_radius=12, height=42, font=("Segoe UI", 13, "bold"), command=self.run_repricing).pack(anchor="w", pady=12)

        self.status_var = tk.StringVar(value="System Ready")
        ttk.Label(frm, textvariable=self.status_var).pack(anchor="w")

        cols = ["Tracking", "Zone", "Shipment Type", "Entered Weight", "Billed Weight", "Dim Weight",
                "Billable Weight", "Length", "Width", "Height", "CAL Base Rate",
                "Shipping Charge Correction Ground Undeliverable Return",
                "Residential Adjustment", "Commercial Adjustment", "Direct Delivery Only",
                "Shipping Charge Correction Large Package Surcharge",
                "CAL Fuel", "Residential",
                "DAS Commercial", "DAS Residential", "DAS Extended Commercial", "DAS Extended Residential",
                "Remote Area Commercial", "Remote Area Residential", "Remote Area - AK", "Remote Area - HI",
                "Additional Handling Surcharge - Weight",
                "Additional Handling Surcharge - Dimension",
                "Additional Handling Surcharge - Packaging", "Large Package Surcharge", "Over Maximum Size Surcharge",
                "Undeliverable Return", "Address Correction",
                "Declared Value", "Return To Sender", "2nd Mile Road Base", "Signature", "Adult Signature", "Package Protection", "CAL Total", "UPS Invoice Total"]

        self.preview = ttk.Treeview(frm, columns=cols, show="headings", height=20)

        for c in cols:
            self.preview.heading(c, text=c)
            self.preview.column(c, width=135)

        self.preview.pack(fill="both", expand=True)


    # =========================================================
    # TAB 8: CODE LOOKUP  (what does each raw-invoice code mean?)
    # =========================================================
    # Built-in reference of common UPS Charge Description Codes.
    UPS_CODE_REFERENCE = {
        "003": "Freight / Transportation Charge",
        "205": "Service Charge",
        "FSC": "Fuel Surcharge",
        "FTP": "Third Party Billing Service",
        "RES": "Residential Surcharge",
        "RDR": "Delivery Area Surcharge (Residential)",
        "RDC": "Delivery Area Surcharge (Commercial)",
        "LDR": "Delivery Area Surcharge - Extended (Residential)",
        "LDC": "Delivery Area Surcharge - Extended (Commercial)",
        "CAR": "Remote Area Surcharge",
        "AHW": "Additional Handling - Weight",
        "AHL": "Additional Handling - Longest Side",
        "AHG": "Additional Handling - Length + Girth",
        "AHC": "Additional Handling - Packaging",
        "LPS": "Large Package Surcharge - Length + Girth",
        "LPR": "Large Package Surcharge Residential - Length + Girth",
        "LCR": "Large Package Surcharge Residential - Cubic Volume",
        "COM": "Commercial Adjustment",
        "EVS": "Declared Value",
        "ERL": "Returns Electronic Label",
        "ALP": "Returns Print Label",
        "ISW": "Return To Sender - Web Request",
        "IFW": "Reschedule Delivery - Web Request",
        "IRW": "Reroute - Web Request (Address Change)",
        "ADC": "Address Correction",
        "SIG": "Signature Required",
        "ASR": "Adult Signature Required",
        "SHD": "Package Protection",
        "DDO": "Direct Delivery Only",
        "UNA": "Undeliverable Return",
    }

    # Merge the master ACC Codes table in. The ACC table is authoritative where
    # the two disagree (e.g. ISW = Intercept Service Web), so the Code Lookup
    # tab shows the same wording as the audit output.
    UPS_CODE_REFERENCE.update(UPS_ACC_CODE_REFERENCE)

    def _build_code_lookup_tab(self):
        frm = ttk.Frame(self.tab_code_lookup)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Label(
            frm,
            text="🔎 UPS Charge Code Lookup",
            font=("Segoe UI", 13, "bold")
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(
            frm,
            text="Reference of what each charge code means. "
                 "Click 'Scan Loaded Invoice' to list the codes that actually "
                 "appear in the imported raw invoice.",
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(0, 8))

        btn_row = ttk.Frame(frm)
        btn_row.pack(fill="x", pady=4)
        ttk.Button(btn_row, text="Scan Loaded Invoice",
                   command=self.scan_invoice_codes).pack(side="left", padx=4)
        ttk.Button(btn_row, text="Show Full Reference",
                   command=self.show_code_reference).pack(side="left", padx=4)

        cols = ("Code", "Surcharge / Meaning", "In Invoice", "Count")
        self.code_tree = ttk.Treeview(frm, columns=cols, show="headings",
                                      height=22)
        for c, w in zip(cols, (90, 420, 100, 90)):
            self.code_tree.heading(c, text=c)
            self.code_tree.column(
                c, width=w,
                anchor="w" if c == "Surcharge / Meaning" else "center")
        self.code_tree.pack(fill="both", expand=True, pady=6)

        self.show_code_reference()

    def show_code_reference(self):
        """Populate the table with the full built-in code reference."""
        self.code_tree.delete(*self.code_tree.get_children())
        for code in sorted(self.UPS_CODE_REFERENCE):
            meaning = self.UPS_CODE_REFERENCE[code]
            self.code_tree.insert("", "end",
                                  values=(code, meaning, "", ""))

    def scan_invoice_codes(self):
        """Read the loaded raw invoice and list which codes appear, with the
        description seen in the file and how many times."""
        path = self.billing_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Missing",
                                 "Select a billing file on Import Files first.")
            return
        try:
            if path.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str, header=None,
                                 low_memory=False).fillna("")
            else:
                df = pd.read_excel(path, dtype=str, header=None).fillna("")

            CODE_COL = 44   # Charge Description Code
            DESC_COL = 45   # Charge Description
            from collections import defaultdict, Counter
            seen = defaultdict(Counter)
            for _, r in df.iterrows():
                code = str(r.iloc[CODE_COL]).strip() \
                    if len(r) > CODE_COL else ""
                desc = str(r.iloc[DESC_COL]).strip() \
                    if len(r) > DESC_COL else ""
                if code and code.lower() not in ("nan", "charge description code"):
                    seen[code][desc] += 1

            self.code_tree.delete(*self.code_tree.get_children())
            if not seen:
                self.set_status("No charge codes found in the invoice.")
                return
            # Unknown codes sort to the TOP. Alphabetical order buried them
            # among 40+ known codes, so the one thing worth acting on was the
            # hardest thing to spot.
            known_now = set(self.UPS_CODE_REFERENCE)
            known_now.update(UPS_ACC_CODE_REFERENCE.keys())
            known_now.update(getattr(self, "dynamic_surcharge_mapping", {}) or {})
            known_now.update(getattr(self, "ahs_lps_code_registry", {}) or {})

            unknown_codes = [c for c in sorted(seen) if c not in known_now]
            known_codes = [c for c in sorted(seen) if c in known_now]

            for code in unknown_codes + known_codes:
                total = sum(seen[code].values())
                invoice_desc = seen[code].most_common(1)[0][0]
                if code in unknown_codes:
                    meaning = f"⚠ UNKNOWN -- {invoice_desc or 'no description'}"
                else:
                    meaning = self.UPS_CODE_REFERENCE.get(
                        code, invoice_desc or "(unknown)")
                self.code_tree.insert(
                    "", "end",
                    values=(code, meaning, "Yes", total))

            if unknown_codes:
                self.status_var.set(
                    f"Scanned invoice: {len(seen)} codes, "
                    f"{len(unknown_codes)} UNKNOWN -> {', '.join(unknown_codes[:10])}")
                messagebox.showwarning(
                    "Unknown codes found",
                    f"{len(unknown_codes)} charge code(s) in this invoice are "
                    f"not registered:\n\n  "
                    + "\n  ".join(
                        f"{c}  ({sum(seen[c].values())} rows)  "
                        f"{seen[c].most_common(1)[0][0][:40]}"
                        for c in unknown_codes[:15])
                    + "\n\nThey will NOT be priced until you add them under\n"
                      "Pricing Rules -> Custom Surcharges.")
            else:
                self.status_var.set(
                    f"Scanned invoice: {len(seen)} distinct charge codes, "
                    f"all recognised.")
        except Exception as e:
            messagebox.showerror("Scan Error", str(e))


    # =========================================================
    # FILE PICKERS
    # =========================================================
    def _pick_billing(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])
        if path:
            self.billing_path.set(path)

    def _pick_config(self):
        path = filedialog.asksaveasfilename(initialfile=DEFAULT_CONFIG_PATH, defaultextension=".json", filetypes=[("JSON", "*.json")])
        if path:
            self.config_path.set(path)

    # =========================================================
    # TEMPLATE FUNCTIONS
    # =========================================================
    def export_base_rate_template(self):
        service = self.base_rate_service.get()
        path = filedialog.asksaveasfilename(initialfile=f"{service}_Base_Rate_template.xlsx", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        rows = [{"Service": service, "Weight": weight} for weight in range(1, 151)]
        for z in [2, 3, 4, 5, 6, 7, 8, 44, 45, 46]:
            for row in rows:
                row[f"Zone {z}"] = 0

        pd.DataFrame(rows).to_excel(path, index=False)
        self.set_status(f"{service} Base Rate template exported:\n{path}")
    def import_base_rate_template(self):
        # Allow selecting multiple base rate files at once (e.g. Commercial +
        # Residential). Tables are MERGED, not overwritten, so reclassification
        # (RES <-> COM) can look up both services.
        paths = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx *.xls")])
        if not paths:
            return
        self._load_base_rate_files(paths, quiet=False)

    def _load_base_rate_files(self, paths, quiet=False):
        """Shared by the Import button and the startup reload. Rate tables are
        DataFrames, too big for the config file, so the config remembers the
        FILE PATHS and reloads them -- otherwise every session starts with no
        rates at all."""
        try:
            if not hasattr(self, "rate_tables") or not isinstance(self.rate_tables, dict):
                self.rate_tables = {}

            all_dfs = []
            imported_services = []
            for path in paths:
                df = pd.read_excel(path).fillna("")

                required = ["Service", "Weight"]
                missing = [c for c in required if c not in df.columns]
                if missing:
                    raise Exception(
                        f"Base rate file missing column(s): {missing}\nFile: {path}")

                # Merge: each service name becomes/updates a table entry.
                for service_name, sub_df in df.groupby("Service"):
                    key = str(service_name).strip()
                    self.rate_tables[key] = sub_df.copy()
                    imported_services.append(key)

                all_dfs.append(df)

            # Keep a COMBINED frame for preview/round-trip save. Accumulate
            # across separate imports too (so importing Commercial then
            # Residential shows BOTH, not just the last one). De-duplicate by
            # Service so re-importing the same service refreshes rather than
            # stacking duplicate rows.
            new_df = pd.concat(all_dfs, ignore_index=True)
            if isinstance(getattr(self, "last_base_rate_df", None), pd.DataFrame) \
                    and not self.last_base_rate_df.empty:
                newly = set(new_df["Service"].astype(str).str.strip())
                kept = self.last_base_rate_df[
                    ~self.last_base_rate_df["Service"].astype(str).str.strip().isin(newly)
                ]
                self.last_base_rate_df = pd.concat([kept, new_df], ignore_index=True)
            else:
                self.last_base_rate_df = new_df

            self._populate_rate_preview(self.last_base_rate_df.head(500))

            cached = ", ".join(sorted(set(self.rate_tables.keys())))
            self.base_rate_paths = [str(p) for p in paths]
            _msg = (f"Base rates imported and merged.\n\n"
                    f"Newly loaded: {', '.join(sorted(set(imported_services)))}\n"
                    f"All cached services: {cached}")
            if quiet:
                self.set_status(_msg.replace("\n", "  "))
            else:
                self.save_config()
                self.confirm("Import complete", _msg)

        except Exception as e:
            if quiet:
                self.set_status(f"Base rate reload failed: {e}")
            else:
                messagebox.showerror("Base rate import failed",
                                     f"Base rate import failed\n\nError: {e}")

    def export_billing_sample(self):
        path = filedialog.asksaveasfilename(initialfile="UPS_Billing_Sample.xlsx", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        df = pd.DataFrame([{"Tracking Number": "1ZTEST0001", "Zone": 4, "Entered Weight": 10, "Billed Weight": 15, "Package Dimensions": "20 x 12 x 10", "Charge Description Code": "FRT", "Charge Description": "Ground", "Net Amount": 0}])
        df.to_excel(path, index=False)
        self.set_status(f"Billing sample exported:\n{path}")
    # =========================================================
    # CONFIG (Thread-Safe)
    # =========================================================
    def open_url(self, url):
        try:
            webbrowser.open(url)
            self.set_status(f"Opened {url}")
        except Exception as e:
            self.set_status(f"Could not open browser: {e}")

    def remember_import(self, kind, path):
        """Round-trip: export defaults to the file the same registry was last
        imported from, instead of a fixed template name that overwrites the
        previous export every time."""
        if not hasattr(self, "last_import_paths"):
            self.last_import_paths = {}
        if path:
            self.last_import_paths[kind] = str(path)

    def export_name_for(self, kind, fallback):
        p = getattr(self, "last_import_paths", {}).get(kind, "")
        if not p:
            return fallback, ""
        raw = str(p).replace("\\", "/")
        name = os.path.basename(raw)
        return (name or fallback), (os.path.dirname(raw) or "")

    def confirm(self, title, text):
        """Imports DO get a dialog. Everything else goes to the status line.
        An import that changes the rate card silently is indistinguishable
        from one that failed, so this is the one place a modal earns its
        interruption."""
        self.set_status(str(text).replace("\n", "  "))
        try:
            messagebox.showinfo(title, text)
        except Exception:
            pass

    # =========================================================
    # PROFIT REPORT
    # =========================================================
    def run_profit_report(self):
        """A standalone workbook: what was billed to the client against what
        UPS billed in, per charge line, with each line's share.

        It reads a repriced report rather than recomputing, so it can be run
        on any past report -- including after the app has been reopened, when
        nothing is left in memory."""
        # Both files are usually already known: the invoice is the one on the
        # Import Files tab, and the report is the one just generated. Asking
        # for them again turned one click into two file hunts.
        def _known(p):
            p = str(p or "").strip()
            return p if p and os.path.exists(p) else ""

        src_path = _known(getattr(self, "last_report_path", ""))
        inv_path = _known(self.billing_path.get())

        # Nothing remembered? The report is almost always sitting next to the
        # invoice with "_Repriced" in the name -- take the newest one rather
        # than opening a file dialog for something already on disk.
        if not src_path and inv_path:
            try:
                _dir = os.path.dirname(inv_path) or "."
                _cands = [os.path.join(_dir, f) for f in os.listdir(_dir)
                          if "_repriced" in f.lower()
                          and f.lower().endswith((".xlsx", ".xlsm"))
                          and not f.lower().endswith("_statement.xlsx")
                          and not f.startswith("~$")]
                if _cands:
                    src_path = max(_cands, key=os.path.getmtime)
            except OSError:
                pass
        reused = []
        if src_path:
            reused.append(os.path.basename(src_path))
        else:
            src_path = filedialog.askopenfilename(
                title="The repriced report you sent the client",
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"),
                           ("All", "*.*")])
            if not src_path:
                return
        if inv_path:
            reused.append(os.path.basename(inv_path))
        else:
            inv_path = filedialog.askopenfilename(
                title="The UPS invoice it was priced from",
                initialdir=os.path.dirname(str(src_path).replace("\\", "/")) or ".",
                filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.csv"),
                           ("All", "*.*")])
            if not inv_path:
                return
        if reused:
            self.set_status("Using " + "  +  ".join(reused))
        try:
            self._build_profit_report(src_path, inv_path)
        except ValueError as e:
            messagebox.showwarning("Cannot build profit report", str(e))
        except Exception as e:
            messagebox.showerror("Cannot build profit report",
                                 f"{type(e).__name__}: {e}")

    # Fixed positions, used when the file has no header row.
    BILLING_COLS = {"Tracking Number": 20, "Net Amount": 52,
                    "Invoice Number": 5, "Invoice Date": 4,
                    "Lead Shipment Number": 13,
                    "Charge Classification Code": 43,
                    "Charge Description Code": 44,
                    "Charge Description": 45}

    @classmethod
    def read_billing_file(cls, path):
        """Returns (frame, has_header).

        UPS ships these as .csv as often as .xlsx, AND the CSV often has no
        header row at all. Reading a headerless file with header=0 eats the
        first shipment line -- it becomes the column names -- so one parcel
        silently loses a charge. That is exactly how a $16.84 freight line
        went missing from a tracking's cost."""
        p = str(path).lower()
        read = ((lambda h: pd.read_csv(path, header=h, dtype=object,
                                       low_memory=False))
                if p.endswith((".csv", ".txt"))
                else (lambda h: pd.read_excel(path, header=h, dtype=object)))

        df = read(0)
        names = {str(c).strip().lower() for c in df.columns}
        if "tracking number" in names and "net amount" in names:
            return df, True
        # No header: re-read keeping every row, address by position.
        return read(None), False

    def ups_cost_by_tracking(self, invoice_path):
        """Net amount per tracking straight from the UPS invoice, summed over
        every layer (SHP + RTN + ADJ) because a correction layer is real
        money too. Also returns the account-level rows, which carry no
        tracking number and therefore no revenue."""
        raw, has_header = self.read_billing_file(invoice_path)

        # By header name when the file has one, by fixed position when it
        # does not. Position alone breaks if a column is added or dropped;
        # names alone break on the headerless exports UPS also produces.
        def col(name, required=True):
            if has_header:
                lookup = {str(c).strip().lower(): c for c in raw.columns}
                if name.lower() in lookup:
                    return raw[lookup[name.lower()]]
            else:
                idx = self.BILLING_COLS.get(name)
                if idx is not None and idx < raw.shape[1]:
                    return raw.iloc[:, idx]
            if required:
                raise ValueError(
                    f"That file has no \"{name}\" column.\n\n"
                    f"It has {raw.shape[1]} columns"
                    + (" with headers." if has_header else " and no header row.")
                    + "\nIt does not look like a UPS billing file.")
            return pd.Series([""] * len(raw), index=raw.index)

        trk = col("Tracking Number")
        net = pd.to_numeric(col("Net Amount"), errors="coerce").fillna(0)
        inv_no = col("Invoice Number").astype(str).str.strip()
        inv_dt = col("Invoice Date")
        lead = col("Lead Shipment Number", required=False)
        cls_c = col("Charge Classification Code", required=False)
        code_c = col("Charge Description Code", required=False)
        desc_c = col("Charge Description", required=False)

        # A missing tracking number can arrive as NaN, pd.NA, "nan" or
        # "None" depending on the pandas string dtype in play. Left as NA it
        # survives the mask AND is then silently dropped by groupby, so its
        # cost disappears and the reported profit is overstated. Normalise to
        # "" first and group with dropna=False.
        key = trk.astype(str).str.strip()
        key = key.where(key.notna(), "")
        key = key.mask(key.str.lower().isin(["nan", "none", "<na>", "nat"]), "")

        df = pd.DataFrame({"Tracking": key, "Net": net,
                           "InvNo": inv_no, "InvDate": inv_dt})
        has = df["Tracking"].ne("")
        by = df[has].groupby("Tracking", as_index=True, dropna=False).agg(
            Net=("Net", "sum"), InvNo=("InvNo", "first"),
            InvDate=("InvDate", "first"))
        account_total = float(df[~has]["Net"].sum())

        # Account-level rows, itemised. They carry no tracking number and
        # therefore no revenue, so they are pure cost on the statement.
        _m = ~has.values
        account_rows = pd.DataFrame({
            "Tracking": lead[_m].astype(str).replace(
                {"nan": "", "None": "", "<NA>": ""}),
            "Charge Description": desc_c[_m].astype(str),
            "Charge Description Code": code_c[_m].astype(str),
            "Charge Classification": cls_c[_m].astype(str),
            "Amount": net[_m].round(2),
        }) if bool(_m.any()) else pd.DataFrame(
            columns=["Tracking", "Charge Description",
                     "Charge Description Code", "Charge Classification",
                     "Amount"])

        # Charge codes this build has no rate for. Same membership test the
        # repricing run uses, so the two agree.
        known = set(UPS_ACC_CODE_REFERENCE.keys())
        known.update({
            "FRT", "FSC", "RES", "RDC", "RDR", "LDC", "LDR", "CAC", "CAR",
            "AHG", "AHL", "AHS", "AHW", "AHD", "AHB", "AHP", "AHC",
            "LPS", "LPR", "LCT", "LCR", "OVR", "EVS", "ISW", "SIG", "ASG",
            "SHD", "DDO", "IRW", "IFW", "ERL", "ALP", "FTP",
        })
        known.update(getattr(self, "dynamic_surcharge_mapping", {}).keys())
        known.update(getattr(self, "ahs_lps_code_registry", {}).keys())
        known.update(getattr(self, "dynamic_surcharge_code_map", {}).keys())

        ar = cls_c.astype(str).str.strip().str.upper()
        as_c = code_c.astype(str).str.strip().str.upper()
        is_unknown = ar.isin(BILLABLE_AR_CODES) & ~as_c.isin(known)
        unknown_rows = pd.DataFrame({
            "Tracking": key[is_unknown.values],
            "Charge Description": desc_c[is_unknown.values].astype(str),
            "AS Charge Code": as_c[is_unknown.values],
            "AR Classification": ar[is_unknown.values],
            "Net Amount": net[is_unknown.values].round(2),
        }) if bool(is_unknown.any()) else pd.DataFrame(
            columns=["Tracking", "Charge Description", "AS Charge Code",
                     "AR Classification", "Net Amount"])

        # The two buckets must add back to the invoice, to the penny.
        _check = float(net.sum()) - float(by["Net"].sum()) - account_total
        if abs(_check) > 0.01:
            raise ValueError(
                f"Cost does not reconcile to the invoice: {_check:,.2f} "
                "unaccounted for. Refusing to report a profit from it.")
        return by, account_total, account_rows, unknown_rows

    def _build_profit_report(self, src_path, invoice_path):
        """Emits the weekly statement layout already in use:

            <MMDDYY>   Invoice Number | Invoice Date | Tracking |
                       Billed amount | UPS | Profit | Margin
            Overview   Invoice Number | Invoice Date | UPS |
                       Charge Customer | Profit | Margin

        One sheet per invoice number found, plus the Overview across them.
        Margin is a fraction so Excel's % format shows it directly."""
        if str(src_path).lower().endswith((".csv", ".txt")):
            book = {"Shipment Detail": pd.read_csv(src_path, low_memory=False)}
        else:
            book = pd.read_excel(src_path, sheet_name=None)
        sheet = next((n for n in book if n.strip().lower() == "shipment detail"),
                     None)
        if sheet is None:
            raise ValueError(
                "No \"Shipment Detail\" sheet in that file.\n\n"
                "Run Generate Reprice Report first, then point this at the "
                "report it produced.")
        fin = book[sheet].copy()
        for need in ("Total", "Tracking"):
            if need not in fin.columns:
                raise ValueError(f"That sheet has no \"{need}\" column.")

        (cost_by, account_total, account_rows,
         unknown_rows) = self.ups_cost_by_tracking(invoice_path)

        fin["Tracking"] = fin["Tracking"].astype(str).str.strip()
        fin["_sell"] = pd.to_numeric(fin["Total"], errors="coerce").fillna(0)
        fin["_cost"] = fin["Tracking"].map(cost_by["Net"]).fillna(0.0)
        # Coverage. A tracking billed to the client but absent from the
        # invoice silently costs 0 and inflates the profit; one on the
        # invoice but not on the report is revenue never charged. Both are
        # money, so both are reported instead of quietly dropped.
        _rep_keys = set(fin["Tracking"])
        _inv_keys = set(cost_by.index)
        _no_cost = sorted(_rep_keys - _inv_keys)
        _not_billed = sorted(_inv_keys - _rep_keys)
        _no_cost_amt = float(fin[fin["Tracking"].isin(_no_cost)]["_sell"].sum())
        _not_billed_amt = float(cost_by.loc[_not_billed, "Net"].sum()) \
            if _not_billed else 0.0

        matched = int(fin["Tracking"].isin(cost_by.index).sum())
        if matched == 0:
            raise ValueError(
                "No tracking number in that report appears in that invoice.\n\n"
                "Check you picked the invoice this report was priced from.")

        # Always from the invoice. The client report deliberately no longer
        # carries them, and the invoice is the authoritative source anyway.
        fin["Invoice Number"] = fin["Tracking"].map(cost_by["InvNo"]).fillna("")
        fin["Invoice Date"] = fin["Tracking"].map(cost_by["InvDate"])

        def sheet_name_for(inv_date, inv_no, used):
            """MMDDYY, the same convention as the existing tabs."""
            base = ""
            try:
                d = pd.to_datetime(inv_date, errors="coerce")
                if not pd.isna(d):
                    base = d.strftime("%m%d%y")
            except Exception:
                pass
            if not base:
                base = str(inv_no)[-6:] or "Detail"
            name = base
            n = 2
            while name in used:
                name = f"{base}_{n}"
                n += 1
            used.add(name)
            return name

        # A tracking with no invoice number is not on the invoice at all.
        # It already shows up under Unmatched; letting it form its own group
        # produced a nameless extra sheet and a row of NaN on the Overview.
        _inv_key = fin["Invoice Number"].astype(str).str.strip()
        _named = fin[_inv_key.ne("") & _inv_key.str.lower().ne("nan")]

        groups, used = [], set()
        for inv_no, g in _named.groupby(
                _named["Invoice Number"].astype(str).str.strip(), sort=True):
            inv_date = g["Invoice Date"].iloc[0] if len(g) else ""
            sell = float(g["_sell"].sum())
            cost = float(g["_cost"].sum())
            detail = pd.DataFrame({
                "Invoice Number": g["Invoice Number"],
                "Invoice Date": g["Invoice Date"],
                "Tracking": g.get("Tracking", ""),
                "Billed amount": g["_sell"].round(2),
                "UPS": g["_cost"].round(2),
                "Profit": (g["_sell"] - g["_cost"]).round(2),
                "Margin": ((g["_sell"] - g["_cost"]) /
                           g["_sell"].where(g["_sell"] != 0)).fillna(0).round(6),
            })
            groups.append({
                "sheet": sheet_name_for(inv_date, inv_no, used),
                "detail": detail,
                "overview": {
                    "Invoice Number": inv_no,
                    "Invoice Date": inv_date,
                    "UPS": round(cost, 2),
                    "Charge Customer": round(sell, 2),
                    "Profit": round(sell - cost, 2),
                    "Margin": round((sell - cost) / sell, 6) if sell else 0.0,
                }})

        overview = pd.DataFrame([g["overview"] for g in groups])
        t_sell = float(fin["_sell"].sum())
        t_cost = float(fin["_cost"].sum())
        # One row per invoice, nothing else. Account charges and unmatched
        # trackings live on their own sheet.
        t_cost += account_total

        base = os.path.splitext(os.path.basename(
            str(src_path).replace("\\", "/")))[0]
        _suggest = f"{base}_Statement.xlsx"

        out = filedialog.asksaveasfilename(
            title="Save weekly statement",
            initialfile=_suggest,
            initialdir=os.path.dirname(str(src_path).replace("\\", "/")) or ".",
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not out:
            return

        det_cols = ["Invoice Number", "Invoice Date", "Tracking",
                    "Billed amount", "UPS", "Profit", "Margin"]
        ov_cols = ["Invoice Number", "Invoice Date", "UPS",
                   "Charge Customer", "Profit", "Margin"]
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            for g in groups:
                g["detail"][det_cols].to_excel(w, sheet_name=g["sheet"],
                                               index=False)
            overview[ov_cols].to_excel(w, sheet_name="Overview", index=False)
            # Account charges and unmatched trackings are the same thing from
            # the operator's point of view: money on one side with nothing
            # facing it on the other. One sheet, one column saying which.
            _other = []
            for _, _r in account_rows.iterrows():
                _other.append({
                    "Tracking": _r["Tracking"],
                    "Type": "account charge (no tracking)",
                    "Charge Description": _r["Charge Description"],
                    "Code": _r["Charge Description Code"],
                    "Amount": _r["Amount"]})
            for t in _no_cost:
                _other.append({
                    "Tracking": t, "Type": "billed to client, not on invoice",
                    "Charge Description": "", "Code": "",
                    "Amount": round(float(
                        fin[fin["Tracking"] == t]["_sell"].sum()), 2)})
            for t in _not_billed:
                _other.append({
                    "Tracking": t, "Type": "on invoice, not billed to client",
                    "Charge Description": "", "Code": "",
                    "Amount": round(float(cost_by.loc[t, "Net"]), 2)})
            if _other:
                pd.DataFrame(_other)[["Tracking", "Type", "Charge Description",
                                      "Code", "Amount"]].to_excel(
                    w, sheet_name="Account Charges", index=False)
            if not unknown_rows.empty:
                unknown_rows.to_excel(w, sheet_name="Unknown Codes",
                                      index=False)

        # Margin as a real percentage, money to 2dp
        try:
            from openpyxl import load_workbook as _lw
            wbk = _lw(out)
            for ws in wbk.worksheets:
                _hrow = 1
                heads = [c.value for c in ws[_hrow]]
                for idx, h in enumerate(heads, start=1):
                    if h == "Margin":
                        fmt = "0.00%"
                    elif h in ("Billed amount", "UPS", "Profit",
                               "Charge Customer", "Amount", "Net Amount"):
                        fmt = "#,##0.00"
                    elif h == "Invoice Date":
                        fmt = "mm/dd/yyyy"
                    else:
                        continue
                    for row in ws.iter_rows(min_row=_hrow + 1, min_col=idx,
                                            max_col=idx):
                        row[0].number_format = fmt
                ws.freeze_panes = f"A{_hrow + 1}"
            wbk.save(out)
        except Exception as e:
            print("Statement formatting skipped:", e)

        self.confirm(
            "Weekly statement",
            f"Statement saved.\n\n{os.path.basename(out)}\n\n"
            f"{len(groups)} invoice sheet(s) + Overview\n\n"
            f"Charge Customer  {t_sell:,.2f}\n"
            f"UPS              {t_cost:,.2f}\n"
            + (f"Unmatched        {len(_no_cost)} billed / "
               f"{len(_not_billed)} on invoice "
               f"({_no_cost_amt + _not_billed_amt:,.2f}) "
               f"-- see Account Charges\n"
               if (_no_cost or _not_billed) else "")
            + f"Profit           {t_sell - t_cost:,.2f}"
            f"  ({(t_sell - t_cost) / t_sell * 100 if t_sell else 0:.2f}%)")

    def set_status(self, text):
        """One place to say what happened. Replaces the modal pop-ups that
        interrupted every import and every report."""
        try:
            self.status_var.set(str(text))
        except Exception:
            pass
        print(text)

    def save_config(self, silent=False):
        """Thread-safe config save"""
        with _config_lock:
            serializable_table = {}

            for key, value in self.accessorial_rate_table.items():
                if isinstance(key, tuple):
                    serializable_key = "|||".join([str(x) for x in key])
                else:
                    serializable_key = str(key)
                serializable_table[serializable_key] = value

            cfg = {
                "global_rules": {
                    "dim_factor": self.dim_factor.get(),
                    "fuel_percent": self.fuel_percent.get(),
                    "use_oversize_longest": self.use_oversize_longest.get(),
                    "use_oversize_weight": self.use_oversize_weight.get(),
                    "use_oversize_cubic": self.use_oversize_cubic.get(),
                    "use_oversize_lg": self.use_oversize_lg.get(),
                    "oversize_longest_side": self.oversize_longest_side.get(),
                    "oversize_actual_weight": self.oversize_actual_weight.get(),
                    "oversize_cubic_inches": self.oversize_cubic_inches.get(),
                    "oversize_length_girth": self.oversize_length_girth.get(),
                    "oversize_min_billable_weight": self.oversize_min_billable_weight.get(),
                    "ahs_weight_threshold": self.ahs_weight_threshold.get(),
                    "ahs_longest_side": self.ahs_longest_side.get(),
                    "ahs_second_side": self.ahs_second_side.get(),
                    "ahs_cubic_inches": self.ahs_cubic_inches.get(),
                    "ahs_lg_limit": self.ahs_lg_limit.get(),
                    "ahs_min_billable_weight": self.ahs_min_billable_weight.get(),
                },
                # These were never written out, so the service dropdown and the
                # OVR fee reset every session, and the base rate files had to
                # be re-imported by hand each time.
                # Remembered so a reopened session does not make you hunt for
                # the same two files again.
                "billing_path": self.billing_path.get(),
                "last_report_path": getattr(self, "last_report_path", ""),
                "base_rate_service": self.base_rate_service.get(),
                "base_rate_paths": list(getattr(self, "base_rate_paths", [])),
                "accessorial_rate_table": serializable_table,
                "fuel_schedule": self.fuel_schedule,
                "dynamic_surcharge_mapping": self.dynamic_surcharge_mapping,
                "dynamic_surcharge_code_map": self.dynamic_surcharge_code_map,
                "ahs_lps_code_registry": self.ahs_lps_code_registry,
                "accessorial_charges": {
                    "residential_fee": self.residential_fee.get(),
                    "das_com_fee": self.das_com_fee.get(),
                    "das_res_fee": self.das_res_fee.get(),
                    "das_ext_com_fee": self.das_ext_com_fee.get(),
                    "das_ext_res_fee": self.das_ext_res_fee.get(),
                    "remote_com_fee": self.remote_com_fee.get(),
                    "remote_res_fee": self.remote_res_fee.get(),
                    "ahs_weight_fee": self.ahs_weight_fee.get(),
                    "ahs_dimension_fee": self.ahs_dimension_fee.get(),
                    "ahs_packaging_fee": self.ahs_packaging_fee.get(),
                    "oversize_fee": self.oversize_fee.get(),
                    "ovr_fee": self.ovr_fee.get(),
                    "signature_fee": self.signature_fee.get(),
                    "package_protection_fee": self.package_protection_fee.get(),
                    "declared_value_markup": self.declared_value_markup.get(),
                }
            }

            path = self.config_path.get().strip() or default_config_path()
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(cfg, f, ensure_ascii=False, indent=2)
            except (IOError, OSError) as e:
                if silent:
                    print("Auto-save failed:", e)
                    return
                messagebox.showerror("Save Failed", f"Unable to save settings:\n\n{e}")
                raise

            self.set_status("Settings saved.")
    def load_config(self, silent=False):
        """Thread-safe config load"""
        with _config_lock:
            path = self.config_path.get().strip()
            if not path or not os.path.exists(path):
                if silent:
                    return
                messagebox.showerror("Error", "Settings file not found.")
                return

            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
            except (IOError, OSError, json.JSONDecodeError) as e:
                if silent:
                    self.set_status(f"Auto-load skipped: {e}")
                    return
                messagebox.showerror("Load Failed", f"Unable to load settings:\n{e}")
                return

            rules = cfg.get("global_rules", {})
            accessorials = cfg.get("accessorial_charges", {})
            raw_accessorial_table = cfg.get("accessorial_rate_table", {})

            restored_accessorial_table = {}

            for key, value in raw_accessorial_table.items():
                if "|||" in str(key):
                    restored_key = tuple(str(key).split("|||"))
                else:
                    restored_key = key
                restored_accessorial_table[restored_key] = value

            self.accessorial_rate_table = restored_accessorial_table

            # Saved config WINS for any code it defines, but the built-in ACC
            # defaults are re-seeded for codes the config never heard of.
            # Without this merge an older config file silently removes every
            # newly added ACC code again.
            merged_dynamic = {
                code: dict(cfg_default)
                for code, cfg_default in DEFAULT_DYNAMIC_SURCHARGE_MAPPING.items()
            }
            merged_dynamic.update(cfg.get("dynamic_surcharge_mapping", {}) or {})
            self.dynamic_surcharge_mapping = merged_dynamic

            self.dynamic_surcharge_code_map = cfg.get("dynamic_surcharge_code_map", {})

            restored_fuel_schedule, fuel_problems = normalize_fuel_schedule(
                cfg.get("fuel_schedule", []) or []
            )
            self.fuel_schedule = restored_fuel_schedule
            if fuel_problems:
                print("Fuel schedule warnings on load:", fuel_problems)

            # load_config is triggered by the "Load Settings" button, i.e. AFTER
            # the Rules tab was built, so the table has to be redrawn here or the
            # restored ranges stay invisible.
            try:
                self.refresh_fuel_schedule_tree()
            except Exception as e:
                print("Fuel schedule table refresh failed:", e)

            saved_registry = cfg.get("ahs_lps_code_registry", {}) or {}
            merged_registry = {
                code: dict(reg_cfg)
                for code, reg_cfg in self.ahs_lps_code_registry.items()
            }
            merged_registry.update(saved_registry)
            self.ahs_lps_code_registry = merged_registry

            for key, var in {
                "dim_factor": self.dim_factor,
                "fuel_percent": self.fuel_percent,
                "oversize_longest_side": self.oversize_longest_side,
                "oversize_actual_weight": self.oversize_actual_weight,
                "oversize_cubic_inches": self.oversize_cubic_inches,
                "oversize_length_girth": self.oversize_length_girth,
                "oversize_min_billable_weight": self.oversize_min_billable_weight,
                "ahs_weight_threshold": self.ahs_weight_threshold,
                "ahs_longest_side": self.ahs_longest_side,
                "ahs_second_side": self.ahs_second_side,
                "ahs_cubic_inches": self.ahs_cubic_inches,
                "ahs_lg_limit": self.ahs_lg_limit,
                "ahs_min_billable_weight": self.ahs_min_billable_weight,
            }.items():
                if key in rules:
                    var.set(rules[key])

            for key, var in {
                "residential_fee": self.residential_fee,
                "das_com_fee": self.das_com_fee,
                "das_res_fee": self.das_res_fee,
                "das_ext_com_fee": self.das_ext_com_fee,
                "das_ext_res_fee": self.das_ext_res_fee,
                "remote_com_fee": self.remote_com_fee,
                "remote_res_fee": self.remote_res_fee,
                "ahs_weight_fee": self.ahs_weight_fee,
                "ahs_dimension_fee": self.ahs_dimension_fee,
                "ahs_packaging_fee": self.ahs_packaging_fee,
                "oversize_fee": self.oversize_fee,
                "signature_fee": self.signature_fee,
                "package_protection_fee": self.package_protection_fee,
                "declared_value_markup": self.declared_value_markup,
            }.items():
                if key in accessorials:
                    var.set(accessorials[key])

            self.render_accessorial_tables()

            try:
                if hasattr(self, "dynamic_surcharge_tree"):
                    self.load_dynamic_surcharge_registry()
            except Exception as e:
                print("Dynamic surcharge UI rebuild failed:", e)

            try:
                if hasattr(self, "ahs_lps_code_tree"):
                    self.load_ahs_lps_registry()
            except Exception as e:
                print("AHS/LPS code UI rebuild failed:", e)

            try:
                if hasattr(self, "surcharge_code_tree"):
                    self.load_surcharge_codes()
            except Exception as e:
                print("Surcharge codes UI rebuild failed:", e)

            try:
                self.acc_canvas.update_idletasks()
            except Exception:
                pass


            # Only restore paths that still exist -- a stale one would send
            # the file dialog to a folder that has moved.
            _bp = str(cfg.get("billing_path") or "")
            if _bp and os.path.exists(_bp):
                self.billing_path.set(_bp)
            _lr = str(cfg.get("last_report_path") or "")
            self.last_report_path = _lr if _lr and os.path.exists(_lr) else ""

            _svc = cfg.get("base_rate_service")
            if _svc:
                self.base_rate_service.set(_svc)
            _ovr = (cfg.get("accessorial_charges") or {}).get("ovr_fee")
            if _ovr is not None:
                self.ovr_fee.set(str(_ovr))

            # Reload the base rate workbooks this config was saved with, so a
            # session starts ready to rate instead of with an empty table.
            _paths = [p for p in (cfg.get("base_rate_paths") or [])
                      if p and os.path.exists(p)]
            self.base_rate_paths = _paths
            if _paths:
                try:
                    self._load_base_rate_files(_paths, quiet=True)
                except Exception as e:
                    self.set_status(f"Base rate reload failed: {e}")

            self.set_status("Settings loaded."
                            + (f"  Base rates reloaded from {len(_paths)} file(s)."
                               if _paths else ""))


    # =========================================================
    # EXPORT ACCESSORIAL CHARGES TEMPLATE (v2.4 - CLEANED)
    # =========================================================
    def export_acc_template(self):

        from openpyxl import Workbook
        from openpyxl.styles import Font

        path = filedialog.asksaveasfilename(
            initialfile=self.export_name_for("acc", "UPS_ACC_TEMPLATE_V4.xlsx")[0],
            initialdir=self.export_name_for("acc", "UPS_ACC_TEMPLATE_V4.xlsx")[1] or ".",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            return

        wb = None
        try:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

            ws1 = wb.create_sheet("AHS_LPS")

            headers1 = ["Fee Type", "Shipment Type", "Zone 2", "Zone 3", "Zone 4", "Zone 5", "Zone 6", "Zone 7", "Zone 8", "Zone 44", "Zone 45", "Zone 46"]

            for c, h in enumerate(headers1, start=1):
                cell = ws1.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)

            rows1 = [
                ["AHS Weight", "Ground Commercial"],
                ["AHS Weight", "Ground Residential"],
                ["AHS Dimension", "Ground Commercial"],
                ["AHS Dimension", "Ground Residential"],
                ["AHS Packaging", "Ground Commercial"],
                ["AHS Packaging", "Ground Residential"],
                ["Large Package", "Ground Commercial"],
                ["Large Package", "Ground Residential"],
            ]

            for r, row in enumerate(rows1, start=2):
                for c, val in enumerate(row, start=1):
                    ws1.cell(row=r, column=c, value=val)

            ws3 = wb.create_sheet("FLAT_CHARGES")

            headers3 = ["Fee Type", "Amount"]

            for c, h in enumerate(headers3, start=1):
                cell = ws3.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)

            rows3 = [
                ["DAS Commercial", 0],
                ["DAS Residential", 0],
                ["DAS Extended Commercial", 0],
                ["DAS Extended Residential", 0],
                ["Remote Area Commercial", 0],
                ["Remote Area Residential", 0],
                ["Remote Area - AK", 0],
                ["Remote Area - HI", 0],
                ["Residential Surcharge", 0],
                ["Return To Sender", 0],
                ["Reroute", 0],
                ["Reschedule Delivery", 0],
                ["Returns Electronic Label", 0],
                ["Returns Print Label", 0],
                ["Signature", 0],
                ["Adult Signature", 0],
                ["Package Protection", 0],
                ["Address Correction", 0],
                ["Direct Delivery Only", 0],
                ["Over Maximum Size Surcharge", 0],
                ["Declared Value Markup %", 0],
            ]

            added_dynamic = set()
            
            for as_code, cfg in self.dynamic_surcharge_mapping.items():
                ar_value = normalize_ar_code(cfg.get("ar", ""))
                
                if ar_value not in BILLABLE_AR_CODES:
                    continue
                
                surcharge_name = str(cfg.get("name", "")).strip()

                if surcharge_name == "" or surcharge_name in added_dynamic:
                    continue

                already_exists = any(str(existing[0]).strip() == surcharge_name for existing in rows3)

                if not already_exists:
                    rows3.append([
                        surcharge_name,
                        self.to_amount(self.accessorial_rate_table.get(surcharge_name, 0))
                    ])
                    added_dynamic.add(surcharge_name)

            for r, row in enumerate(rows3, start=2):
                for c, val in enumerate(row, start=1):
                    ws3.cell(row=r, column=c, value=val)

            for ws in wb.worksheets:
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 34

            wb.save(path)

            self.set_status(f"Surcharge rates exported:\n{path}")
        except Exception as e:
            messagebox.showerror("Surcharge rates export failed", f"Error: {str(e)}")
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            gc.collect()

    # =========================================================
    # SAFE NUMBER CONVERTER
    # =========================================================
    def to_amount(self, value, context=""):

        if value is None:
            return 0.0

        try:

            if isinstance(value, str):
                value = value.replace("$", "")
                value = value.replace(",", "")
                value = value.strip()

                if value == "":
                    return 0.0

            return float(value)

        except Exception:

            if not hasattr(self, "warning_logs"):
                self.warning_logs = []

            if context:
                self.warning_logs.append(
                    f"Invalid numeric value [{value}] in {context}"
                )

            return 0.0

    # =========================================================
    # IMPORT ACCESSORIAL CHARGES TEMPLATE
    # =========================================================
    def import_acc_template(self):

        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])

        if not path:
            return
        self.remember_import("acc", path)

        try:
            self.warning_logs = []

            with managed_workbook(path) as wb_import:
                sheet_names = wb_import.sheetnames

                def read_sheet_rows(sheet_name):
                    ws = wb_import[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))

                    if not rows:
                        return [], []

                    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                    data_rows = []

                    for raw_row in rows[1:]:
                        row_dict = {}
                        for idx, header in enumerate(headers):
                            if header:
                                row_dict[header] = raw_row[idx] if idx < len(raw_row) else ""
                        data_rows.append(row_dict)

                    return headers, data_rows

                if "AHS_LPS" not in sheet_names:
                    raise Exception("Missing required sheet: AHS_LPS")

                headers_ahs, rows_ahs = read_sheet_rows("AHS_LPS")

                for col in ["Fee Type", "Shipment Type"]:
                    if col not in headers_ahs:
                        raise Exception(f"AHS_LPS missing required column: {col}")

                for row in rows_ahs:
                    fee_type = str(row.get("Fee Type", "")).strip()
                    shipment_type = str(row.get("Shipment Type", "")).strip()

                    if fee_type == "" or shipment_type == "":
                        continue

                    lookup_key = (fee_type, shipment_type)
                    self.accessorial_rate_table[lookup_key] = {}

                    for col in headers_ahs:
                        col_name = str(col).strip()
                        if col_name.startswith("Zone"):
                            zone_key = col_name.replace("Zone", "").strip()
                            self.accessorial_rate_table[lookup_key][zone_key] = self.to_amount(
                                row.get(col, 0),
                                f"AHS_LPS | {fee_type} | {shipment_type} | {zone_key}"
                            )

                if "FLAT_CHARGES" not in sheet_names:
                    raise Exception("Missing required sheet: FLAT_CHARGES")

                headers_flat, rows_flat = read_sheet_rows("FLAT_CHARGES")

                for col in ["Fee Type", "Amount"]:
                    if col not in headers_flat:
                        raise Exception(f"FLAT_CHARGES missing required column: {col}")

                for row in rows_flat:
                    fee_type = str(row.get("Fee Type", "")).strip()

                    if fee_type == "":
                        continue

                    amount = self.to_amount(row.get("Amount", 0), f"FLAT_CHARGES | {fee_type}")
                    self.accessorial_rate_table[fee_type] = amount

            gc.collect()

            self.render_accessorial_tables()
            self.acc_canvas.update_idletasks()

            try:
                self.save_config()
            except Exception as e:
                print("Auto save config after ACC import failed:", e)

            if self.warning_logs:
                preview_warning = "\n\n".join(self.warning_logs[:5])
                messagebox.showwarning(
                    "Surcharge rates imported with warnings",
                    f"Import completed.\nWarning Count: {len(self.warning_logs)}\nPreview:\n{preview_warning}"
                )
            else:
                self.confirm("Import complete",
                             f"Surcharge rates imported.\n\n"
                             f"{os.path.basename(path)}")
        except Exception as e:
            error_msg = f"Surcharge rates import failed\n\nError: {str(e)}"
            messagebox.showerror("Surcharge rates import failed", error_msg)


    # =========================================================
    # DYNAMIC SURCHARGE METHODS
    # =========================================================
    def add_dynamic_surcharge_row(self):
        self.open_dynamic_surcharge_popup()

    def delete_dynamic_surcharge_row(self):
        selected = self.dynamic_surcharge_tree.selection()
        for item in selected:
            self.dynamic_surcharge_tree.delete(item)
        try:
            self.save_dynamic_surcharge_registry()
        except Exception as e:
            print("Auto save after deleting surcharge failed:", e)

    def save_dynamic_surcharge_registry(self):
        """Save dynamic surcharge mapping to memory + config"""

        tree = getattr(self, "dynamic_surcharge_tree", None)
        if tree is None or not tree.winfo_exists():
            return

        self.dynamic_surcharge_mapping = {}

        for item in self.dynamic_surcharge_tree.get_children():
            vals = self.dynamic_surcharge_tree.item(item, "values")
            if not vals:
                continue

            as_code = str(vals[0]).strip().upper()
            if as_code == "":
                continue

            self.dynamic_surcharge_mapping[as_code] = {
                "ar": normalize_ar_code(vals[1]),
                "name": str(vals[2]).strip(),
                "type": normalize_dynamic_type(vals[3]),
                "fuel_eligible": str(vals[4]).strip().upper() == "YES",
                "enabled": str(vals[5]).strip().upper() == "YES"
            }

        for as_code, cfg in self.dynamic_surcharge_mapping.items():
            surcharge_name = cfg.get("name", "")
            ar_value = normalize_ar_code(cfg.get("ar", ""))
            
            if ar_value in BILLABLE_AR_CODES and surcharge_name != "" and surcharge_name not in self.accessorial_rate_table:
                self.accessorial_rate_table[surcharge_name] = 0

        try:
            self.render_accessorial_tables()
            self.acc_canvas.update_idletasks()
        except Exception:
            pass

        try:
            self.save_config()
        except Exception as e:
            print("Auto save config after dynamic surcharge registry failed:", e)

        self.set_status(f"Custom surcharges saved.\nTotal surcharges: {len(self.dynamic_surcharge_mapping)}")
    def load_dynamic_surcharge_registry(self):
        """Restore dynamic surcharge mapping from memory"""

        tree = getattr(self, "dynamic_surcharge_tree", None)
        if tree is None or not tree.winfo_exists():
            return

        self.dynamic_surcharge_tree.delete(*self.dynamic_surcharge_tree.get_children())

        for as_code, cfg in self.dynamic_surcharge_mapping.items():
            self.dynamic_surcharge_tree.insert("", "end", values=(
                as_code,
                cfg.get("ar", "ACC"),
                cfg.get("name", ""),
                normalize_dynamic_type(cfg.get("type", "REPRICE")),
                "YES" if cfg.get("fuel_eligible", True) else "NO",
                "YES" if cfg.get("enabled", True) else "NO"
            ))

        for as_code, cfg in self.dynamic_surcharge_mapping.items():
            surcharge_name = cfg.get("name", "")
            ar_value = normalize_ar_code(cfg.get("ar", ""))
            
            if ar_value in BILLABLE_AR_CODES and surcharge_name != "" and surcharge_name not in self.accessorial_rate_table:
                self.accessorial_rate_table[surcharge_name] = 0

        try:
            self.render_accessorial_tables()
            self.acc_canvas.update_idletasks()
        except Exception as e:
            print("Accessorial table render after load failed:", e)


    def open_dynamic_surcharge_popup(self):

        popup = tk.Toplevel(self.root)
        popup.title("Add Custom Surcharge")
        popup.geometry("620x320")

        fields = {}

        container = ttk.Frame(popup)
        container.pack(fill="both", expand=True, padx=16, pady=16)

        # AS Code / Surcharge Name stay free text; AR is a fixed vocabulary
        # from the invoice's Charge Classification column, so it is a dropdown
        # (a typo like "AAC" used to save silently and then never price).
        for idx, label in enumerate(["AS Code", "Surcharge Name"]):
            row_i = 0 if label == "AS Code" else 2
            ttk.Label(container, text=label).grid(row=row_i, column=0, sticky="w", pady=8)
            entry = ttk.Entry(container, width=30)
            entry.grid(row=row_i, column=1, padx=8, pady=8)
            fields[label] = entry

        ttk.Label(container, text="AR").grid(row=1, column=0, sticky="w", pady=8)
        ar_var = tk.StringVar(value=AR_DROPDOWN_CHOICES[0])
        ar_combo = ttk.Combobox(
            container, textvariable=ar_var,
            values=AR_DROPDOWN_CHOICES,
            state="readonly", width=40)
        ar_combo.grid(row=1, column=1, padx=8, pady=8)
        ttk.Label(container, text="FRT / FSC are computed by the tool",
                  font=("Segoe UI", 8)).grid(row=1, column=2, sticky="w", padx=6)

        ttk.Label(container, text="Type").grid(row=3, column=0, sticky="w", pady=8)
        type_var = tk.StringVar(value=DYNAMIC_TYPE_CHOICES[0])
        type_combo = ttk.Combobox(container, textvariable=type_var,
                                  values=DYNAMIC_TYPE_CHOICES,
                                  state="readonly", width=40)
        type_combo.grid(row=3, column=1, padx=8, pady=8)

        ttk.Label(container, text="Fuel Eligible").grid(row=4, column=0, sticky="w", pady=8)
        fuel_var = tk.StringVar(value="YES")
        fuel_combo = ttk.Combobox(container, textvariable=fuel_var, values=["YES", "NO"], state="readonly", width=27)
        fuel_combo.grid(row=4, column=1, padx=8, pady=8)

        ttk.Label(container, text="Enabled").grid(row=5, column=0, sticky="w", pady=8)
        enabled_var = tk.StringVar(value="YES")
        enabled_combo = ttk.Combobox(container, textvariable=enabled_var, values=["YES", "NO"], state="readonly", width=27)
        enabled_combo.grid(row=5, column=1, padx=8, pady=8)

        def save_popup():
            as_code = fields["AS Code"].get().strip().upper()
            ar_value = normalize_ar_code(ar_var.get())
            name = fields["Surcharge Name"].get().strip()

            if as_code == "":
                messagebox.showwarning("Missing AS Code", "AS Code cannot be blank.")
                return
            if name == "":
                messagebox.showwarning(
                    "Missing name",
                    "A custom surcharge needs a Surcharge Name -- that name is the\n"
                    "report column and the key used to look up its rate.")
                return

            existing = {
                str(self.dynamic_surcharge_tree.item(i, "values")[0]).strip().upper()
                for i in self.dynamic_surcharge_tree.get_children()
            }
            if as_code in existing:
                messagebox.showwarning(
                    "Duplicate code",
                    f"{as_code} is already in the list. Delete the old row first.")
                return

            values = (as_code, ar_value, name,
                      normalize_dynamic_type(type_var.get()),
                      fuel_var.get(), enabled_var.get())
            self.dynamic_surcharge_tree.insert("", "end", values=values)
            try:
                self.save_dynamic_surcharge_registry()
            except Exception as e:
                print("Auto registry save from popup failed:", e)
            popup.destroy()

        ttk.Button(container, text="Save", command=save_popup).grid(row=6, column=0, columnspan=2, pady=16)


    # =========================================================
    # RUN REPRICING - v2.7 FULL IMPLEMENTATION
    # =========================================================
    def run_repricing(self):
        """
        v2.7 Repricing Logic - v7 PATCH APPLIED

        THREE CRITICAL FIXES:
        1. DDO: REMOVE else: ddo_charge = 0 (DELETE IT)
        2. Commercial Adjustment: NO max(..., 0) to preserve negative values
        3. Final assignment: NO conditional > 0 check
        """

        self.preview.delete(*self.preview.get_children())
        self.base_rate_diagnostics = []

        try:

            billing_file = self.billing_path.get().strip()

            if not billing_file:
                raise Exception(
                    "Please select a billing file first."
                )

            if not getattr(self, "rate_tables", None):
                raise Exception(
                    "No base rate table imported.\n\n"
                    "Go to '📁 Import Files' and import the base rate "
                    "template(s) before generating the report, otherwise "
                    "freight cannot be calculated.")

            self.status_var.set(
                "Processing billing file..."
            )

            # =====================================================
            # LOAD BILLING DATA
            # =====================================================

            if billing_file.lower().endswith(".csv"):
                df = pd.read_csv(
                    billing_file,
                    dtype=str,
                    low_memory=False
                )
            else:
                df = pd.read_excel(
                    billing_file,
                    dtype=str
                )

            df = df.fillna("")

            # ---- Validate the billing file looks like a UPS raw invoice ----
            if df.shape[0] < 1:
                raise Exception(
                    "The selected billing file has no data rows.\n\n"
                    "Please select a valid UPS raw invoice file.")
            if df.shape[1] < 53:
                raise Exception(
                    "This file does not look like a UPS raw invoice.\n\n"
                    f"Expected at least 53 columns but found {df.shape[1]}.\n"
                    "Please select the correct raw invoice file (the wide UPS "
                    "billing export, not a summary or template).")
            # quick content sanity: must contain at least one SHP/RTN/ADJ row
            try:
                cat_codes = df.iloc[:, 34].astype(str).str.upper()
                if not cat_codes.isin(["SHP", "RTN", "ADJ"]).any():
                    raise Exception(
                        "No SHP / RTN / ADJ rows found in this file.\n\n"
                        "The Charge Category Code column (column 35) does not "
                        "contain shipment data. Please check you selected the "
                        "correct UPS raw invoice.")
            except IndexError:
                raise Exception(
                    "This file does not have the expected UPS raw invoice "
                    "layout. Please select the correct file.")

            # Column 20 "Tracking Number" (Excel col U), NOT column 13
            # "Lead Shipment Number". On a multi-package shipment every
            # package shares one lead number, so grouping by 13 merged
            # physically separate parcels -- with their own weights, dims and
            # base rates -- into a single repriced row and dropped the rest.
            # Column 13 is also not always a tracking number at all: on pickup
            # adjustments it holds the pickup record number, and on import
            # entries it holds the brokerage batch number.
            TRACKING_COL = 20
            DATE_COL = 11
            # Her weekly statement groups by INVOICE, not by transaction date.
            INVOICE_DATE_COL = 4
            INVOICE_NUMBER_COL = 5
            REF1_COL = 15
            REF2_COL = 16
            PAYOR_COL = 17
            PACKAGE_QTY_COL = 18
            PACKAGE_ID_COL = 20
            ENTERED_WEIGHT_COL = 26
            BILLED_WEIGHT_COL = 28
            BILLED_WEIGHT_TYPE_COL = 31
            PACKAGE_DIM_TEXT_COL = 32
            ZONE_COL = 33
            AI_COL = 34
            AJ_COL = 35
            AK_COL = 36
            AR_COL = 43
            AS_COL = 44
            AT_COL = 45
            INCENTIVE_COL = 51
            BA_COL = 52
            AA_COL = 222
            AC_COL = 224
            HR_DIM_COL = 225

            def safe_get(row, idx, default=""):
                try:
                    if len(row) > idx:
                        return row.iloc[idx]
                    return default
                except (IndexError, KeyError, AttributeError):
                    return default

            def normalize_zone(value):
                s = str(value).strip()
                if s.endswith(".0"):
                    s = s[:-2]
                s = s.lstrip("0") or "0"
                return s

            def parse_dim(value):
                """Parse dimension string"""
                s = str(value).strip()

                if s == "" or s.lower() == "nan":
                    return (0.0, 0.0, 0.0)

                s = re.sub(r"\s+", "", s)
                s = s.upper().replace("X", "x").replace("*", "x").replace("×", "x")

                try:
                    if "x" in s:
                        parts = s.split("x")
                        nums = []
                        for p in parts:
                            if p.strip() == "":
                                continue
                            nums.append(float(p))

                        if len(nums) >= 3:
                            nums = sorted(nums[:3], reverse=True)
                            return (round(nums[0], 2), round(nums[1], 2), round(nums[2], 2))

                    nums = re.findall(r"\d+\.?\d*", s)

                    if len(nums) >= 3:
                        dims = sorted([float(nums[0]), float(nums[1]), float(nums[2])], reverse=True)
                        return (round(dims[0], 2), round(dims[1], 2), round(dims[2], 2))

                except (ValueError, IndexError):
                    pass

                return (0.0, 0.0, 0.0)

            def cubic(dims):
                return float(dims[0]) * float(dims[1]) * float(dims[2])

            _SERVICE_PATTERNS = [
                    ("surepost", "SurePost"),
                    # Shipping Charge Correction (SCC) descriptions: match the
                    # residential variant BEFORE the bare "ground" so it isn't
                    # mis-classified. Bare "Shipping Charge Correction Ground"
                    # (no res/com word) defaults to Ground Commercial.
                    ("shipping charge correction ground residential", "Ground Residential"),
                    ("shipping charge correction ground commercial", "Ground Commercial"),
                    ("shipping charge correction ground return to sender", "Ground Commercial"),
                    ("shipping charge correction ground undeliverable", "Ground Commercial"),
                    ("shipping charge correction ground", "Ground Commercial"),
                    # Returns / undeliverable / RTS. UPS writes these with no
                    # res/com word at all, so they used to fall through to
                    # base_rate_service.get() -- i.e. whatever the Import
                    # Files dropdown happened to show. A commercial return
                    # was then rated as residential.
                    # Residential variants first, same as the SCC block above.
                    ("returns ground residential", "Ground Residential"),
                    ("returns ground", "Ground Commercial"),
                    ("ground undeliverable return residential", "Ground Residential"),
                    ("ground undeliverable return", "Ground Commercial"),
                    ("ground return to sender residential", "Ground Residential"),
                    ("ground return to sender", "Ground Commercial"),
                    ("zone adjustment ground residential", "Ground Residential"),
                    ("zone adjustment ground", "Ground Commercial"),
                    ("shipping charge correction", "Ground Commercial"),
                    ("ground residential", "Ground Residential"),
                    ("residential", "Ground Residential"),
                    ("ground commercial", "Ground Commercial"),
                    ("standard to canada", "Standard"),
                    ("3 day", "3 Day Select"),
                    ("2nd day air am", "2nd Day Air AM"),
                    ("2nd day air", "2nd Day Air"),
                    ("next day air early", "Next Day Air Early"),
                    ("next day air saver", "Next Day Air Saver"),
                    ("next day air", "Next Day Air"),
                    ("worldwide express plus", "Worldwide Express Plus"),
                    ("worldwide express", "Worldwide Express"),
                    ("worldwide saver", "Worldwide Saver"),
                    ("worldwide expedited", "Worldwide Expedited"),
                    ("worldwide economy", "Worldwide Economy"),
                    ("standard", "Standard"),
            ]

            def service_is_recognised(text_value):
                """True when the description matches a known service pattern.
                When it does not, normalize_service falls back to the Import
                Files dropdown -- an unrelated UI control -- so the caller must
                be able to tell a real classification from that guess."""
                s = str(text_value).strip().lower()
                return any(p in s for p, _n in _SERVICE_PATTERNS)

            def normalize_service(text_value, ai_value=""):
                s = str(text_value).strip().lower()

                services = _SERVICE_PATTERNS


                for pattern, service_name in services:
                    if pattern in s:
                        return service_name

                return self.base_rate_service.get()

            def lookup_base_rate(service, zone, billable_weight, tracking=""):
                """
                v2.6 Base Rate lookup:
                1. CEILING weight
                2. Exact weight lookup only
                3. If not found, return 0 and log Weight Not Found
                """
                try:
                    if not hasattr(self, "rate_tables") or not self.rate_tables:
                        msg = f"[{tracking}] No imported base rate table"
                        self.base_rate_diagnostics.append(msg)
                        return 0.0, msg

                    if service not in self.rate_tables:
                        service = self.base_rate_service.get()

                    if service not in self.rate_tables:
                        msg = f"[{tracking}] No rate table for service: {service}"
                        self.base_rate_diagnostics.append(msg)
                        return 0.0, msg

                    df_rate = self.rate_tables[service].copy()
                    zone_col = f"Zone {normalize_zone(zone)}"

                    if zone_col not in df_rate.columns:
                        msg = f"[{tracking}] Missing rate column: {zone_col}"
                        self.base_rate_diagnostics.append(msg)
                        return 0.0, msg

                    try:
                        weight_used = int(math.ceil(float(billable_weight)))
                    except Exception:
                        msg = f"[{tracking}] Invalid weight for base rate lookup: {billable_weight}"
                        self.base_rate_diagnostics.append(msg)
                        return 0.0, msg

                    df_rate["Weight_NUM"] = pd.to_numeric(df_rate["Weight"], errors="coerce")
                    df_rate = df_rate.dropna(subset=["Weight_NUM"])
                    match = df_rate[df_rate["Weight_NUM"].astype(int) == weight_used]

                    if match.empty:
                        msg = (
                            f"[{tracking}] Weight Not Found in Base Rate Template: "
                            f"service={service}, zone={zone_col}, weight={weight_used}"
                        )
                        self.base_rate_diagnostics.append(msg)
                        return 0.0, msg

                    value = match.iloc[0].get(zone_col, 0)
                    return self.to_amount(value), ""

                except (KeyError, ValueError, TypeError) as rate_error:
                    msg = f"[{tracking}] Rate lookup error: {str(rate_error)}"
                    self.base_rate_diagnostics.append(msg)
                    return 0.0, msg

            def normalize_zone_lookup(zone):
                z = str(zone).strip()
                if z.endswith(".0"):
                    z = z[:-2]
                if z in ["2", "3", "4", "5", "6", "7", "8", "44", "45", "46"]:
                    return z
                return "8"

            def accessorial_lookup(fee_type, shipment_type, zone):
                table = getattr(self, "accessorial_rate_table", {})
                zone_key = normalize_zone_lookup(zone)

                if fee_type == "Residential":
                    # Row renamed to plain "Residential Surcharge". Older ACC
                    # templates still carry the long "(Skip if Residential
                    # Service)" wording, so fall back to it when the short name
                    # is absent -- an existing template must not silently
                    # reprice residential at 0.
                    fee_type = "Residential Surcharge"
                    legacy = "Residential Surcharge (Skip if Residential Service)"
                    if fee_type not in table and legacy in table:
                        fee_type = legacy

                if fee_type in table and isinstance(table[fee_type], (int, float)):
                    return self.to_amount(table[fee_type])

                lookup_key = (fee_type, shipment_type)

                if lookup_key in table and isinstance(table[lookup_key], dict):
                    return self.to_amount(table[lookup_key].get(zone_key, 0))

                if fee_type in table and isinstance(table[fee_type], dict):
                    return self.to_amount(table[fee_type].get(zone_key, 0))

                return 0.0

            # =====================================================
            # LOAD RAW INVOICE
            # =====================================================
            if billing_file.lower().endswith(".csv"):
                raw_df = pd.read_csv(billing_file, header=None,
                                     dtype=str, low_memory=False).fillna("")
            else:
                raw_df = pd.read_excel(billing_file, header=None,
                                       dtype=str).fillna("")

            if raw_df.empty:
                raise Exception("Billing file is empty.")

            shipment_groups = {}
            account_rows = []
            unknown_rows = []
            warning_rows = []

            ups_invoice_total = 0.0

            # =====================================================
            # GROUP RAW SHIPMENT ROWS FIRST
            # =====================================================
            for row_index, row in raw_df.iterrows():

                tracking = str(safe_get(row, TRACKING_COL, "")).strip()

                ai = str(safe_get(row, AI_COL, "")).strip().upper()
                aj = str(safe_get(row, AJ_COL, "")).strip().upper()
                ak = str(safe_get(row, AK_COL, "")).strip().upper()
                ar = str(safe_get(row, AR_COL, "")).strip().upper()
                as_code = str(safe_get(row, AS_COL, "")).strip().upper()
                at = str(safe_get(row, AT_COL, "")).strip()

                ba = self.to_amount(safe_get(row, BA_COL, 0))

                if ba != 0:
                    ups_invoice_total += ba

                # MIS (miscellaneous) and MSC (billing adjustment) are
                # account-level whether or not UPS attached a tracking number.
                # Leaving a stray tracking number in charge of them put a
                # whole week's adjustment on a single parcel.
                is_account_level = (tracking == "" or ai == "MIS"
                                    or ar == "MSC")

                if is_account_level:
                    account_rows.append({
                        "Row": row_index + 1,
                        "Charge Description": at if at else as_code,
                        "Charge Description Code": as_code,
                        "Charge Classification Code": ar,
                        "Type Code 1": ai,
                        "Type Detail Code 1": aj,
                        "Type Detail Value 1": ak,
                        "Amount": ba
                    })
                    continue

                # Layer = Charge Category Code (SHP / RTN / ADJ). Group by
                # (tracking, layer) so each layer is repriced separately and
                # shown side by side in the INV sheet.
                layer = ai if ai in ("SHP", "RTN", "ADJ") else "SHP"
                shipment_groups.setdefault((tracking, layer), []).append(row)

            # =====================================================
            # CONSOLIDATE + REPRICE
            # =====================================================
            preview_cols = [
                "Tracking", "Layer", "Zone", "Shipment Type", "Entered Weight", "Billed Weight", "Dim Weight",
                "Billable Weight", "Length", "Width", "Height", "CAL Base Rate",
                "Shipping Charge Correction Ground Undeliverable Return",
                "Residential Adjustment", "Commercial Adjustment", "Direct Delivery Only",
                "Shipping Charge Correction Large Package Surcharge",
                "CAL Fuel", "Residential",
                "DAS Commercial", "DAS Residential", "DAS Extended Commercial", "DAS Extended Residential",
                "Remote Area Commercial", "Remote Area Residential", "Remote Area - AK", "Remote Area - HI",
                "Additional Handling Surcharge - Weight",
                "Additional Handling Surcharge - Dimension",
                "Additional Handling Surcharge - Packaging", "Large Package Surcharge", "Over Maximum Size Surcharge",
                "Undeliverable Return", "Address Correction",
                "Declared Value", "Return To Sender", "2nd Mile Road Base", "Signature", "Adult Signature", "Package Protection", "CAL Total", "UPS Invoice Total"
            ]

            # Custom ACC surcharges become real columns, inserted just before
            # the totals so CAL Total / UPS Invoice Total stay rightmost.
            dyn_acc_cols = self.dynamic_acc_names(exclude=preview_cols)
            dyn_acc_nonfuel = self.dynamic_acc_nonfuel_names()
            dyn_acc_report_only = self.dynamic_acc_report_only_names()
            # Columns that DO count toward CAL Total (and therefore fuel).
            dyn_acc_billable = [c for c in dyn_acc_cols
                                if c not in dyn_acc_report_only]

            if dyn_acc_cols:
                _tail = ["CAL Total", "UPS Invoice Total"]
                preview_cols = [c for c in preview_cols if c not in _tail] + dyn_acc_cols + _tail

            output_rows = []

            # =====================================================
            # PRE-PASS: SHP/RTN freight A per tracking
            #   For each tracking, compute the A-side freight (HR dims + AA
            #   entered weight) of its SHP/RTN layers. An ADJ with a matching
            #   SHP/RTN in the same invoice reuses this A directly instead of
            #   recomputing from the ADJ row's own HR/AA.
            # =====================================================
            try:
                _pp_dim_factor = float(self.dim_factor.get())
            except (ValueError, TypeError):
                _pp_dim_factor = 300.0
            try:
                _pp_ahs_min = float(self.ahs_min_billable_weight.get())
            except (ValueError, TypeError):
                _pp_ahs_min = 15.0

            shp_rtn_freight_A = {}
            # Per-tracking SHP/RTN dimensions, so an ADJ whose own rows carry
            # no dimensions can reuse the same shipment's dimensions (spec
            # rule: ADJ dimension falls back to the matching SHP/RTN).
            shp_rtn_dims = {}
            # Per-tracking SHP/RTN entered & billed weights, so an ADJ whose own
            # rows carry no weight can reuse the matching shipment's weight.
            shp_rtn_weights = {}
            # Per-tracking SHP/RTN service (Ground Residential / Commercial), so
            # an ADJ "Shipping Charge Correction Ground" row (which carries no
            # RES/COM signal of its own) keeps the original shipping channel
            # instead of defaulting to Commercial. SCC only corrects the amount;
            # it does not change the service.
            shp_rtn_service = {}
            for (pp_tracking, pp_layer), pp_rows in shipment_groups.items():
                if pp_layer not in ("SHP", "RTN"):
                    continue
                pp_zone = ""
                pp_hr = []
                pp_ag = []
                pp_entered = 0.0
                pp_billed = 0.0
                pp_service = ""
                for r in pp_rows:
                    if pp_zone in ("", "0"):
                        # Skip the placeholder 000 zone (see main loop note).
                        _pz = normalize_zone(safe_get(r, ZONE_COL, ""))
                        if _pz not in ("", "0"):
                            pp_zone = _pz
                        elif pp_zone == "":
                            pp_zone = _pz
                    ar = str(safe_get(r, AR_COL, "")).strip().upper()
                    asc = str(safe_get(r, AS_COL, "")).strip().upper()
                    at = str(safe_get(r, AT_COL, "")).strip()
                    ai_ = str(safe_get(r, AI_COL, "")).strip().upper()
                    if pp_service == "" and (asc == "FRT" or ar == "FRT"):
                        pp_service = normalize_service(at, ai_)
                    ew = self.to_amount(safe_get(r, ENTERED_WEIGHT_COL, 0))
                    if ew > pp_entered:
                        pp_entered = ew
                    bw_ = self.to_amount(safe_get(r, BILLED_WEIGHT_COL, 0))
                    if bw_ > pp_billed:
                        pp_billed = bw_
                    d_hr = parse_dim(safe_get(r, HR_DIM_COL, ""))
                    d_ag = parse_dim(safe_get(r, PACKAGE_DIM_TEXT_COL, ""))
                    if cubic(d_hr) > 0:
                        pp_hr.append(d_hr)
                    if cubic(d_ag) > 0:
                        pp_ag.append(d_ag)
                if pp_service == "":
                    pp_service = self.base_rate_service.get()
                # SHP/RTN A dimension: HR first, fall back to AG
                if pp_hr:
                    dimA = sorted(pp_hr, key=cubic, reverse=True)[0]
                elif pp_ag:
                    dimA = sorted(pp_ag, key=cubic, reverse=True)[0]
                else:
                    dimA = (0.0, 0.0, 0.0)
                dwA = round(cubic(dimA) / _pp_dim_factor, 2) if _pp_dim_factor > 0 else 0.0
                # No dimensions at all -> use billed weight as the A weight
                pp_weight = pp_entered if cubic(dimA) > 0 else max(pp_billed, pp_entered)
                bwA = math.ceil(max(pp_weight, dwA, 0.0))
                baseA, _ = lookup_base_rate(pp_service, pp_zone, bwA, pp_tracking)
                # Accumulate across SHP and RTN of the same tracking
                shp_rtn_freight_A[pp_tracking] = shp_rtn_freight_A.get(pp_tracking, 0.0) + baseA

                # Remember this shipment's best HR / AG dims for ADJ fallback.
                best_hr = sorted(pp_hr, key=cubic, reverse=True)[0] if pp_hr else (0.0, 0.0, 0.0)
                best_ag = sorted(pp_ag, key=cubic, reverse=True)[0] if pp_ag else (0.0, 0.0, 0.0)
                prev = shp_rtn_dims.get(pp_tracking, ((0.0, 0.0, 0.0), (0.0, 0.0, 0.0)))
                keep_hr = best_hr if cubic(best_hr) >= cubic(prev[0]) else prev[0]
                keep_ag = best_ag if cubic(best_ag) >= cubic(prev[1]) else prev[1]
                shp_rtn_dims[pp_tracking] = (keep_hr, keep_ag)

                # Remember best (max) entered/billed weight for ADJ fallback.
                pw = shp_rtn_weights.get(pp_tracking, (0.0, 0.0))
                shp_rtn_weights[pp_tracking] = (
                    max(pw[0], pp_entered), max(pw[1], pp_billed))

                # Remember the original service for the ADJ SCC fallback.
                # Residential wins if any SHP/RTN layer of this tracking is
                # residential (RTN layers are often generic).
                prev_svc = shp_rtn_service.get(pp_tracking, "")
                if not prev_svc or "RESIDENTIAL" in str(pp_service).upper():
                    shp_rtn_service[pp_tracking] = pp_service

            for (tracking, layer), rows in shipment_groups.items():

                row_result = {col: 0 for col in preview_cols}
                row_result["Tracking"] = tracking
                row_result["Layer"] = layer

                # FIX: Initialize per shipment
                ddo_charge = 0
                residential_adj = 0
                commercial_adj = 0

                ag_dims = []
                hr_dims = []

                max_entered_weight = 0.0
                max_billed_weight = 0.0
                group_acc_codes = set()
                service_was_guessed = False
                billed_weight_type = None
                zone = ""
                zone_din = ""          # zone seen only on DIN rows
                shipment_type = ""
                shipment_description = ""
                _layer_has_freight = False

                # Shipment-level static info
                for row in rows:
                    # A DIN leg (delivery intercept / return to sender) is a
                    # SEPARATE movement with its own zone: one parcel had SHP
                    # zone 6 and a DIN leg at zone 5, and because the DIN row
                    # came first the whole shipment was rated at 5.
                    #
                    # But a layer can be ENTIRELY DIN. Skipping it outright
                    # then leaves the layer with no zone at all, no base rate
                    # is found, and the row comes out 0. So DIN is held aside
                    # and used only if nothing better turns up.
                    _aj_now = str(safe_get(row, AJ_COL, "")).strip().upper()
                    if _aj_now == "DIN":
                        if not zone_din:
                            _zd = normalize_zone(safe_get(row, ZONE_COL, ""))
                            if _zd not in ("", "0"):
                                zone_din = _zd
                    if zone in ("", "0") and _aj_now != "DIN":
                        # "0" is not a real UPS zone: accessorial-only ADJ rows
                        # (e.g. a standalone Reschedule Delivery fee) carry a
                        # 000 zone. Keep scanning until a genuine zone appears,
                        # otherwise the whole shipment gets rated at zone 0 and
                        # loses its base rate.
                        _z = normalize_zone(safe_get(row, ZONE_COL, ""))
                        if _z not in ("", "0"):
                            zone = _z
                        elif zone == "":
                            zone = _z

                    ar = str(safe_get(row, AR_COL, "")).strip().upper()
                    as_code = str(safe_get(row, AS_COL, "")).strip().upper()
                    at = str(safe_get(row, AT_COL, "")).strip()
                    ai = str(safe_get(row, AI_COL, "")).strip().upper()
                    aj = str(safe_get(row, AJ_COL, "")).strip().upper()

                    # WEIGHT: whole-group nonzero max. Adjustment rows
                    # (RES/COM) can legitimately carry the weight basis (e.g.
                    # a Residential Adjustment row holds AC), so we do not
                    # restrict capture to 003 freight rows.
                    entered_w = self.to_amount(safe_get(row, ENTERED_WEIGHT_COL, 0))
                    billed_w = self.to_amount(safe_get(row, BILLED_WEIGHT_COL, 0))

                    if entered_w > max_entered_weight:
                        max_entered_weight = entered_w

                    if billed_w > max_billed_weight:
                        max_billed_weight = billed_w

                    # Surcharge codes present anywhere in this group. Used
                    # below to tell a genuine UPS minimum-billable weight
                    # apart from a stray artifact.
                    if ar == "ACC" and as_code:
                        group_acc_codes.add(as_code)

                    # UPS Billed Weight Type: 9 = dimensional weight governs,
                    # 2 = actual weight, 3 = surcharge-minimum weight,
                    # 8 = corrected (SCC) basis. Capture the LAST non-empty
                    # value in the group so an SCC re-basis (8) supersedes the
                    # original shipment's basis (3/9).
                    _bwt = str(safe_get(row, BILLED_WEIGHT_TYPE_COL, "")).strip()
                    if _bwt and _bwt.lower() != "nan":
                        try:
                            billed_weight_type = int(float(_bwt))
                        except (ValueError, TypeError):
                            pass

                    # capture descriptive fields (first non-empty per group)
                    if not row_result.get("Date"):
                        dv = safe_get(row, DATE_COL, "")
                        if str(dv).strip():
                            row_result["Date"] = str(dv).strip()
                    if not row_result.get("Invoice Number"):
                        _iv = safe_get(row, INVOICE_NUMBER_COL, "")
                        if str(_iv).strip():
                            row_result["Invoice Number"] = str(_iv).strip()
                    if not row_result.get("Invoice Date"):
                        _id = safe_get(row, INVOICE_DATE_COL, "")
                        if str(_id).strip():
                            row_result["Invoice Date"] = _id
                    if not row_result.get("Ref1"):
                        rv = safe_get(row, REF1_COL, "")
                        if str(rv).strip():
                            row_result["Ref1"] = str(rv).strip()
                    if not row_result.get("Ref2"):
                        rv2 = safe_get(row, REF2_COL, "")
                        if str(rv2).strip():
                            row_result["Ref2"] = str(rv2).strip()

                    # UPS files several accessorials as AR=FRT with their own
                    # wording -- "Address Correction Ground" is a $6.31 fee,
                    # not a service. Taking it as the shipment description
                    # makes normalize_service() fail, marks the lane
                    # unsupported, and the charge is never priced.
                    _desc_is_accessorial = (
                        ai == "ADJ" and aj in ACCESSORIAL_ADJ_DETAIL_CODES)

                    if (shipment_description == ""
                            and (as_code == "FRT" or ar == "FRT")
                            and not _desc_is_accessorial):
                        shipment_description = at
                        shipment_type = normalize_service(at, ai)
                        # Record when the description matched nothing. In that
                        # case shipment_type is NOT a classification -- it is
                        # whatever the Import Files "Currently Editing"
                        # dropdown happens to be set to. Reporting that as if
                        # it were the real service is how "Ground
                        # Hundredweight" ended up shown as "Ground
                        # Residential".
                        if str(at).strip() and not service_is_recognised(at):
                            service_was_guessed = True

                    # Track whether THIS layer carries a freight line at all. An
                    # ADJ made up purely of accessorial corrections (e.g. a
                    # standalone Reschedule Delivery fee) has none, so it never
                    # re-rated the base/size charges and must not overwrite the
                    # shipment's values during the final merge.
                    if ar == "FRT" or as_code == "FRT":
                        _layer_has_freight = True

                    ag = safe_get(row, PACKAGE_DIM_TEXT_COL, "")
                    hr = safe_get(row, HR_DIM_COL, "")
                    d_ag = parse_dim(ag)
                    d_hr = parse_dim(hr)

                    if cubic(d_ag) > 0:
                        ag_dims.append(d_ag)
                    if cubic(d_hr) > 0:
                        hr_dims.append(d_hr)


                # Nothing but DIN rows in this layer -- use their zone rather
                # than leaving the layer unzoned and unrated.
                if zone in ("", "0") and zone_din:
                    zone = zone_din

                if shipment_type == "":
                    shipment_type = self.base_rate_service.get()
                    if str(shipment_description).strip():
                        service_was_guessed = True

                # RTN (returns) channel is always Commercial.
                if layer == "RTN":
                    shipment_type = "Ground Commercial"

                # =====================================================
                # DIMENSION SELECTION (A/C TWO-PASS MODEL)
                #   A (SHP/RTN) freight = HR dimensions + AA (entered) weight
                #   C (correct)  freight = AG dimensions + AC (billed) weight
                #   B (ADJ)      freight = C - A
                # =====================================================
                selected_dim_ag = (0.0, 0.0, 0.0)
                selected_dim_hr = (0.0, 0.0, 0.0)

                if ag_dims:
                    selected_dim_ag = sorted(ag_dims, key=cubic, reverse=True)[0]

                if hr_dims:
                    selected_dim_hr = sorted(hr_dims, key=cubic, reverse=True)[0]

                # A uses HR dims; C uses AG dims.
                #   SHP/RTN: HR is the package dimension; if HR is empty fall
                #            back to AG (some packages only carry AG).
                #   ADJ:     keep the two passes strictly separate (A=HR, C=AG)
                #            so B = C - A is meaningful.
                # Pure ADJ: an ADJ layer whose tracking has no SHP/RTN layer in
                # this invoice. It is a standalone credit/debit document.
                pure_adj_layer = (layer == "ADJ" and tracking not in shp_rtn_dims)

                # ADJ carrying NO dimensions of its own (both AG "Package
                # Dimension" and HR "Detail Keyed Dim" are 0/empty):
                # skip the base rate lookup entirely.
                # Only applies to ADJ-only trackings: no SHP/RTN layer exists in
                # this invoice for the same tracking. If a SHP/RTN exists, its
                # dimensions are reused via the fallback below and the base rate
                # is calculated normally.
                adj_no_own_dims = (
                    pure_adj_layer
                    and cubic(selected_dim_ag) <= 0
                    and cubic(selected_dim_hr) <= 0
                )

                # NEW: ADJ carrying NO weight of its own (both AA "Entered
                # Weight" and AC "Billed Weight" are 0) -> skip the base rate
                # lookup, exactly like the no-dimension case above. Same scope:
                # only for ADJ-only trackings (no SHP/RTN in this invoice to
                # fall back on). e.g. 1ZXXXXXXXXXXXXXXXX.
                adj_no_own_weight = (
                    pure_adj_layer
                    and max_entered_weight <= 0
                    and max_billed_weight <= 0
                )

                # Either condition alone suppresses the base rate.
                adj_skip_base = adj_no_own_dims or adj_no_own_weight

                if layer == "ADJ":
                    # ADJ's own rows may carry no dimensions. Fall back to the
                    # matching SHP/RTN shipment's dimensions (spec rule 4).
                    fb_hr, fb_ag = shp_rtn_dims.get(
                        tracking, ((0.0, 0.0, 0.0), (0.0, 0.0, 0.0)))
                    if cubic(selected_dim_hr) <= 0:
                        selected_dim_hr = fb_hr
                    if cubic(selected_dim_ag) <= 0:
                        # C dimension: prefer AG; if neither ADJ nor SHP has AG,
                        # use the SHP HR dimension as the package size.
                        selected_dim_ag = fb_ag if cubic(fb_ag) > 0 else fb_hr
                    dim_A = selected_dim_hr
                    dim_C = selected_dim_ag
                else:
                    dim_A = selected_dim_hr if cubic(selected_dim_hr) > 0 else selected_dim_ag
                    dim_C = selected_dim_ag if cubic(selected_dim_ag) > 0 else selected_dim_hr

                # A uses original entered weight (AA); C uses billed weight (AC)
                weight_A = max_entered_weight
                weight_C = max_billed_weight

                # ADJ whose own rows carry no weight: reuse the matching
                # SHP/RTN shipment's entered/billed weight (spec rule, same as
                # the dimension fallback).
                if layer == "ADJ":
                    fb_ent, fb_bill = shp_rtn_weights.get(
                        tracking, (0.0, 0.0))
                    if weight_A <= 0:
                        weight_A = fb_ent if fb_ent > 0 else fb_bill
                    if weight_C <= 0:
                        weight_C = fb_bill if fb_bill > 0 else fb_ent

                # SHP/RTN whose entered weight is 0 (e.g. RTN returns carry the
                # weight only in the billed column): use the billed weight as
                # the A weight so billable isn't understated to the dim weight.
                if layer != "ADJ" and weight_A <= 0:
                    weight_A = max(max_billed_weight, max_entered_weight)

                # The C package (audited-correct) drives surcharge triggers and
                # the displayed dimensions/weights.
                selected_dim = dim_C
                has_ag_dimension = cubic(dim_C) > 0

                length, width, height = selected_dim
                cubic_inches = cubic(selected_dim)

                try:
                    dim_factor = float(self.dim_factor.get())
                except (ValueError, TypeError):
                    dim_factor = 300.0

                dim_weight = round(cubic_inches / dim_factor, 2) if dim_factor > 0 else 0.0

                # Hoisted: minimum-billable thresholds are needed by the
                # billed-weight sanitizer below, before the trigger sections.
                try:
                    oversize_min_cfg = float(self.oversize_min_billable_weight.get())
                except (ValueError, TypeError):
                    oversize_min_cfg = 90.0
                try:
                    ahs_min_cfg = float(self.ahs_min_billable_weight.get())
                except (ValueError, TypeError):
                    ahs_min_cfg = 15.0

                # =====================================================
                # FINAL TRACKING WEIGHT (C package, for surcharge triggers)
                # =====================================================
                # FIX: UPS's Billed Weight column already has the LPS/AHS
                # minimum baked in (e.g. a 30 lb parcel billed as Large
                # Package shows 90). Feeding that back in as "actual weight"
                # double-applies the minimum and falsely fires the AHS-Weight
                # trigger (>50). Trust the billed weight only when it is
                # corroborated by a real weight source: the entered/scale
                # weight or the dim weight. Otherwise it is a surcharge
                # artifact and must be discarded.
                # FIX: an ADJ layer carries no entered weight of its own, so
                # max_entered_weight is 0 here and any billed weight that
                # happens to equal a minimum-billable threshold (e.g. a real
                # 40 lb parcel vs. ahs_min=40) was wrongly discarded as an
                # artifact. Corroborate against the matching SHP/RTN layer's
                # real weight, the same fallback weight_A/weight_C already use.
                _real_weight = max_entered_weight
                if layer == "ADJ":
                    _fb_ent, _fb_bill = shp_rtn_weights.get(
                        tracking, (0.0, 0.0))
                    _real_weight = max(_real_weight, _fb_ent)
                # A billed weight that equals a configured minimum is NOT
                # automatically an artifact -- UPS applies those same minimums
                # for real. Two pieces of invoice evidence say it is genuine:
                #
                #   * Billed Weight Type 3 means "surcharge-minimum weight",
                #     i.e. UPS itself states the weight came from a minimum.
                #   * The group carries the surcharge that owns that minimum
                #     (LPS family for the oversize minimum, AHS family for the
                #     AHS minimum).
                #
                # Without this, a 10 lb parcel billed at 90 lb under a Large
                # Package minimum was reverted to 10 lb and repriced there,
                # producing a large phantom overcharge. These rows carry no
                # dimensions, so large_package_trigger cannot re-apply the
                # minimum later and the weight was simply lost.
                _LPS_FAMILY = {"LPS", "LPR", "LCC", "LCR", "LWC", "LWR", "LCT"}
                _AHS_FAMILY = {"AHS", "AHW", "AHG", "AHL", "AHV", "AHC",
                               "AHP", "AHD", "AHB"}

                def _minimum_is_genuine(candidate):
                    if billed_weight_type == 3:
                        return True
                    if (oversize_min_cfg > 0
                            and math.isclose(candidate, oversize_min_cfg, abs_tol=0.01)
                            and (group_acc_codes & _LPS_FAMILY)):
                        return True
                    if (ahs_min_cfg > 0
                            and math.isclose(candidate, ahs_min_cfg, abs_tol=0.01)
                            and (group_acc_codes & _AHS_FAMILY)):
                        return True
                    return False

                _min_artifacts = (oversize_min_cfg, ahs_min_cfg)
                _looks_like_min = (
                    weight_C > _real_weight
                    and weight_C > dim_weight
                    and any(math.isclose(weight_C, _m, abs_tol=0.01)
                            for _m in _min_artifacts if _m > 0)
                )

                # Floor carried forward so the minimum is not lost downstream:
                # these rows have no dimensions, so large_package_trigger
                # cannot re-derive it later.
                minimum_billable_floor = 0.0

                if _looks_like_min and not _minimum_is_genuine(weight_C):
                    # weight_C is purely a minimum-billable artifact.
                    weight_C = _real_weight
                elif _looks_like_min:
                    minimum_billable_floor = weight_C

                actual_weight = max(weight_C, _real_weight)

                minimum_billable = 0.0

                # Large Package trigger
                try:
                    oversize_longest = float(self.oversize_longest_side.get())
                    oversize_cubic = float(self.oversize_cubic_inches.get())
                    oversize_lg = float(self.oversize_length_girth.get())
                    oversize_weight = float(self.oversize_actual_weight.get())
                    oversize_min = float(self.oversize_min_billable_weight.get())
                except (ValueError, TypeError):
                    oversize_longest, oversize_cubic, oversize_lg, oversize_weight, oversize_min = 96, 17280, 130, 110, 90

                length_girth = length + 2 * (width + height)
                length_girth = round(length_girth, 2)

                # LARGE PACKAGE weight trigger uses the ACTUAL (scale) weight,
                # per UPS rules -- NOT the dim-inflated billed weight. On a
                # dimensional-weight-governed (type 9) correction the billed
                # weight can be pushed far above the real scale weight (e.g. a
                # 43 lb parcel billed at 123 lb from its dim weight). Feeding
                # that into the 110 lb LPS threshold falsely triggers a Large
                # Package Surcharge when the parcel is neither oversize by
                # dimension nor >110 lb by scale. Prefer the matched SHP/RTN
                # entered (scale) weight; then this layer's own entered; only
                # fall back to billed when no scale weight exists at all.
                _fb_ent_lps, _fb_bill_lps = shp_rtn_weights.get(
                    tracking, (0.0, 0.0))
                if _fb_ent_lps > 0:
                    _lps_scale_weight = _fb_ent_lps
                elif max_entered_weight > 0:
                    _lps_scale_weight = max_entered_weight
                elif _fb_bill_lps > 0:
                    _lps_scale_weight = _fb_bill_lps
                else:
                    _lps_scale_weight = max_billed_weight

                lps_longest_trigger = self.use_oversize_longest.get() and length > oversize_longest
                lps_cubic_trigger = self.use_oversize_cubic.get() and cubic_inches > oversize_cubic
                lps_lg_trigger = self.use_oversize_lg.get() and length_girth > float(oversize_lg) and not math.isclose(length_girth, float(oversize_lg), abs_tol=0.01)
                # On a type-9 pure ADJ there is no trustworthy scale weight at
                # all (both columns are dim-driven billing weights), so the
                # weight path cannot fire -- the size triggers above decide.
                if pure_adj_layer and billed_weight_type == 9:
                    _lps_scale_weight = 0.0
                lps_weight_trigger = self.use_oversize_weight.get() and _lps_scale_weight > oversize_weight

                large_package_trigger = lps_longest_trigger or lps_cubic_trigger or lps_lg_trigger or lps_weight_trigger

                if large_package_trigger:
                    minimum_billable = max(minimum_billable, oversize_min)

                # =====================================================
                # AHS TRIGGERS (WEIGHT + DIMENSION + PACKAGING)
                # =====================================================
                try:
                    ahs_weight_threshold = float(self.ahs_weight_threshold.get())
                    ahs_longest = float(self.ahs_longest_side.get())
                    ahs_second = float(self.ahs_second_side.get())
                    ahs_cubic = float(self.ahs_cubic_inches.get())
                    ahs_lg_limit = float(self.ahs_lg_limit.get())
                    ahs_min = float(self.ahs_min_billable_weight.get())
                except (ValueError, TypeError):
                    ahs_weight_threshold = 50
                    ahs_longest = 48
                    ahs_second = 30
                    ahs_cubic = 10368
                    ahs_lg_limit = 105
                    ahs_min = 15

                # AHS-Weight is assessed on the ACTUAL (scale) weight per UPS
                # rules -- NOT billable or dimensional weight, and NOT the
                # SCC-corrected billing weight.
                #
                # The trustworthy scale weight is the ORIGINAL shipment's entered
                # weight. A pitfall: an ADJ (SCC) row's own "entered" column is
                # often the CORRECTED BILLING weight carried over from the
                # shipment's billed column (e.g. a 41 lb parcel whose billed
                # weight was dim-inflated to 95 shows entered=95 on the ADJ row).
                # Using that would falsely fire the 50 lb weight trigger. So:
                #   * If a matched SHP/RTN exists, use ITS entered weight (the
                #     real scale weight) -- never the ADJ's entered.
                #   * SHP/RTN with entered=0 (undeliverable returns) carry the
                #     real weight in the billed column; use billed there.
                #   * A pure ADJ (no origin shipment) has no trustworthy scale
                #     weight; on Billed Weight Type 9 (dim weight governs) UPS
                #     charges AHS by dimension, so suppress the weight trigger.
                _fb_ent_w, _fb_bill_w = shp_rtn_weights.get(
                    tracking, (0.0, 0.0))
                if layer == "ADJ":
                    if _fb_ent_w > 0:
                        _ahs_scale_weight = _fb_ent_w
                    elif _fb_bill_w > 0:
                        _ahs_scale_weight = _fb_bill_w
                    else:
                        # Pure ADJ with no origin: entered is the only signal but
                        # it is an SCC billing artifact under type 9.
                        _ahs_scale_weight = max_entered_weight
                else:
                    _ahs_scale_weight = max_entered_weight
                    if _ahs_scale_weight <= 0:
                        _ahs_scale_weight = max_billed_weight

                if pure_adj_layer and billed_weight_type == 9:
                    ahs_weight_trigger = False
                else:
                    ahs_weight_trigger = _ahs_scale_weight > ahs_weight_threshold
                ahs_dimension_trigger = (
                    (length > ahs_longest) or
                    (width > ahs_second) or
                    (cubic_inches > ahs_cubic) or
                    (length_girth > ahs_lg_limit)
                )
                ahs_packaging_trigger = False

                # ----- A-package (HR dims) triggers, for pure-ADJ C-A -----
                # When a prior-period ADJ carries both AG (C) and HR (A), the
                # size-driven surcharges are charged as C minus A. Compute the
                # A-package triggers so an already-present surcharge nets to 0.
                a_len, a_wid, a_hgt = dim_A
                a_cubic = cubic(dim_A)
                a_length_girth = round(a_len + 2 * (a_wid + a_hgt), 2)
                a_actual_weight = max(weight_A, max_entered_weight)
                # The A-package LPS weight trigger must use the same true scale
                # weight rule as the C side (see _lps_scale_weight above). On a
                # type-9 (dim-weight-governed) SCC both the entered and billed
                # columns hold corrected BILLING weights, not a real scale weight
                # -- e.g. a parcel whose dim weight is 55.8 shows entered 111 /
                # billed 121. Feeding 111 into the 110 lb threshold makes the A
                # package look oversize too, so a genuine C-only Large Package
                # (LG 131 vs A's 129) nets to zero and the surcharge silently
                # disappears.
                if pure_adj_layer and billed_weight_type == 9:
                    _a_lps_weight = 0.0
                else:
                    _a_lps_weight = a_actual_weight
                a_lps_trigger = (
                    (self.use_oversize_longest.get() and a_len > oversize_longest) or
                    (self.use_oversize_cubic.get() and a_cubic > oversize_cubic) or
                    (self.use_oversize_lg.get() and a_length_girth > float(oversize_lg)
                     and not math.isclose(a_length_girth, float(oversize_lg), abs_tol=0.01)) or
                    (self.use_oversize_weight.get() and _a_lps_weight > oversize_weight))
                # Same rule as the C package: suppress AHS - Weight only for a
                # pure ADJ governed by type-9 dim weight, so the C - A net is
                # consistent. (A-package triggers only matter in the pure-ADJ
                # both-dims C-A path.)
                if pure_adj_layer and billed_weight_type == 9:
                    a_ahs_weight_trigger = False
                else:
                    a_ahs_weight_trigger = a_actual_weight > ahs_weight_threshold
                a_ahs_dimension_trigger = (
                    (a_len > ahs_longest) or (a_wid > ahs_second) or
                    (a_cubic > ahs_cubic) or (a_length_girth > ahs_lg_limit))

                # AHS minimum billable weight applies only when the package
                # is billed on ACTUAL weight. When dimensional weight governs
                # (Billed Weight Type 9) UPS does not lift the weight to the
                # AHS minimum -- verified against every AHS-triggering package
                # on this invoice. The surcharge fee is still charged either
                # way; only the weight floor is skipped.
                # Type 9 = dimensional weight governs; type 8 = SCC re-basis
                # onto the corrected scale weight. In neither case does UPS
                # lift the billed weight to the AHS minimum.
                _ahs_min_applies = (billed_weight_type not in (8, 9))
                if (ahs_weight_trigger or ahs_dimension_trigger) and _ahs_min_applies:
                    minimum_billable = max(minimum_billable, ahs_min)

                # ---- Build human-readable trigger reasons (full names) ----
                lps_reasons = []
                if lps_longest_trigger:
                    lps_reasons.append("Large Package Surcharge - Longest Side")
                if lps_lg_trigger:
                    lps_reasons.append(
                        "Large Package Surcharge - Length + Girth")
                if lps_cubic_trigger:
                    lps_reasons.append(
                        "Large Package Surcharge - Cubic Volume")
                if lps_weight_trigger:
                    lps_reasons.append("Large Package Surcharge - Weight")
                lps_reason_text = "; ".join(lps_reasons)

                ahs_reasons = []
                if ahs_weight_trigger:
                    ahs_reasons.append("Additional Handling Surcharge - Weight")
                if length > ahs_longest:
                    ahs_reasons.append(
                        "Additional Handling Surcharge - Longest Side")
                if width > ahs_second:
                    ahs_reasons.append(
                        "Additional Handling Surcharge - Second Longest Side")
                if length_girth > ahs_lg_limit:
                    ahs_reasons.append(
                        "Additional Handling Surcharge - Length + Girth")
                if cubic_inches > ahs_cubic:
                    ahs_reasons.append(
                        "Additional Handling Surcharge - Cubic Volume")
                ahs_reason_text = "; ".join(ahs_reasons)

                # Billed Weight Type gate: UPS only applies dimensional weight
                # when the basis code says so (9). A type-8 row is an SCC
                # re-basis onto the corrected/scale weight, and type 2 is
                # actual weight -- in both cases dim weight must NOT be
                # applied. Unknown/absent type falls back to the old
                # behaviour (apply dim weight) so nothing regresses.
                _dim_applies = billed_weight_type not in (2, 8)
                _eff_dim_weight = dim_weight if _dim_applies else 0.0

                billable_weight = math.ceil(max(actual_weight, _eff_dim_weight,
                                                minimum_billable,
                                                minimum_billable_floor))

                if (ahs_weight_trigger or ahs_dimension_trigger) and _ahs_min_applies \
                        and billable_weight < ahs_min:
                    billable_weight = math.ceil(ahs_min)

                # =====================================================
                # RECLASSIFICATION PRE-SCAN
                # Driver = ADJ FRT charge code: COM -> Commercial, RES -> Residential.
                #   RES adjustment = was Commercial, corrected to Residential
                #   COM adjustment = was Residential, corrected to Commercial
                # orig_service = service for A (original); reclass_service = C.
                # =====================================================
                # A "Shipping Charge Correction Ground" ADJ carries no RES/COM
                # signal, so normalize_service() defaults it to Commercial. That
                # is wrong: an SCC only re-bills / refunds the base rate, it does
                # NOT change the shipping channel. When the original SHP/RTN for
                # the same tracking is in this invoice, inherit its service.
                # (A genuine COM/RES reclassification ADJ is detected below and
                # overrides this.)
                if layer == "ADJ" and tracking in shp_rtn_service:
                    _orig_svc = shp_rtn_service.get(tracking, "")
                    if _orig_svc:
                        shipment_type = _orig_svc
                        # Inherited from the shipment's own SHP/RTN line, so
                        # it is the real service -- not the dropdown guess.
                        # Leaving the flag set marked the lane unsupported and
                        # replaced Shipment Type with the raw ADJ wording.
                        service_was_guessed = False

                orig_service = shipment_type
                reclass_service = shipment_type
                if layer == "ADJ":
                    for _r in rows:
                        _ar = str(safe_get(_r, AR_COL, "")).strip().upper()
                        _as = str(safe_get(_r, AS_COL, "")).strip().upper()
                        _at = str(safe_get(_r, AT_COL, "")).strip().upper()
                        if _ar == "FRT" and (_as == "COM" or "COMMERCIAL ADJUSTMENT" in _at):
                            reclass_service = "Ground Commercial"
                            orig_service = "Ground Residential"   # was residential
                            break
                        if _ar == "FRT" and (_as == "RES" or "RESIDENTIAL ADJUSTMENT" in _at):
                            reclass_service = "Ground Residential"
                            orig_service = "Ground Commercial"     # was commercial
                            break

                # Cross-period pure ADJ: when the original SHP isn't in this
                # invoice and the description was the generic "SCC Ground"
                # (defaults to Commercial), infer service from the ADJ's own
                # residential/commercial accessorial codes (LPR/RDR/LDR/CAR ->
                # residential; LPS/RDC/LDC/CAC -> commercial). Interim rule until
                # cross-period lookup exists. Only when not reclassified.
                if (layer == "ADJ" and orig_service == reclass_service
                        and tracking not in shp_rtn_freight_A):
                    pre_res = pre_com = False
                    for _r in rows:
                        _as = str(safe_get(_r, AS_COL, "")).strip().upper()
                        if _as in ACC_RESIDENTIAL_CODES:
                            pre_res = True
                        elif _as in ACC_COMMERCIAL_CODES:
                            pre_com = True
                    if pre_res and not pre_com:
                        orig_service = reclass_service = "Ground Residential"
                        shipment_type = "Ground Residential"
                    elif pre_com and not pre_res:
                        orig_service = reclass_service = "Ground Commercial"
                        shipment_type = "Ground Commercial"

                # =====================================================
                # A / C FREIGHT (TWO-PASS BASE RATE)
                #   freight_A: HR dims + AA (entered) weight, ORIGINAL service
                #   freight_C: AG dims + AC (billed) weight, RECLASSIFIED service
                #   v2: ADJ base rate = freight_C (corrected), not a difference.
                # =====================================================
                def freight_for(dims, weight, tag, svc, ahs_floor, lps_floor):
                    cu = cubic(dims)
                    dw = round(cu / dim_factor, 2) if dim_factor > 0 else 0.0
                    # Per-package minimum billable: LPS forces its min (e.g. 90),
                    # else AHS forces its min (e.g. 40). Each package (A vs C)
                    # uses ITS OWN triggers, not a shared value.
                    pkg_min = 0.0
                    if lps_floor:
                        pkg_min = max(pkg_min, oversize_min)
                    elif ahs_floor:
                        pkg_min = max(pkg_min, ahs_min)
                    bw = math.ceil(max(weight, dw, pkg_min))
                    base, warn = lookup_base_rate(svc, zone, bw, tracking)
                    if warn:
                        rate_notes.append(f"{tag}: weight {bw} not found ({svc}, zone {zone})")
                    return base, bw

                rate_notes = []
                if adj_skip_base:
                    # ADJ with no Package Dimension (AG) + no Detail Keyed Dim
                    # (HR), OR with no entered weight (AA) + no billed weight
                    # (AC): no base rate lookup at all.
                    freight_A = freight_C = 0.0
                    billable_A = billable_C = billable_weight
                else:
                    # A package: HR dims, original service, A-package triggers.
                    freight_A, billable_A = freight_for(
                        dim_A, weight_A, "A", orig_service,
                        (a_ahs_weight_trigger or a_ahs_dimension_trigger),
                        a_lps_trigger)
                    # C package: AG dims, reclassified service, C-package triggers.
                    freight_C, billable_C = freight_for(
                        dim_C, weight_C, "C", reclass_service,
                        (ahs_weight_trigger or ahs_dimension_trigger),
                        large_package_trigger)

                pure_adj_both_dims = False
                if layer == "ADJ":
                    matched = tracking in shp_rtn_freight_A
                    # Pure ADJ (no SHP/RTN this invoice) that carries BOTH its
                    # own AG and HR: this is a prior-period correction where HR
                    # is the invisible original (A) and AG is the corrected (C).
                    # The adjustment = C - A.
                    #
                    # Guard: UPS stamps a 1x1x1 placeholder in the HR (or AG)
                    # column when there is no real prior dimension. A 1x1x1 "A
                    # package" is not a genuine original to subtract -- treating
                    # it as both-dims makes the tool compute C - A and (worse)
                    # can suppress size surcharges entirely. Require BOTH dims to
                    # be real parcels (longest side > 1 inch) before doing C - A.
                    def _real_parcel(d):
                        return max(d) > 1.0 and cubic(d) > 1.0
                    pure_adj_both_dims = (
                        not matched
                        and _real_parcel(dim_A)
                        and _real_parcel(dim_C))
                    if adj_no_own_dims:
                        cal_base = 0.0
                        notes_match = "No ADJ dimension (AG & HR = 0) - base rate skipped"
                    elif adj_no_own_weight:
                        cal_base = 0.0
                        notes_match = "No ADJ weight (AA & AC = 0) - base rate skipped"
                    elif pure_adj_both_dims:
                        cal_base = round(freight_C - freight_A, 2)
                        notes_match = "pure ADJ C-A (both dims)"
                    elif (not matched
                          and max_entered_weight > 0 and max_billed_weight > 0
                          and abs(max_billed_weight - max_entered_weight) > 0.001):
                        # Pure ADJ (no SHP/RTN this invoice) with a 1x1x1
                        # placeholder in HR, but carrying BOTH an original entered
                        # weight (AA) and a corrected/billed weight (AC). This is
                        # a dimensional/weight re-rate: UPS re-measured the parcel
                        # and bills the DIFFERENCE between the corrected charge and
                        # what was originally paid -- not the full corrected rate.
                        #
                        # Use the RAW entered/billed weights here, NOT weight_C /
                        # weight_A: when the billed weight equals the AHS minimum
                        # (e.g. 40 lb), an earlier guard rewrites weight_C down to
                        # the real weight, which would collapse the delta to 0 and
                        # fall through to a full-rate charge. Both legs use the
                        # corrected (AG) dimensions and the same AHS/LPS floors --
                        # only the weight differs -- so the delta is purely the
                        # weight re-rate, matching the invoice's FRT line.
                        # The CORRECTED leg uses the AG (re-measured) dimensions
                        # and billed weight, with the AHS/LPS floors the
                        # re-measurement triggered. The ORIGINAL leg uses the
                        # ORIGINAL dimensions (HR -- typically the 1x1x1
                        # placeholder, i.e. no effective size) and the entered
                        # weight, with NO floors. This matters for type-9
                        # (dim-weight-governed) corrections: the corrected AG size
                        # produces a large dim weight, but the ORIGINAL charge was
                        # based only on the entered actual weight -- so the delta
                        # is rate(corrected dim/billed) - rate(original entered),
                        # matching the invoice FRT line. Using AG dims on the
                        # original leg would inflate its dim weight to the
                        # corrected value and collapse the delta to 0.
                        _c_floor_ahs = ahs_weight_trigger or ahs_dimension_trigger
                        freight_hi, _ = freight_for(
                            dim_C, max_billed_weight, "C", reclass_service,
                            _c_floor_ahs, large_package_trigger)
                        freight_lo, _ = freight_for(
                            dim_A, max_entered_weight, "A", reclass_service,
                            False, False)
                        cal_base = round(freight_hi - freight_lo, 2)
                        notes_match = "pure ADJ weight re-rate (billed-entered delta)"
                    else:
                        cal_base = freight_C
                        notes_match = ("corrected (matched SHP/RTN)" if matched
                                       else "corrected (no SHP/RTN match)")
                    billable_weight = billable_C
                else:
                    cal_base = freight_A
                    billable_weight = billable_A
                    notes_match = ""

                # ----- Build Notes column (warnings for manual review) -----
                notes = []

                # UNSUPPORTED SERVICE / ZONE GUARD
                # The rate tables only cover US domestic zones (2-8, 44-46).
                # International services such as "Standard to Canada" ship on
                # zone 53 and have their own tariff, which this tool does not
                # hold. Silently falling back to Ground Commercial / zone 8
                # produces a plausible-looking but completely wrong figure, so
                # flag the row loudly instead and let the user handle it
                # manually.
                _supported_zones = {"2", "3", "4", "5", "6", "7", "8",
                                    "44", "45", "46"}
                _raw_desc = str(shipment_description).strip()
                _desc_l = _raw_desc.lower()
                _intl_markers = ("to canada", "canada", "worldwide",
                                 "international", "export", "import")
                _is_intl_desc = any(m in _desc_l for m in _intl_markers)
                _zone_unsupported = (str(zone).strip() not in _supported_zones)

                # Hundredweight (Ground / Air CWT) is rated on the TOTAL weight
                # of the whole multi-package shipment at a per-hundredweight
                # rate, not per parcel. normalize_service has no pattern for
                # it, so it used to fall through to base_rate_service.get() --
                # meaning the service was decided by whatever happened to be
                # selected in the Import Files dropdown, and each package was
                # then repriced as an ordinary parcel. That produces a large
                # phantom overcharge. Flag it instead.
                _is_hundredweight = ("hundredweight" in _desc_l
                                     or "cwt" in _desc_l.split())

                _unsupported_lane = bool(_zone_unsupported or _is_intl_desc
                                         or _is_hundredweight)

                if _unsupported_lane:
                    _bits = []
                    if _raw_desc:
                        _bits.append(_raw_desc)
                    if _zone_unsupported:
                        _bits.append(f"zone {zone}")
                    if _is_hundredweight:
                        _bits.append("hundredweight - rated on total shipment "
                                     "weight, not per parcel")
                    # The row is not rated, so do not present a service that
                    # was only guessed from the dropdown as if it were real.
                    if service_was_guessed:
                        shipment_type = (str(shipment_description).strip()
                                         or shipment_type)
                    notes.append(
                        "*** UNSUPPORTED SERVICE/ZONE - NOT RATED ("
                        + ", ".join(_bits)
                        + ") - no rate table for this lane; verify manually ***")

                if service_was_guessed and not _unsupported_lane:
                    notes.append(
                        f"Service '{str(shipment_description).strip()}' not "
                        f"recognised - defaulted to '{shipment_type}' from the "
                        f"Import Files dropdown; verify")

                # Dimension-missing flags
                if layer == "ADJ":
                    # HR only matters for A when A is recomputed (no match)
                    if not matched and cubic(dim_A) <= 0:
                        notes.append("No HR dimension (A)")
                    if cubic(dim_C) <= 0:
                        notes.append("No AG dimension (C)")
                    if notes_match:
                        notes.append(notes_match)
                else:
                    if cubic(dim_A) <= 0:
                        notes.append("No HR dimension")
                # Rate-not-found flags
                notes.extend(rate_notes)
                row_result["Notes"] = "; ".join(notes)

                rate_warning = ""

                # Fuel % follows the SHIP DATE, not the invoice. UPS republishes
                # the percentage weekly, so an invoice that spans a rate change
                # needs a different percent per shipment.
                try:
                    default_fuel_percent = float(self.fuel_percent.get())
                except (ValueError, TypeError):
                    default_fuel_percent = 0.0

                fuel_pct, fuel_note = fuel_percent_for_date(
                    getattr(self, "fuel_schedule", []),
                    row_result.get("Date", ""),
                    default_fuel_percent
                )

                if fuel_note:
                    notes.append(fuel_note)
                    row_result["Notes"] = "; ".join(notes)

                cal_fuel = 0

                # Start final static columns
                row_result["Zone"] = zone
                row_result["Shipment Type"] = shipment_type
                # ---- diagnostics (help pinpoint calc issues) ----
                row_result["_dbg_reclass_service"] = reclass_service
                row_result["_dbg_orig_service"] = orig_service
                row_result["_dbg_freight_A"] = round(freight_A, 2)
                row_result["_dbg_freight_C"] = round(freight_C, 2)
                row_result["_dbg_billable_A"] = billable_A
                row_result["_dbg_billable_C"] = billable_C
                row_result["_dbg_pure_ca"] = 1 if pure_adj_both_dims else 0
                row_result["_dbg_dimA"] = f"{dim_A[0]}x{dim_A[1]}x{dim_A[2]}"
                row_result["_dbg_dimC"] = f"{dim_C[0]}x{dim_C[1]}x{dim_C[2]}"
                row_result["Entered Weight"] = round(max_entered_weight, 2)
                row_result["Billed Weight"] = round(max_billed_weight, 2)
                row_result["Dim Weight"] = round(dim_weight, 2)
                row_result["Billable Weight"] = billable_weight
                row_result["Length"] = round(length, 2)
                row_result["Width"] = round(width, 2)
                row_result["Height"] = round(height, 2)
                row_result["New Weight"] = billable_weight
                row_result["Length+Girth"] = round(length_girth, 2)
                row_result["VOL"] = round(cubic_inches, 2)
                # LPS and AHS are mutually exclusive (LPS wins). Write the
                # full reason name, not just a flag.
                if large_package_trigger:
                    row_result["LPS"] = lps_reason_text or "Large Package Surcharge"
                    row_result["AHS"] = ""
                elif ahs_weight_trigger or ahs_dimension_trigger:
                    row_result["AHS"] = (ahs_reason_text or
                                         "Additional Handling Surcharge")
                    row_result["LPS"] = ""
                else:
                    row_result["AHS"] = ""
                    row_result["LPS"] = ""
                row_result["CAL Base Rate"] = round(cal_base, 2)
                row_result["_has_freight"] = 1 if _layer_has_freight else 0
                row_result["CAL Fuel"] = round(cal_fuel, 2)
                row_result["Shipping Charge Correction Ground Undeliverable Return"] = 0
                row_result["Residential Adjustment"] = 0
                row_result["Commercial Adjustment"] = 0
                row_result["Direct Delivery Only"] = 0
                row_result["Shipping Charge Correction Large Package Surcharge"] = 0

                ups_total_for_tracking = 0.0

                # =====================================================
                # AUDIT FLAGS
                # =====================================================
                ups_charged_ahs_weight = False
                ups_charged_ahs_dimension = False
                ups_charged_ahs_packaging = False
                ups_charged_lps = False
                ups_charged_ovr = False

                system_should_lps = large_package_trigger
                system_should_ovr = False

                # Raw surcharge consolidation
                raw_residential_net = 0.0
                raw_das_com_net = 0.0
                raw_das_res_net = 0.0
                raw_das_ext_com_net = 0.0
                raw_das_ext_res_net = 0.0
                raw_remote_com_net = 0.0
                raw_remote_res_net = 0.0
                raw_remote_ak_net = 0.0
                raw_remote_hi_net = 0.0
                raw_declared_value_net = 0.0
                raw_rts_net = 0.0
                raw_signature_net = 0.0
                raw_adult_signature_net = 0.0
                raw_package_protection_net = 0.0
                raw_reroute_seen = False
                raw_reschedule_seen = False
                raw_ret_elabel_seen = False
                raw_ret_plabel_seen = False

                raw_residential_seen = False
                raw_das_com_seen = False
                raw_das_res_seen = False
                raw_das_ext_com_seen = False
                raw_das_ext_res_seen = False
                raw_remote_com_seen = False
                raw_remote_res_seen = False
                raw_remote_ak_seen = False
                raw_remote_hi_seen = False
                # Residential/commercial code presence (for inferring service of
                # a cross-period pure ADJ whose original SHP isn't in this file).
                res_code_seen = False
                com_code_seen = False

                raw_ahs_weight_net = 0.0
                raw_ahs_dimension_net = 0.0
                raw_ahs_packaging_net = 0.0
                raw_lps_net = 0.0
                raw_ovr_net = 0.0

                undeliverable_return_net = 0.0
                undeliverable_return_seen = False

                address_correction_net = 0.0
                address_correction_seen = False

                residential_adjustment_net = 0.0
                residential_adjustment_seen = False
                commercial_adjustment_seen = False
                shipment_original_service = shipment_type

                # v2.6 dynamic correction source rows
                scc_ground_undeliverable_seen = False
                scc_source_entered_weight = 0.0
                scc_source_billed_weight = 0.0
                scc_source_service = shipment_type
                scc_source_zone = zone

                skip_2nd_mile = False

                dynamic_surcharge_seen_codes = set()
                dynamic_surcharge_seen_amounts = {}
                dynamic_surcharge_code_matched = {}

                # =====================================================
                # PROCESS RAW ROWS
                # =====================================================
                for row in rows:
                    ai = str(safe_get(row, AI_COL, "")).strip().upper()
                    aj = str(safe_get(row, AJ_COL, "")).strip().upper()
                    ak = str(safe_get(row, AK_COL, "")).strip().upper()
                    ar = str(safe_get(row, AR_COL, "")).strip().upper()
                    as_code = str(safe_get(row, AS_COL, "")).strip().upper()
                    at = str(safe_get(row, AT_COL, "")).strip()
                    at_upper = at.upper()
                    ba = self.to_amount(safe_get(row, BA_COL, 0))

                    # Reclassification (COM/RES Adjustment) rows frequently carry
                    # net = 0. Detect them BEFORE the zero-net skip below, or the
                    # service switch is missed and residential is double-charged.
                    if ai == "ADJ" and ar == "FRT":
                        if "RESIDENTIAL ADJUSTMENT" in at_upper or as_code == "RES":
                            residential_adjustment_seen = True
                        elif "COMMERCIAL ADJUSTMENT" in at_upper or as_code == "COM":
                            commercial_adjustment_seen = True

                    # A REPRICE custom surcharge is priced from the contract
                    # table, not from what UPS billed -- so it must be
                    # registered even when UPS shows 0.00 (waived lines,
                    # marker rows). It used to fall to the zero-net skip
                    # below, so a code with a rate filled in was silently
                    # worth nothing. Registered here, priced later.
                    if (ar in BILLABLE_AR_CODES
                            and as_code in self.dynamic_surcharge_mapping):
                        _dyn_cfg = self.dynamic_surcharge_mapping.get(as_code)
                        if (isinstance(_dyn_cfg, dict)
                                and _dyn_cfg.get("enabled", True)
                                and normalize_dynamic_type(
                                    _dyn_cfg.get("type", "REPRICE")) == "REPRICE"):
                            dynamic_surcharge_seen_codes.add(as_code)

                    if ba == 0:
                        continue

                    # Account-level money never belongs to a parcel, even when
                    # UPS put a tracking number on the row. Counting it here
                    # wrecked that parcel's UPS total and its margin.
                    if ai == "MIS" or ar == "MSC":
                        continue

                    ups_total_for_tracking += ba

                    # FIX #1: DDO - Delete else: ddo_charge = 0
                    if str(as_code).strip().upper() == "DDO":
                        try:
                            ddo_charge = float(
                                self.accessorial_rate_table.get(
                                    "Direct Delivery Only",
                                    0
                                )
                            )
                        except (ValueError, TypeError):
                            ddo_charge = 0

                    # v2.6: Shipping Charge Correction Ground Undeliverable Return
                    # Do NOT use UPS BA. Calculate by Base Rate Template:
                    # Billed Weight Base Rate - Entered Weight Base Rate.
                    if ai == "ADJ" and ar == "FRT" and "SHIPPING CHARGE CORRECTION" in at_upper and "UNDELIVERABLE" in at_upper:
                        scc_ground_undeliverable_seen = True
                        # Fallback to ADJ row values only if no RTN row is available later.
                        if scc_source_entered_weight <= 0:
                            scc_source_entered_weight = self.to_amount(safe_get(row, ENTERED_WEIGHT_COL, 0))
                        if scc_source_billed_weight <= 0:
                            scc_source_billed_weight = self.to_amount(safe_get(row, BILLED_WEIGHT_COL, 0))
                        continue

                    # RTN - Ground Undeliverable Return
                    elif ai == "RTN" and ar == "FRT" and "GROUND UNDELIVERABLE RETURN" in at_upper:
                        undeliverable_return_seen = True
                        undeliverable_return_net += ba
                        # When RTN + ADJ SCC both exist, SCC must use this RTN row AA/AC.
                        scc_source_entered_weight = self.to_amount(safe_get(row, ENTERED_WEIGHT_COL, 0))
                        scc_source_billed_weight = self.to_amount(safe_get(row, BILLED_WEIGHT_COL, 0))
                        scc_source_service = normalize_service(at, ai)
                        scc_source_zone = normalize_zone(safe_get(row, ZONE_COL, "")) or zone

                    # ADC - Address Correction
                    elif ai == "ADJ" and aj == "ADC" and ar == "FRT":
                        address_correction_seen = True
                        address_correction_net += ba

                    # FIX #2: Residential / Commercial Adjustment - NO max(..., 0)
                    elif ai == "ADJ" and ar == "FRT" and "RESIDENTIAL ADJUSTMENT" in at_upper:
                        residential_adjustment_seen = True
                        residential_adjustment_net += ba

                    elif ai == "ADJ" and ar == "FRT" and ("COMMERCIAL ADJUSTMENT" in at_upper or as_code == "COM"):
                        commercial_adjustment_seen = True

                    elif ai == "ADJ" and aj == "DIN" and ar == "FRT":
                        skip_2nd_mile = True
                        continue

                    elif as_code == "FRT" or as_code == "FSC":
                        continue

                    elif as_code == "RES":
                        raw_residential_seen = True
                        raw_residential_net += ba
                        res_code_seen = True

                    elif as_code == "RDC":
                        raw_das_com_seen = True
                        raw_das_com_net += ba
                        com_code_seen = True

                    elif as_code == "RDR":
                        raw_das_res_seen = True
                        raw_das_res_net += ba
                        res_code_seen = True

                    elif as_code == "LDC":
                        raw_das_ext_com_seen = True
                        raw_das_ext_com_net += ba
                        com_code_seen = True

                    elif as_code == "LDR":
                        raw_das_ext_res_seen = True
                        raw_das_ext_res_net += ba
                        res_code_seen = True

                    elif as_code == "CAC":
                        raw_remote_com_seen = True
                        raw_remote_com_net += ba
                        com_code_seen = True

                    elif as_code == "CAR":
                        raw_remote_res_seen = True
                        raw_remote_res_net += ba
                        res_code_seen = True

                    elif as_code == "HIC":
                        # Hawaii remote area, commercial. Same surcharge and same
                        # rate row as CAC per the ACC Codes table.
                        raw_remote_com_seen = True
                        raw_remote_com_net += ba
                        com_code_seen = True

                    elif as_code == "HIR":
                        # Hawaii remote area, residential -> same row as CAR.
                        raw_remote_res_seen = True
                        raw_remote_res_net += ba
                        res_code_seen = True

                    elif as_code == "AKR":
                        # Alaska remote area. Priced from its OWN rate row
                        # ("Remote Area - AK"), NOT the 48-state remote rate --
                        # UPS charges Alaska materially higher.
                        raw_remote_ak_seen = True
                        raw_remote_ak_net += ba

                    elif as_code in self.dynamic_surcharge_code_map:
                        linked_cfg = self.dynamic_surcharge_code_map[as_code]
                        surcharge_name = linked_cfg.get("surcharge", "").strip()

                        if surcharge_name != "":
                            dynamic_surcharge_code_matched[surcharge_name] = ba

                    elif as_code in self.ahs_lps_code_registry:
                        reg_cfg = self.ahs_lps_code_registry[as_code]
                        charge_type = reg_cfg.get("charge_type", "AHS - Weight")

                        if "Weight" in charge_type:
                            ups_charged_ahs_weight = True
                            raw_ahs_weight_net += ba
                        elif "Dimension" in charge_type:
                            ups_charged_ahs_dimension = True
                            raw_ahs_dimension_net += ba
                        elif "Packaging" in charge_type:
                            ups_charged_ahs_packaging = True
                            raw_ahs_packaging_net += ba
                        elif "Large Package" in charge_type:
                            ups_charged_lps = True
                            raw_lps_net += ba

                        # The registry branch runs BEFORE the LPS/LPR branch
                        # below, so the residential / commercial signal carried
                        # by the code (LPR/LCR/LWR vs LPS/LCT/LCC/LWC) must be
                        # recorded here too, otherwise a registry-matched code
                        # silently loses it.
                        if as_code in ACC_RESIDENTIAL_CODES:
                            res_code_seen = True
                        elif as_code in ACC_COMMERCIAL_CODES:
                            com_code_seen = True

                    elif as_code == "AHW":
                        # AHW = Additional Handling - Weight (actual weight over
                        # the 50 lb threshold).
                        ups_charged_ahs_weight = True
                        raw_ahs_weight_net += ba

                    elif as_code in ["AHG", "AHL", "AHD", "AHB", "AHS"]:
                        # Dimension-based Additional Handling codes:
                        #   AHG = Length+Girth over 105"
                        #   AHL = Longest Side over 48"
                        #   AHS = Second Longest Side over 30"
                        #   AHD / AHB = other dimension bases
                        # NOTE: AHS is "Addl. Handling second long. Side" -- a
                        # DIMENSION criterion, not weight. UPS lists the second
                        # longest side alongside longest-side and length+girth as
                        # the three dimension triggers; weight is a separate rule.
                        ups_charged_ahs_dimension = True
                        raw_ahs_dimension_net += ba

                    elif as_code in ["AHP", "AHC"]:
                        ups_charged_ahs_packaging = True
                        raw_ahs_packaging_net += ba

                    elif ar in ["LPS", "LPR", "LCT", "LCR"] or as_code in ["LPS", "LPR", "LCT", "LCR"]:
                        ups_charged_lps = True
                        raw_lps_net += ba
                        if as_code in ("LPR", "LCR"):
                            res_code_seen = True
                        elif as_code in ("LPS", "LCT"):
                            com_code_seen = True

                    elif as_code in ["OVR", "OML"]:
                        # OML is the Over Maximum Limit variant of OVR.
                        ups_charged_ovr = True
                        raw_ovr_net += ba

                    elif as_code == "EVS":
                        raw_declared_value_net += ba

                    elif as_code == "ISW":
                        raw_rts_net += ba

                    elif as_code in ["SIG", "DCS", "CSG"]:
                        # DCS = Delivery Confirmation Signature,
                        # CSG = the commercial variant. Both are the signature
                        # surcharge, priced from the "Signature" rate row.
                        raw_signature_net += ba

                    elif as_code in ["ASG", "ADS"]:
                        # ADS is the code UPS uses on these invoices for Adult
                        # Signature Required.
                        raw_adult_signature_net += ba

                    elif as_code == "SHD":
                        raw_package_protection_net += ba

                    elif as_code == "IRW":
                        raw_reroute_seen = True

                    elif as_code == "IFW":
                        raw_reschedule_seen = True

                    elif as_code == "ERL":
                        raw_ret_elabel_seen = True

                    elif as_code == "ALP":
                        raw_ret_plabel_seen = True

                    elif as_code == "FTP":
                        # Third Party Billing Service marker (net 0); known,
                        # not a billable surcharge.
                        pass

                    elif ar in BILLABLE_AR_CODES and as_code in self.dynamic_surcharge_mapping:
                        dynamic_surcharge_seen_codes.add(as_code)
                        # COPY INVOICE / REPORT ONLY need the billed figure.
                        dynamic_surcharge_seen_amounts[as_code] = (
                            dynamic_surcharge_seen_amounts.get(as_code, 0.0) + ba)

                    else:
                        handled_raw_codes = {
                            "FRT", "FSC", "RES", "RDC", "RDR", "LDC", "LDR",
                            "CAC", "CAR", "AHG", "AHL", "AHS", "AHW", "AHD", "AHB", "AHP", "AHC",
                            "LPS", "LPR", "LCT", "LCR", "OVR", "EVS", "ISW", "SIG", "ASG", "SHD", "DDO",
                            "IRW", "IFW", "ERL", "ALP", "FTP",
                        }

                        # Every AR=ACC code in the master ACC Codes table counts
                        # as known, so only genuinely NEW UPS codes surface in
                        # the Unknown report.
                        handled_raw_codes.update(UPS_ACC_CODE_REFERENCE.keys())

                        handled_raw_codes.update(self.dynamic_surcharge_mapping.keys())
                        handled_raw_codes.update(self.ahs_lps_code_registry.keys())
                        handled_raw_codes.update(self.dynamic_surcharge_code_map.keys())

                        # Unknown-code check now covers BRK / GOV / MSC too --
                        # it used to look at ACC only, so an unregistered
                        # brokerage or duty code was silently unpriced AND
                        # silently unreported.
                        if ar in BILLABLE_AR_CODES and as_code not in handled_raw_codes:
                            unknown_rows.append({
                                "Tracking": tracking,
                                "AT Description": at if at else as_code,
                                "AI Type Code 1": ai,
                                "AS Charge Code": as_code,
                                "AR Charge Classification": ar,
                                "BA Net Amount": round(ba, 2)
                            })

                # =====================================================
                # FINAL OVR DECISION
                # =====================================================
                system_should_ovr = raw_ovr_net != 0

                # =====================================================
                # RECLASSIFICATION (per spec): the ADJ FRT charge code is the
                # authoritative signal for service reclassification.
                #   COM (Commercial Adjustment) -> Ground Commercial
                #   RES (Residential Adjustment) -> Ground Residential
                # Set shipment_type here so ALL downstream surcharge lookups
                # (AHS/LPS/DAS/Residential) use the post-reclassification table.
                # =====================================================
                reclassified = False
                if commercial_adjustment_seen:
                    shipment_type = "Ground Commercial"
                    row_result["Shipment Type"] = shipment_type
                    reclassified = True
                elif residential_adjustment_seen:
                    shipment_type = "Ground Residential"
                    row_result["Shipment Type"] = shipment_type
                    reclassified = True
                # Marker for Final aggregation: scenario A (reclass) vs B.
                row_result["_reclassified"] = 1 if reclassified else 0

                # =====================================================
                # RESIDENTIAL CORRECTION CONSOLIDATION
                # =====================================================
                residential_net = round(raw_residential_net, 2)

                if reclassified and commercial_adjustment_seen:
                    if pure_adj_layer:
                        # Pure ADJ = standalone correction document. Originally
                        # billed Residential, corrected to Commercial, so the
                        # Residential Surcharge is REFUNDED: negative template
                        # rate (e.g. 2.6 -> -2.6). Use the residential table.
                        row_result["Residential"] = -accessorial_lookup(
                            "Residential", "Ground Residential", zone)
                    else:
                        # Reclassified to commercial: residential surcharge is removed.
                        row_result["Residential"] = 0
                elif reclassified and residential_adjustment_seen:
                    if pure_adj_layer:
                        # Mirror image: originally Commercial, corrected to
                        # Residential, so the Residential Surcharge is CHARGED.
                        row_result["Residential"] = accessorial_lookup(
                            "Residential", "Ground Residential", zone)
                    # Reclassified to residential: charge the residential
                    # surcharge ONLY if the invoice actually carried it. If the
                    # original never had a residential surcharge, don't invent
                    # one just because the service became residential.
                    elif raw_residential_seen:
                        row_result["Residential"] = accessorial_lookup("Residential", shipment_type, zone)
                    else:
                        row_result["Residential"] = 0
                elif raw_residential_seen and residential_net == 0:
                    row_result["Residential"] = 0
                    shipment_type = "Ground Commercial"
                    row_result["Shipment Type"] = shipment_type

                elif residential_net != 0:
                    row_result["Residential"] = accessorial_lookup("Residential", shipment_type, zone)
                else:
                    row_result["Residential"] = 0

                # =====================================================
                # DAS / RAS / REMOTE AREA CORRECTION CONSOLIDATION
                # =====================================================
                das_com_net = round(raw_das_com_net, 2)
                das_res_net = round(raw_das_res_net, 2)
                any_das_seen = (raw_das_res_seen or raw_das_com_seen)

                if reclassified and commercial_adjustment_seen:
                    # Reclassified to commercial.
                    row_result["DAS Residential"] = 0
                    if pure_adj_layer:
                        # Pure ADJ = standalone correction document. Originally
                        # billed Residential, corrected to Commercial. The
                        # residential DAS that was charged is now REFUNDED, and
                        # (if the shipment is still in a DAS area) the commercial
                        # DAS applies instead. Repriced from YOUR rate table
                        # (not copied from the invoice): net = commercial DAS
                        # rate - residential DAS rate. Since residential DAS
                        # (1.97) > commercial DAS (1.35), this nets negative,
                        # matching the invoice's DAS refund. Mirrors the
                        # Residential Surcharge refund logic above.
                        if any_das_seen:
                            row_result["DAS Commercial"] = round(
                                accessorial_lookup("DAS Commercial", "Ground Commercial", zone)
                                - accessorial_lookup("DAS Residential", "Ground Residential", zone), 2)
                    elif any_das_seen:
                        # Matched SHP/RTN reprice: charge full commercial DAS.
                        row_result["DAS Commercial"] = accessorial_lookup("DAS Commercial", shipment_type, zone)
                elif reclassified and residential_adjustment_seen:
                    # Reclassified to residential: commercial DAS removed,
                    # residential DAS charged.
                    row_result["DAS Commercial"] = 0
                    if pure_adj_layer:
                        # Mirror image: originally Commercial, corrected to
                        # Residential. Repriced net = residential DAS rate -
                        # commercial DAS rate (nets positive, a charge).
                        if any_das_seen:
                            row_result["DAS Residential"] = round(
                                accessorial_lookup("DAS Residential", "Ground Residential", zone)
                                - accessorial_lookup("DAS Commercial", "Ground Commercial", zone), 2)
                    elif any_das_seen:
                        row_result["DAS Residential"] = accessorial_lookup("DAS Residential", shipment_type, zone)
                elif raw_das_res_seen and raw_das_com_seen and das_res_net > 0 and das_com_net < 0:
                    shipment_type = "Ground Commercial"
                    row_result["Shipment Type"] = shipment_type
                    row_result["DAS Residential"] = 0
                    row_result["DAS Commercial"] = accessorial_lookup("DAS Commercial", shipment_type, zone)
                else:
                    if raw_das_com_seen and das_com_net != 0:
                        row_result["DAS Commercial"] = accessorial_lookup("DAS Commercial", shipment_type, zone)

                    if raw_das_res_seen and das_res_net != 0:
                        row_result["DAS Residential"] = accessorial_lookup("DAS Residential", shipment_type, zone)

                any_das_ext_seen = (raw_das_ext_com_seen or raw_das_ext_res_seen)
                if reclassified and commercial_adjustment_seen:
                    row_result["DAS Extended Residential"] = 0
                    if any_das_ext_seen:
                        row_result["DAS Extended Commercial"] = accessorial_lookup("DAS Extended Commercial", shipment_type, zone)
                elif reclassified and residential_adjustment_seen:
                    row_result["DAS Extended Commercial"] = 0
                    if any_das_ext_seen:
                        row_result["DAS Extended Residential"] = accessorial_lookup("DAS Extended Residential", shipment_type, zone)
                else:
                    if raw_das_ext_com_seen and round(raw_das_ext_com_net, 2) != 0:
                        row_result["DAS Extended Commercial"] = accessorial_lookup("DAS Extended Commercial", shipment_type, zone)

                    if raw_das_ext_res_seen and round(raw_das_ext_res_net, 2) != 0:
                        row_result["DAS Extended Residential"] = accessorial_lookup("DAS Extended Residential", shipment_type, zone)

                if raw_remote_com_seen and round(raw_remote_com_net, 2) != 0:
                    row_result["Remote Area Commercial"] = accessorial_lookup("Remote Area Commercial", shipment_type, zone)

                if raw_remote_res_seen and round(raw_remote_res_net, 2) != 0:
                    row_result["Remote Area Residential"] = accessorial_lookup("Remote Area Residential", shipment_type, zone)

                # Alaska (AKR) is applied AFTER the system-driven reset block
                # below, which zeroes "Remote Area - AK" unconditionally.

                if raw_declared_value_net != 0:
                    row_result["Declared Value"] = raw_declared_value_net

                # =====================================================
                # UNDELIVERABLE RETURN + SCC Ground Undeliverable Return
                # The RTN "Ground Undeliverable Return" freight is the return
                # shipment's base rate and is already captured in CAL Base Rate
                # via the normal freight path. Do NOT charge it again in a
                # separate "Undeliverable Return" column (that double-counts).
                # =====================================================

                if scc_ground_undeliverable_seen:
                    if pure_adj_both_dims:
                        # Pure ADJ (no SHP/RTN this period) with both AG and HR:
                        # the adjustment is the C-A base computed above (with dim
                        # weight). Show it as the Base Rate correction.
                        row_result["CAL Base Rate"] = round(cal_base, 2)
                    elif matched:
                        # RTN/SHP is the主體 and the ADJ corrects it: the base is
                        # the FULL corrected freight (freight_C at the corrected
                        # billable weight), NOT a billed-minus-entered difference.
                        row_result["CAL Base Rate"] = round(freight_C, 2)
                    else:
                        # No usable dims and no matching shipment: difference of
                        # billed vs entered base rate, by weight only.
                        entered_for_scc = scc_source_entered_weight if scc_source_entered_weight > 0 else max_entered_weight
                        billed_for_scc = scc_source_billed_weight if scc_source_billed_weight > 0 else max_billed_weight
                        scc_service = scc_source_service or shipment_type
                        scc_zone = scc_source_zone or zone

                        entered_br, _ = lookup_base_rate(scc_service, scc_zone, entered_for_scc, tracking)
                        billed_br, _ = lookup_base_rate(scc_service, scc_zone, billed_for_scc, tracking)

                        row_result["CAL Base Rate"] = round(billed_br - entered_br, 2)

                # =====================================================
                # ADDRESS CORRECTION
                # =====================================================
                if address_correction_seen:
                    row_result["Address Correction"] = accessorial_lookup("Address Correction", shipment_type, zone)

                # FIX #2: Calculate Residential/Commercial Adjustment - NO max(..., 0)
                if residential_adjustment_seen or commercial_adjustment_seen:
                    commercial_base, _ = lookup_base_rate("Ground Commercial", zone, billable_weight, tracking)
                    residential_base, _ = lookup_base_rate("Ground Residential", zone, billable_weight, tracking)

                    if residential_adjustment_seen:
                        residential_adj = residential_base - commercial_base

                    if commercial_adjustment_seen:
                        commercial_adj = commercial_base - residential_base

                # =====================================================
                # RETURN TO SENDER: 2ND MILE ROAD BASE RATE
                # =====================================================
                if raw_rts_net != 0:
                    second_mile_base, _ = lookup_base_rate(shipment_type, zone, billable_weight, tracking)
                    row_result["2nd Mile Road Base"] = round(second_mile_base, 2)
                    row_result["Return To Sender"] = accessorial_lookup("Return To Sender", shipment_type, zone)

                if raw_signature_net != 0:
                    row_result["Signature"] = accessorial_lookup("Signature", shipment_type, zone)

                if raw_adult_signature_net != 0:
                    row_result["Adult Signature"] = accessorial_lookup("Adult Signature", shipment_type, zone)

                if raw_package_protection_net != 0:
                    row_result["Package Protection"] = accessorial_lookup("Package Protection", shipment_type, zone)

                # Flat, reclassification-independent charges (carry by invoice
                # presence; amount from the surcharge template FLAT_CHARGES).
                if raw_reroute_seen:
                    row_result["Reroute"] = accessorial_lookup("Reroute", shipment_type, zone)
                if raw_reschedule_seen:
                    row_result["Reschedule Delivery"] = accessorial_lookup("Reschedule Delivery", shipment_type, zone)
                if raw_ret_elabel_seen:
                    row_result["Returns Electronic Label"] = accessorial_lookup("Returns Electronic Label", shipment_type, zone)
                if raw_ret_plabel_seen:
                    row_result["Returns Print Label"] = accessorial_lookup("Returns Print Label", shipment_type, zone)

                # =====================================================
                # v2.6: Dimension Adjustment / SCC Base Rate Adjustment removed
                # =====================================================

                # =====================================================
                # SYSTEM-DRIVEN SURCHARGE ASSIGNMENT
                # =====================================================
                row_result["Additional Handling Surcharge - Weight"] = 0
                row_result["Additional Handling Surcharge - Dimension"] = 0
                row_result["Additional Handling Surcharge - Packaging"] = 0
                row_result["Large Package Surcharge"] = 0
                row_result["Over Maximum Size Surcharge"] = 0
                row_result["Remote Area - AK"] = 0
                row_result["Remote Area - HI"] = 0

                # AKR = Alaska remote area. Set AFTER the reset above, from its
                # own "Remote Area - AK" rate row (Alaska is charged materially
                # higher than the 48-state remote rate, so it must NOT reuse
                # Remote Area Commercial / Residential).
                if raw_remote_ak_seen and round(raw_remote_ak_net, 2) != 0:
                    row_result["Remote Area - AK"] = accessorial_lookup(
                        "Remote Area - AK", shipment_type, zone)

                final_apply_lps = bool(system_should_lps)
                final_apply_ovr = bool(system_should_ovr)

                # =====================================================
                # ADJ LPR -> AHS CONVERSION
                # If ADJ has negative LPR and positive AHS, credit LPR and charge AHS by ACC Template.
                # =====================================================
                lpr_to_ahs_conversion = (
                    raw_lps_net < -0.01 and
                    (raw_ahs_weight_net > 0.01 or raw_ahs_dimension_net > 0.01 or raw_ahs_packaging_net > 0.01)
                )

                if lpr_to_ahs_conversion:
                    row_result["Shipping Charge Correction Large Package Surcharge"] = round(raw_lps_net, 2)
                    final_apply_lps = False

                # =====================================================
                # AHS PRIORITY (UPS OFFICIAL ORDER)
                # =====================================================
                # Per UPS 2026 rules, when a package meets multiple Additional
                # Handling criteria, only ONE charge is assessed, in this fixed
                # order of precedence (NOT the highest fee):
                #     Weight -> Length+Girth/Dimension -> Length -> Width
                #     -> Packaging
                # The weight-based charge supersedes the size-based ones when it
                # applies. AHS is not assessed at all when a Large Package
                # Surcharge applies (handled by the final_apply_lps gate).
                final_apply_ahs_weight = False
                final_apply_ahs_dimension = False
                final_apply_ahs_packaging = False

                if lpr_to_ahs_conversion:
                    final_apply_ahs_weight = raw_ahs_weight_net > 0.01
                    final_apply_ahs_dimension = raw_ahs_dimension_net > 0.01
                    final_apply_ahs_packaging = raw_ahs_packaging_net > 0.01

                elif not final_apply_lps:
                    # AHS type selection.
                    #
                    # Principle: RESPECT THE INVOICE. When UPS explicitly billed
                    # a specific Additional Handling code on this shipment, that
                    # code is authoritative for the TYPE -- we reprice its amount
                    # from the rate table but do not second-guess weight vs.
                    # dimension vs. packaging by re-deriving it from thresholds.
                    # UPS sees the real parcel; our thresholds are only an
                    # approximation (e.g. a return whose scale weight we can't
                    # see, or packaging we can't detect from dimensions).
                    #   * AHW present            -> Weight
                    #   * AHG/AHL/AHD/AHB present-> Dimension
                    #   * AHP/AHC present        -> Packaging
                    # UPS bills at most one AH per package; if more than one code
                    # somehow appears, fall back to the official precedence
                    # (Weight > Dimension > Packaging).
                    #
                    # Only when the invoice carried NO AH code at all do we rate
                    # by threshold (a genuine reprice of an un-adjusted shipment).
                    # And a pure ADJ that carried no AH code invents nothing --
                    # its AH was already assessed on the original shipment.
                    _adj_carried_ahs = (
                        ups_charged_ahs_weight
                        or ups_charged_ahs_dimension
                        or ups_charged_ahs_packaging)

                    # Which AH type did UPS actually CHARGE (positive net) on
                    # this layer? A positive code is UPS telling us the real
                    # answer for a parcel it physically handled, and it beats our
                    # thresholds -- our scale weight can be wrong. Case in point:
                    # a customer declares 1 lb, UPS re-weighs the parcel, backs
                    # out the AHG and charges AHW instead. Our threshold sees the
                    # declared 1 lb, misses the weight trigger, and would emit a
                    # negative Dimension charge; the AHW code is correct.
                    #
                    # A NEGATIVE code is different: it is UPS removing a charge
                    # (e.g. dimensions corrected below the threshold). There the
                    # thresholds on the corrected parcel give the right answer --
                    # often "no AH at all" -- so we re-rate instead of echoing a
                    # negative surcharge.
                    _billed_ahs_weight = raw_ahs_weight_net > 0.01
                    _billed_ahs_dimension = raw_ahs_dimension_net > 0.01
                    _billed_ahs_packaging = raw_ahs_packaging_net > 0.01
                    _billed_any_ahs = (_billed_ahs_weight
                                       or _billed_ahs_dimension
                                       or _billed_ahs_packaging)

                    # Does THIS layer have its own physical dimensions to rate
                    # from? Guard against the 1x1x1 placeholder UPS puts on rows
                    # with no real measured dimensions (e.g. undeliverable
                    # returns): a genuine parcel's longest side exceeds 1 inch.
                    _have_own_dims = (length > 1.0 and cubic_inches > 1.0)

                    if _billed_any_ahs:
                        # UPS positively charged an AH code -> that type wins,
                        # in the official precedence order.
                        if _billed_ahs_weight:
                            final_apply_ahs_weight = True
                        elif _billed_ahs_dimension:
                            final_apply_ahs_dimension = True
                        elif _billed_ahs_packaging:
                            final_apply_ahs_packaging = True
                    elif _have_own_dims:
                        # No positive AH code (either none at all, or only a
                        # refund). Re-rate the corrected parcel by threshold.
                        # Packaging can never be derived from dimensions, so it
                        # still falls back to the invoice code.
                        if ahs_weight_trigger:
                            final_apply_ahs_weight = True
                        elif ahs_dimension_trigger:
                            final_apply_ahs_dimension = True
                        elif ups_charged_ahs_packaging:
                            final_apply_ahs_packaging = True
                        # else: corrected parcel triggers no AHS -> none.
                    elif ahs_weight_trigger:
                        # No usable dimensions, but AHS-Weight does not need any:
                        # it depends solely on the scale weight. This is the
                        # reclassification case -- a Residential->Commercial
                        # adjustment carries no dims and no AH code of its own,
                        # yet the parcel is still over the weight threshold and
                        # UPS keeps charging Additional Handling (only the
                        # residential surcharge is reversed). Without this branch
                        # the reclass row rates AHS as 0 and wipes out the
                        # shipment's legitimate weight-based charge.
                        final_apply_ahs_weight = True
                    elif _adj_carried_ahs:
                        # No dims to rate; trust the invoice-specified type.
                        if ups_charged_ahs_weight:
                            final_apply_ahs_weight = True
                        elif ups_charged_ahs_dimension:
                            final_apply_ahs_dimension = True
                        elif ups_charged_ahs_packaging:
                            final_apply_ahs_packaging = True
                    else:
                        # No dims and no AH code: nothing to charge.
                        pass

                # For a pure prior-period ADJ carrying both AG (C) and HR (A),
                # the size-driven surcharges are charged as C minus A.
                _pure_ca = bool(pure_adj_both_dims)

                # ...but only when UPS actually adjusted a size-driven charge on
                # this ADJ. A pure freight/fuel correction (no AH and no Large
                # Package code) must NOT produce a C-A delta: those charges were
                # already settled on the original shipment, so re-deriving C and
                # A from thresholds here yields a spurious non-zero (often
                # negative) amount. Gate the C-A netting on the invoice actually
                # carrying one of these codes -- the same rule used for the
                # threshold path above.
                #
                # NOTE: the gate must include the Large Package codes (LPS/LPR/
                # LCT/LCR), not just AH. An ADJ that upgrades a parcel to a Large
                # Package carries LPR but no AH code; gating on AH alone skipped
                # the whole branch and silently dropped the Large Package
                # Surcharge.
                _adj_carried_ahs_ca = (
                    ups_charged_ahs_weight
                    or ups_charged_ahs_dimension
                    or ups_charged_ahs_packaging
                    or ups_charged_lps)

                if _pure_ca and not _adj_carried_ahs_ca:
                    # No AH / Large Package code on this pure ADJ: leave the
                    # size-driven columns at 0.
                    pass
                elif _pure_ca:
                    # AHS is ONE mutually-exclusive charge per package. Compute
                    # C's single AHS amount (the type that applied, highest) and
                    # A's single AHS amount (A's own highest triggered type), and
                    # record the net C - A under the winning type's column. This
                    # avoids double-subtracting when A and C pick different types
                    # (e.g. both trigger weight AND dimension).
                    def c_ahs_amount():
                        if final_apply_ahs_weight:
                            return ("Additional Handling Surcharge - Weight",
                                    accessorial_lookup("AHS Weight", reclass_service, zone))
                        if final_apply_ahs_dimension:
                            return ("Additional Handling Surcharge - Dimension",
                                    accessorial_lookup("AHS Dimension", reclass_service, zone))
                        if final_apply_ahs_packaging:
                            return ("Additional Handling Surcharge - Packaging",
                                    accessorial_lookup("AHS Packaging", reclass_service, zone))
                        return (None, 0.0)

                    def a_ahs_amount():
                        # AHS and LPS are mutually exclusive: if the A package
                        # triggered LPS, it was charged LPS (not AHS), so its AHS
                        # contribution is 0 for the C-A subtraction.
                        if a_lps_trigger:
                            return 0.0
                        cands = []
                        if a_ahs_weight_trigger:
                            cands.append(accessorial_lookup("AHS Weight", orig_service, zone))
                        if a_ahs_dimension_trigger:
                            cands.append(accessorial_lookup("AHS Dimension", orig_service, zone))
                        return max(cands) if cands else 0.0

                    if not final_apply_lps:
                        c_col, c_val = c_ahs_amount()
                        a_val = a_ahs_amount()
                        net = round(c_val - a_val, 2)
                        # Put the net under C's winning column; if C had no AHS
                        # (only A did), record the credit under A's type.
                        target = c_col
                        if target is None and a_val > 0:
                            target = ("Additional Handling Surcharge - Weight"
                                      if a_ahs_weight_trigger else
                                      "Additional Handling Surcharge - Dimension")
                        if target is not None:
                            row_result[target] = net
                        # If the ORIGINAL A package was a Large Package but the
                        # corrected C package is not, refund A's LPS (C has 0
                        # LPS, so C - A = -A). C's new AHS (above) is charged
                        # separately. This is the "downgrade from LPS to AHS".
                        if a_lps_trigger:
                            row_result["Large Package Surcharge"] = round(
                                -accessorial_lookup("Large Package", orig_service, zone), 2)

                    if final_apply_lps:
                        c_amt = accessorial_lookup("Large Package", reclass_service, zone)
                        if a_lps_trigger:
                            c_amt = round(c_amt - accessorial_lookup(
                                "Large Package", orig_service, zone), 2)
                        row_result["Large Package Surcharge"] = c_amt
                        # C is now a Large Package, so it carries NO AHS (LPS and
                        # AHS are mutually exclusive). If the original A package
                        # had AHS, that AHS must be refunded (C has 0 AHS, so
                        # C - A = -A).
                        if not a_lps_trigger:
                            a_val = a_ahs_amount()
                            if a_val > 0:
                                target = ("Additional Handling Surcharge - Weight"
                                          if a_ahs_weight_trigger else
                                          "Additional Handling Surcharge - Dimension")
                                row_result[target] = round(-a_val, 2)

                else:
                    # When the invoice's AH line is a REFUND/REVERSAL (negative
                    # net, e.g. an AHW backed out because the shipment was
                    # upgraded to a Large Package, or a standalone AHG credit),
                    # the repriced amount must be negative too. Reprice the
                    # magnitude from the rate table but carry the invoice's sign.
                    def _signed(fee_col, fee_key, raw_net):
                        amt = accessorial_lookup(fee_key, shipment_type, zone)
                        if raw_net < -0.001:
                            amt = -abs(amt)
                        row_result[fee_col] = amt

                    if final_apply_ahs_weight:
                        _signed("Additional Handling Surcharge - Weight",
                                "AHS Weight", raw_ahs_weight_net)

                    if final_apply_ahs_dimension:
                        _signed("Additional Handling Surcharge - Dimension",
                                "AHS Dimension", raw_ahs_dimension_net)

                    if final_apply_ahs_packaging:
                        _signed("Additional Handling Surcharge - Packaging",
                                "AHS Packaging", raw_ahs_packaging_net)
                        if not final_apply_lps:
                            pkg_name = "Additional Handling Surcharge - Packaging"
                            existing = str(row_result.get("AHS", "")).strip()
                            if existing and pkg_name not in existing:
                                row_result["AHS"] = existing + "; " + pkg_name
                            elif not existing:
                                row_result["AHS"] = pkg_name
                            row_result["LPS"] = ""

                    if final_apply_lps:
                        row_result["Large Package Surcharge"] = \
                            accessorial_lookup("Large Package", shipment_type, zone)

                if final_apply_ovr:
                    row_result["Over Maximum Size Surcharge"] = round(raw_ovr_net, 2)

                # FIX #3: ASSIGN THREE VALUES - NO > 0 conditional check
                row_result["Residential Adjustment"] = round(residential_adj, 2)
                row_result["Commercial Adjustment"] = round(commercial_adj, 2)
                row_result["Direct Delivery Only"] = round(ddo_charge, 2)

                # ===== CUSTOM ACC SURCHARGES =====
                # dynamic_surcharge_seen_codes used to be collected and then
                # dropped, so anything added on the Custom Surcharges tab was
                # silently worth $0. Price each seen code from the surcharge
                # rate table now that shipment_type and zone are final.
                # Two distinct AS codes may map to the same surcharge name
                # (e.g. commercial / residential variants), so accumulate.
                for _dyn_as, _dyn_name, _dyn_fuel_ok, _dyn_mode in self.dynamic_acc_columns():
                    if _dyn_as not in dynamic_surcharge_seen_codes:
                        continue

                    if _dyn_mode == "REPRICE":
                        # Value it from the negotiated rate table.
                        _dyn_amt = accessorial_lookup(_dyn_name, shipment_type, zone)
                    else:
                        # COPY INVOICE and REPORT ONLY both carry the billed
                        # figure -- there is no contract rate to apply. They
                        # differ only in whether CAL Total counts them.
                        _dyn_amt = self.to_amount(
                            dynamic_surcharge_seen_amounts.get(_dyn_as, 0))

                    row_result[_dyn_name] = round(
                        self.to_amount(row_result.get(_dyn_name, 0)) + _dyn_amt, 2)

                # ===== FINAL CHARGE BASE FOR FUEL / TOTAL =====
                # Fuel and Total are computed from the SAME charge columns that
                # appear in the exported detail layout, so the displayed
                # columns always reconcile with Total. (Internal raw-adjustment
                # columns like Commercial/Residential Adjustment are not shown
                # and must not silently drag Total negative.)
                cal_charge_cols = [
                    "CAL Base Rate",
                    "Residential", "DAS Commercial", "DAS Residential",
                    "DAS Extended Commercial", "DAS Extended Residential",
                    "Remote Area Commercial", "Remote Area Residential",
                    "Remote Area - AK", "Remote Area - HI",
                    "Additional Handling Surcharge - Weight",
                    "Additional Handling Surcharge - Dimension",
                    "Additional Handling Surcharge - Packaging",
                    "Large Package Surcharge", "Over Maximum Size Surcharge",
                    "Undeliverable Return", "Address Correction",
                    "Declared Value", "Return To Sender", "2nd Mile Road Base",
                    "Signature", "Adult Signature", "Package Protection", "Direct Delivery Only",
                ] + dyn_acc_billable

                # =====================================================
                # UNSUPPORTED SERVICE/ZONE: DO NOT RATE
                # =====================================================
                # This lane has no rate table (e.g. Standard to Canada on zone
                # 53). Every figure computed above used the Ground Commercial /
                # zone 8 fallback and is therefore meaningless. Clear the
                # calculated charges so the row reports nothing rather than
                # something wrong -- the Notes column carries the warning and
                # the UPS invoice total is still shown for reference.
                if _unsupported_lane:
                    for _c in cal_charge_cols:
                        row_result[_c] = 0
                    row_result["CAL Fuel"] = 0

                # A custom surcharge marked Fuel Eligible = NO still counts
                # toward CAL Total but must not inflate the fuel base.
                fuel_base = sum(
                    self.to_amount(row_result.get(col, 0))
                    for col in cal_charge_cols
                    if col not in dyn_acc_nonfuel
                )

                if not _unsupported_lane:
                    row_result["CAL Fuel"] = round(fuel_base * fuel_pct, 2)

                # =====================================================
                # FINAL CALCULATION
                # =====================================================
                cal_total = round(
                    sum(self.to_amount(row_result.get(col, 0)) for col in cal_charge_cols) + row_result["CAL Fuel"],
                    2
                )

                row_result["CAL Total"] = cal_total
                row_result["UPS Invoice Total"] = round(ups_total_for_tracking, 2)
                row_result["_fuel_pct"] = fuel_pct

                # ===== APPEND row_result TO output_rows =====
                output_rows.append(row_result)

            # =====================================================
            # PREVIEW (table UI removed; status only)
            # =====================================================
            if not output_rows:
                raise Exception(
                    "Repricing produced no rows.\n\n"
                    "This usually means the billing file format is wrong, or "
                    "no base rate table was imported. Please import a Base "
                    "Rate template and check the billing file, then try again.")
            try:
                self.status_var.set(
                    f"Repriced {len(output_rows)} layer rows. Exporting...")
            except Exception:
                pass

            # =====================================================
            # EXPORT
            # =====================================================
            # Default the report name to the invoice it came from, so a
            # folder of reports says which invoice each one belongs to
            # instead of a row of "Save Repricing Result" guesses.
            # normalise separators first: a Windows path handed to a POSIX
            # basename() comes back whole, and the report gets named after
            # the entire path.
            _src_raw = self.billing_path.get().strip().replace("\\", "/")
            _src_name = os.path.splitext(os.path.basename(_src_raw))[0]
            _suggest = (f"{_src_name}_Repriced.xlsx" if _src_name
                        else "Repriced.xlsx")
            _initdir = os.path.dirname(_src_raw) or os.getcwd()

            export_path = filedialog.asksaveasfilename(
                title="Save Repricing Result",
                initialfile=_suggest,
                initialdir=_initdir,
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")]
            )

            if export_path:

                detail_df = pd.DataFrame(output_rows)

                # Charge columns = per-code buckets. Each named column is one
                # charge type; combine overrides per column (per code).
                charge_cols = [
                    "CAL Base Rate", "Residential",
                    "DAS Commercial", "DAS Residential",
                    "DAS Extended Commercial", "DAS Extended Residential",
                    "Remote Area Commercial", "Remote Area Residential",
                    "Remote Area - AK", "Remote Area - HI",
                    "Additional Handling Surcharge - Weight",
                    "Additional Handling Surcharge - Dimension",
                    "Additional Handling Surcharge - Packaging",
                    "Large Package Surcharge", "Over Maximum Size Surcharge",
                    "Direct Delivery Only",
                    "Undeliverable Return", "Address Correction",
                    "Declared Value", "Return To Sender", "2nd Mile Road Base",
                    "Signature", "Adult Signature", "Package Protection",
                    "Reroute", "Reschedule Delivery",
                    "Returns Electronic Label", "Returns Print Label",
                ] + dyn_acc_cols
                id_cols = ["Tracking", "Layer", "Zone", "Shipment Type",
                           "Entered Weight", "Billed Weight", "Dim Weight",
                           "Billable Weight", "Length", "Width", "Height"]

                # rename map: drop "CAL", use friendly names
                rename_map = {
                    "CAL Base Rate": "Base Rate",
                    "CAL Fuel": "Fuel Surcharge",
                    "CAL Total": "Total",
                    "Residential": "Residential Surcharge",
                    "2nd Mile Road Base": "2nd Mile Road Base Rate",
                }

                if not detail_df.empty:
                    for c in (id_cols + charge_cols +
                              ["CAL Fuel", "CAL Total", "_fuel_pct"]):
                        if c not in detail_df.columns:
                            detail_df[c] = 0

                    # ---------- v2: per tracking = ONE final correct row ----------
                    # Size/weight-driven charges are recomputed by the ADJ row
                    # (reclassified service + size/weight triggers). Non-size
                    # charges (Signature, Address Correction, Declared Value,
                    # etc.) carry over from the original SHP/RTN.
                    # Output:
                    #   - has SHP/RTN  -> "Final" sheet (corrected if ADJ exists)
                    #   - ADJ-only     -> "Prior Adj" sheet (修上期)
                    size_driven_cols = {
                        "CAL Base Rate", "Residential",
                        "DAS Commercial", "DAS Residential",
                        "DAS Extended Commercial", "DAS Extended Residential",
                        "Remote Area Commercial", "Remote Area Residential",
                        "Remote Area - AK", "Remote Area - HI",
                        "Additional Handling Surcharge - Weight",
                        "Additional Handling Surcharge - Dimension",
                        "Additional Handling Surcharge - Packaging",
                        "Large Package Surcharge", "Over Maximum Size Surcharge",
                        "2nd Mile Road Base",
                    }
                    # Truly dimension/weight-triggered charges. In scenario B
                    # (no reclassification) these are ALWAYS recomputed from the
                    # ADJ row by thresholds. The remaining size_driven items
                    # (residential / DAS / Remote) are service-driven and only
                    # switch on reclassification (scenario A); otherwise they
                    # carry from SHP.
                    dimension_driven_cols = {
                        "CAL Base Rate",
                        "Additional Handling Surcharge - Weight",
                        "Additional Handling Surcharge - Dimension",
                        "Additional Handling Surcharge - Packaging",
                        "Large Package Surcharge", "Over Maximum Size Surcharge",
                        "2nd Mile Road Base",
                    }

                    final_rows = []
                    for trk, g in detail_df.groupby("Tracking"):
                        shp = g[g["Layer"] == "SHP"]
                        rtn = g[g["Layer"] == "RTN"]
                        adj = g[g["Layer"] == "ADJ"]

                        def layer_charge(layer_df, col):
                            if layer_df.empty:
                                return 0.0
                            return float(pd.to_numeric(
                                layer_df[col], errors="coerce").fillna(0).sum())

                        has_origin = (not shp.empty) or (not rtn.empty)
                        has_adj = not adj.empty

                        # Scenario A vs B: A = ADJ carries a COM/RES reclass.
                        reclassified = False
                        if has_adj and "_reclassified" in adj.columns:
                            reclassified = bool(pd.to_numeric(
                                adj["_reclassified"], errors="coerce")
                                .fillna(0).max() >= 1)

                        def adj_listed(col):
                            # True if the ADJ layer actually carries this charge
                            # (non-zero), i.e. UPS adjusted this specific code.
                            return abs(layer_charge(adj, col)) > 0.0001

                        # Did the ADJ layer actually re-rate freight/size? Check
                        # whether it carries a freight (FRT) line at all. Do NOT
                        # test its Length/Weight columns: when an ADJ has no dims
                        # of its own the tool fills them in from the matched
                        # SHP/RTN, so those columns look populated even for an
                        # accessorial-only correction. An ADJ made purely of
                        # accessorial lines (e.g. a standalone Reschedule
                        # Delivery fee) has no freight line -- its recomputed
                        # base/AHS/LPS are 0 and must not overwrite the
                        # shipment's real values.
                        adj_touches_size = True
                        if has_adj:
                            try:
                                adj_touches_size = float(pd.to_numeric(
                                    adj["_has_freight"], errors="coerce")
                                    .fillna(0).max()) > 0
                            except (KeyError, IndexError):
                                adj_touches_size = True

                        # Does the ADJ layer carry its OWN recomputed dimensions?
                        # An SCC that re-rates size (base/AHS/LPS) has its own
                        # Length/Width/Height. A pure REFUND adjustment (UPS backs
                        # out a previously-billed AHG/AHL with a negative net and
                        # no dimensions of its own) does not. This distinction
                        # decides how dimension-driven charges merge below.
                        adj_has_own_dims = False
                        if has_adj:
                            try:
                                _al = float(pd.to_numeric(
                                    adj["Length"], errors="coerce")
                                    .fillna(0).max())
                                _aw = float(pd.to_numeric(
                                    adj["Width"], errors="coerce")
                                    .fillna(0).max())
                                _ah = float(pd.to_numeric(
                                    adj["Height"], errors="coerce")
                                    .fillna(0).max())
                                adj_has_own_dims = (_al > 0 and _aw > 0
                                                    and _ah > 0)
                            except (KeyError, IndexError):
                                adj_has_own_dims = False

                        # Descriptive info source: corrected uses ADJ (updated)
                        # when present, else the original.
                        info_src0 = (adj if has_adj else
                                     (shp if not shp.empty else rtn))

                        def iv(col, default=""):
                            try:
                                return info_src0[col].iloc[0]
                            except (KeyError, IndexError):
                                return default

                        # The shipment (SHP/RTN) is the主體 for physical fields
                        # like entered/billed weight. Use the origin row for
                        # those when it exists; only a pure ADJ uses the ADJ row.
                        origin_src = (shp if not shp.empty else
                                      (rtn if not rtn.empty else adj))

                        def ov(col, default=""):
                            try:
                                v = origin_src[col].iloc[0]
                                return v
                            except (KeyError, IndexError):
                                return default

                        fuel_pct = 0.0
                        for src in (adj, shp, rtn):
                            if not src.empty:
                                fuel_pct = float(pd.to_numeric(
                                    src["_fuel_pct"], errors="coerce")
                                    .fillna(0).iloc[0])
                                break

                        def fmt_date(raw):
                            # Return a real datetime so Excel stores it as a
                            # numeric date value (not text). The display format
                            # (MM/DD/YYYY) is applied to the column later.
                            s = str(raw).strip()
                            if not s:
                                return s
                            try:
                                dt = pd.to_datetime(s, errors="coerce")
                                if pd.isna(dt):
                                    return s
                                return dt.to_pydatetime()
                            except Exception:
                                return s

                        row = {"Date": fmt_date(iv("Date")), "Tracking": trk,
                               "Invoice Number": (iv("Invoice Number")
                                                  or ov("Invoice Number", "")),
                               "Invoice Date": fmt_date(
                                   iv("Invoice Date")
                                   or ov("Invoice Date", ""))}
                        # Entered/Billed weight come from the shipment (主體);
                        # the rest of the descriptive fields track the corrected
                        # (ADJ) row when present.
                        for f in ("Zone", "Shipment Type", "Dim Weight",
                                  "Billable Weight", "Length", "Width",
                                  "Height"):
                            row[f] = iv(f)
                        # Zone guard: an ADJ layer made up only of accessorial
                        # rows (e.g. a standalone Reschedule Delivery fee) carries
                        # zone 000, which normalizes to "0" -- not a real UPS
                        # zone. Preferring the ADJ row then wipes out the real
                        # zone and the shipment rates at zone 0 (no base rate).
                        # Fall back to the shipment's own zone in that case.
                        _z = str(row.get("Zone", "")).strip()
                        if _z in ("", "0", "0.0"):
                            _z_origin = str(ov("Zone", "")).strip()
                            if _z_origin not in ("", "0", "0.0"):
                                row["Zone"] = _z_origin
                        for f in ("Entered Weight", "Billed Weight"):
                            row[f] = ov(f)

                        base_sum = 0.0
                        for col in charge_cols:
                            name = rename_map.get(col, col)
                            shp_val = layer_charge(shp, col)
                            rtn_val = layer_charge(rtn, col)
                            adj_val = layer_charge(adj, col)
                            origin_val = shp_val + rtn_val

                            if not has_adj:
                                # No adjustment: original as-is.
                                val = origin_val

                            elif not has_origin:
                                # Pure ADJ (修上期): ADJ row already holds the
                                # corrected value (C, or C-A for both-dims).
                                val = adj_val

                            elif reclassified and col in size_driven_cols:
                                # Scenario A: whole service switched; ALL
                                # size-driven (incl. residential/DAS/Remote)
                                # come from the recomputed ADJ row.
                                val = adj_val

                            elif col in dimension_driven_cols:
                                # Scenario B: dimension-triggered charges
                                # (base rate / AHS / LPS / Over-max) come from
                                # the recomputed ADJ row, which re-rates using
                                # ADJ's dims (or SHP dims if ADJ carried none).
                                #
                                # Exception: an ADJ that made NO freight or size
                                # adjustment at all -- no dimensions and no
                                # weight of its own, e.g. a standalone accessorial
                                # correction like a Reschedule Delivery fee -- has
                                # a recomputed value of 0 for these columns.
                                # Taking it would wipe out the shipment's real
                                # base rate. Keep the origin in that case.
                                if adj_touches_size:
                                    val = adj_val
                                else:
                                    val = origin_val

                            else:
                                # Service-driven (residential/DAS/Remote) and
                                # flat charges (Signature, Reroute, RTS, etc.):
                                # ADJ adjusts only what it actually lists; the
                                # rest keeps the SHP/RTN original.
                                if adj_listed(col):
                                    val = origin_val + adj_val
                                else:
                                    val = origin_val

                            val = round(val, 2)
                            base_sum += val
                            row[name] = val

                        row["Fuel Surcharge"] = round(base_sum * fuel_pct, 2)
                        row["Total"] = round(
                            base_sum + row["Fuel Surcharge"], 2)

                        final_rows.append(row)

                    detail_full = detail_df.copy().rename(columns=rename_map)
                    final_df = pd.DataFrame(final_rows)
                else:
                    detail_full = pd.DataFrame()
                    final_df = pd.DataFrame()

                # Fixed detail layout (matches the Shipment Detail reference)
                detail_layout = [
                    # No Invoice Number / Invoice Date: this workbook goes to
                    # the client, and they are UPS's own references. The
                    # statement takes them straight from the invoice instead.
                    "Date", "Tracking", "Zone", "Shipment Type",
                    "Entered Weight", "Billed Weight", "Dim Weight",
                    "Billable Weight", "Length", "Width", "Height",
                    "Base Rate", "Residential Surcharge",
                    "DAS Commercial", "DAS Residential",
                    "DAS Extended Commercial", "DAS Extended Residential",
                    "Remote Area Commercial", "Remote Area Residential",
                    "Remote Area - AK", "Remote Area - HI",
                    "Additional Handling Surcharge - Weight",
                    "Additional Handling Surcharge - Dimension",
                    "Additional Handling Surcharge - Packaging",
                    "Large Package Surcharge", "Over Maximum Size Surcharge",
                    "Undeliverable Return", "Address Correction",
                    "Declared Value", "Return To Sender",
                    "2nd Mile Road Base Rate",
                    "Signature", "Adult Signature", "Package Protection", "Direct Delivery Only",
                    "Reroute", "Reschedule Delivery",
                    "Returns Electronic Label", "Returns Print Label",
                ] + dyn_acc_cols + [
                    "Fuel Surcharge", "Total",
                ]
                # NOTE: no UPS cost, profit or margin here. This report is
                # what the client receives. Cost lives only in the internal
                # statement, which the Profit Report button builds by joining
                # this report back to the UPS invoice.
                detail_money = [
                    "Base Rate", "Residential Surcharge",
                    "DAS Commercial", "DAS Residential",
                    "DAS Extended Commercial", "DAS Extended Residential",
                    "Remote Area Commercial", "Remote Area Residential",
                    "Remote Area - AK", "Remote Area - HI",
                    "Additional Handling Surcharge - Weight",
                    "Additional Handling Surcharge - Dimension",
                    "Additional Handling Surcharge - Packaging",
                    "Large Package Surcharge", "Over Maximum Size Surcharge",
                    "Undeliverable Return", "Address Correction",
                    "Declared Value", "Return To Sender",
                    "2nd Mile Road Base Rate",
                    "Signature", "Adult Signature", "Package Protection", "Direct Delivery Only",
                    "Reroute", "Reschedule Delivery",
                    "Returns Electronic Label", "Returns Print Label",
                ] + dyn_acc_cols + [
                    "Fuel Surcharge", "Total",
                ]
                # NOTE: no UPS cost, profit or margin here. This report is
                # what the client receives. Cost lives only in the internal
                # statement, which the Profit Report button builds by joining
                # this report back to the UPS invoice.

                from openpyxl.utils import get_column_letter
                money_fmt = '$#,##0.00'
                date_fmt = 'mm/dd/yyyy'

                def fmt_sheet(ws, headers, money_set):
                    for ci, h in enumerate(headers, start=1):
                        letter = get_column_letter(ci)
                        if h in money_set:
                            for cell in ws[letter][1:]:
                                cell.number_format = money_fmt
                        elif h == "Date":
                            for cell in ws[letter][1:]:
                                cell.number_format = date_fmt

                with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
                    final_layout = list(detail_layout)
                    final_money = set(detail_money)

                    # Sheet 1 — 總表 (Shipment Detail): every shipped tracking,
                    # one row, ONE sheet with every tracking.
                    final_layout_one = list(final_layout)
                    if not final_df.empty:
                        for col in final_layout_one:
                            if col not in final_df.columns:
                                final_df[col] = (0 if col in final_money
                                                 else "")
                        fin = final_df[final_layout_one]
                    else:
                        fin = pd.DataFrame(columns=final_layout_one)
                    fin.to_excel(writer, sheet_name="Shipment Detail",
                                 index=False)
                    if not fin.empty:
                        fmt_sheet(writer.sheets["Shipment Detail"],
                                  list(fin.columns), final_money)

                    # Sheet 2 -- Unknown Codes. These rows were already being
                    # collected and then thrown away, so an invoice could
                    # contain charge codes the tool cannot price and nothing
                    # would ever say so. Silent under-pricing is the worst
                    # possible failure for an audit tool, so it now gets its
                    # own sheet plus a warning dialog.
                    # NOTE: Account Charges and Unknown Codes used to be
                    # written here. They carry UPS amounts, and this workbook
                    # goes to the client -- so they moved to the internal
                    # statement that the Profit Report button builds.


                _report_notes = []
                if unknown_rows:
                    _report_notes.append(
                        f"{len(unknown_rows)} row(s) with charge codes this "
                        f"tool cannot price -- NOT in CAL Total, see "
                        f"\"Unknown Codes\" sheet")

                if account_rows:
                    _acct_total = sum(self.to_amount(r.get("Amount", 0))
                                      for r in account_rows)
                    _report_notes.append(
                        f"{len(account_rows)} account-level charge(s) "
                        f"${_acct_total:,.2f} -- see \"Account Charges\" sheet")

                # Remember it so Generate Profit Report can use this run
                # directly instead of making you find the file again.
                self.last_report_path = export_path

                # One line, no modal. Anything that needs attention is named
                # here and detailed on its own sheet.
                self.set_status("Report generated."
                                + ("   |   " + "   |   ".join(_report_notes) if _report_notes
                                   else ""))

            if unknown_rows:
                _codes = {str(r.get("AS Charge Code", "")).strip()
                          for r in unknown_rows}
                _codes.discard("")
                self.status_var.set(
                    f"Report generated -- {len(_codes)} UNKNOWN code(s), "
                    f"{len(unknown_rows)} row(s) not priced")
            else:
                self.status_var.set("Report Generated Successfully")

        except Exception as e:

            try:
                self.status_var.set("Error — see message.")
            except Exception:
                pass
            messagebox.showerror("Cannot Generate Report", str(e))


if __name__ == "__main__":

    try:
        root = ctk.CTk()
        app = UPSRepricingTool(root)

        # Settings should survive closing the window. Load whatever was saved
        # last time, and save again on the way out, so nothing has to be
        # re-entered or re-imported by hand.
        try:
            app.load_config(silent=True)
        except Exception as _e:
            print("Auto-load skipped:", _e)

        def _on_close():
            try:
                app.save_config(silent=True)
            except Exception as _e:
                print("Auto-save skipped:", _e)
            root.destroy()

        root.protocol("WM_DELETE_WINDOW", _on_close)
        root.mainloop()

    except Exception as e:

        error_text = traceback.format_exc()

        with open("UPS_TOOL_CRASH_LOG.txt", "w", encoding="utf-8") as f:
            f.write(error_text)

        print(error_text)

        input("\nProgram crashed. Press Enter to exit...")

